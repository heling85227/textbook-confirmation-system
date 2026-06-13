import streamlit as st
import pandas as pd
import io
import math
import base64
from datetime import datetime, date
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook

from database import query_df, execute_sql, PRICE_CALC, PRICE_JOIN
from components import show_header, excel_export, apply_excel_borders, styled_dataframe
from utils import (
    get_filtered_list,
    get_filtered_class_names,
    get_filtered_colleges,
    get_filtered_majors,
    get_filtered_grades,
    safe_int,
    safe_float,
    safe_str,
    safe_field,
    normalize_grade,
    get_class_student_counts,
)


def confirmation_page():
    show_header("学生领书确认表", "按班级确认学生是否需要教材，免领标记后该生费用统计将排除")

    # 页面跟踪：从其他页面切换进来时，重置班级选择
    prev_page = st.session_state.get("current_page", "")
    if prev_page != "confirmation":
        st.session_state.pop("c_class", None)
        st.session_state.pop("c_major", None)
    st.session_state["current_page"] = "confirmation"

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("请先添加学期")
        return

    class_names = get_filtered_class_names()
    colleges = get_filtered_colleges()
    majors = get_filtered_majors()
    if not class_names:
        st.warning("请先在学生管理中导入学生数据")
        return

    tab1, tab2 = st.tabs(["领书确认", "确认汇总"])

    # Tab 1: 领书确认（查看领取情况，支持编辑备注）
    with tab1:
        st.markdown("#### 查看教材领取情况")
        st.caption("勾选「补领」可补发教材，勾选「退书」可退回。有退伍复学等特殊情况，请在「教材发放表」中补录。")

        with st.expander("🔍 筛选条件", expanded=True):
            col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1.5])
            with col1:
                def _reset_c_class():
                    st.session_state.pop("c_class", None)

                def _reset_c_major_and_class():
                    st.session_state.pop("c_major", None)
                    st.session_state.pop("c_class", None)

                c_semester = st.selectbox(
                    "学期 *",
                    [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                    format_func=lambda x: x[1],
                    key="c_sem",
                    on_change=_reset_c_class,
                )
            with col2:
                c_college = st.selectbox(
                    "学院", ["全部"] + colleges, key="c_college", on_change=_reset_c_major_and_class
                )
            with col3:
                # 专业：基于所选学院级联过滤
                if c_college != "全部":
                    c_major_opts = ["全部"] + get_filtered_list(
                        "students", "major", "college = %s", (c_college,)
                    )
                else:
                    c_major_opts = ["全部"] + majors
                c_major = st.selectbox("专业", c_major_opts, key="c_major", on_change=_reset_c_class)
            with col4:
                # 班级：基于所选学院+专业级联过滤
                c_class_where = "1=1"
                c_class_params = []
                if c_college != "全部":
                    c_class_where += " AND college = %s"
                    c_class_params.append(c_college)
                if c_major != "全部":
                    c_class_where += " AND major = %s"
                    c_class_params.append(c_major)
                c_class_opts = (
                    get_filtered_list("students", "class_name", c_class_where, tuple(c_class_params))
                    if c_class_params
                    else class_names
                )
                # 处理从概览跳转到指定班级
                if st.session_state.get("_goto_class"):
                    st.session_state["c_class"] = st.session_state.pop("_goto_class")
                    st.rerun()
                c_class_opts = ["请选择班级"] + c_class_opts
                c_class = st.selectbox("班级 *", c_class_opts, key="c_class")
                st.caption("请先选择班级，再查看领取情况")

        # 班级领取概览：选学期后自动展示（折叠）
        if c_semester and (not c_class or c_class == "请选择班级"):
            sem_id_ov = c_semester[0]
            # 查所有已下发教材的班级及其学生/教材数
            ov_classes = query_df(
                """
                SELECT o.class_name, COUNT(DISTINCT s.id) as stu_count,
                       COUNT(DISTINCT o.textbook_id) as tb_count
                FROM textbook_orders o
                LEFT JOIN students s ON s.class_name = o.class_name
                WHERE o.semester_id = %s
                GROUP BY o.class_name
                ORDER BY o.class_name
            """,
                (sem_id_ov,),
            )

            if not ov_classes.empty:
                # 查各班级的发放记录数
                ov_dist = query_df(
                    """
                    SELECT s.class_name, COUNT(*) as dist_count
                    FROM distributions d
                    JOIN textbooks t ON d.textbook_id = t.id
                    JOIN students s ON d.student_id = s.id
                    WHERE t.semester_id = %s
                    GROUP BY s.class_name
                """,
                    (sem_id_ov,),
                )
                dist_map = {}
                for _, dr in ov_dist.iterrows():
                    dist_map[dr["class_name"]] = int(dr["dist_count"])

                # 查各班级签名数
                ov_sig = query_df(
                    """
                    SELECT s.class_name, COUNT(*) as sig_count
                    FROM signatures sg
                    JOIN students s ON sg.student_id = s.id
                    WHERE sg.semester_id = %s
                    GROUP BY s.class_name
                """,
                    (sem_id_ov,),
                )
                sig_map_ov = {}
                for _, sr in ov_sig.iterrows():
                    sig_map_ov[sr["class_name"]] = int(sr["sig_count"])

                # 批量查询所有班级的学院/专业映射（一次查询替代逐班N+1查询）
                class_meta_rows = query_df(
                    "SELECT DISTINCT class_name, college, major FROM students WHERE class_name IS NOT NULL"
                )
                class_meta = {}
                for _, cm in class_meta_rows.iterrows():
                    class_meta[cm["class_name"]] = {
                        "college": cm.get("college") or "",
                        "major": cm.get("major") or ""
                    }

                # 按学院/专业筛选 + 逐班计算完成率
                ov_rows = []
                for _, oc in ov_classes.iterrows():
                    cn = oc["class_name"]
                    meta = class_meta.get(cn, {})
                    if c_college != "全部" and meta.get("college", "") != c_college:
                        continue
                    if c_major != "全部" and meta.get("major", "") != c_major:
                        continue
                    stu = int(oc["stu_count"])
                    tb = int(oc["tb_count"])
                    total_pairs = stu * tb
                    actual = dist_map.get(cn, 0)
                    pct = round(actual / total_pairs * 100, 1) if total_pairs > 0 else 0
                    sig_cnt = sig_map_ov.get(cn, 0)
                    ov_rows.append(
                        {
                            "班级": cn,
                            "学生": stu,
                            "教材种数": tb,
                            "应发条数": total_pairs,
                            "已发条数": actual,
                            "完成率": pct,
                            "已签名": sig_cnt,
                        }
                    )

                if ov_rows:
                    st.divider()
                    st.markdown("##### 各班级领取情况概览")
                    st.caption("已发条数 = 学生 × 教材的发放记录数，100% 表示该班全部教材均已发放完毕")

                    # ── 班级领取统计卡片 ──
                    total_classes = len(ov_rows)
                    completed = sum(1 for r in ov_rows if r["完成率"] >= 100)
                    partial = sum(1 for r in ov_rows if 0 < r["完成率"] < 100)
                    not_started = sum(1 for r in ov_rows if r["完成率"] == 0)

                    sc1, sc2, sc3, sc4 = st.columns(4)
                    sc1.metric("总班级数", total_classes)
                    sc2.metric("✅ 已全部完成", completed)
                    sc3.metric("⏳ 部分完成", partial)
                    sc4.metric("❌ 未发放", not_started)
                    st.divider()

                    for r in sorted(ov_rows, key=lambda x: x["完成率"]):
                        pct = r["完成率"]
                        if pct >= 100:
                            badge, icon = "", " 已全部完成"
                        elif pct > 0:
                            badge, icon = "", " 部分完成"
                        else:
                            badge, icon = "", " 未发放"
                        with st.expander(
                            f"{badge} {r['班级']} — {icon}（{r['已发条数']}/{r['应发条数']}，{pct}%）✍️{r['已签名']}/{r['学生']}签",
                            expanded=False,
                        ):
                            c1, c2, c3, c4, c5, c6 = st.columns(6)
                            c1.metric("学生数", r["学生"])
                            c2.metric("教材种数", r["教材种数"])
                            c3.metric("应发条数", r["应发条数"])
                            c4.metric("已发条数", r["已发条数"])
                            c5.metric("完成率", f"{pct}%")
                            c6.metric("✍️ 已签名", f"{r['已签名']}/{r['学生']}")
                            # 快速跳转：点击按钮自动选中该班级
                            if st.button(f"查看 {r['班级']} 详情", key=f"goto_{r['班级']}"):
                                st.session_state["_goto_class"] = r["班级"]
                                st.rerun()
                else:
                    if c_college != "全部" or c_major != "全部":
                        st.info("当前筛选条件下暂无已下发教材的班级")
            else:
                st.info("当前学期暂无已下发教材的班级，请先在「征订总表」中执行「一键下发」")

        if c_semester and c_class and c_class != "请选择班级":
            sem_id = c_semester[0]

            # ── 返回概览按钮 ──
            if st.button("← 返回班级概览", key="back_to_overview"):
                st.session_state.pop("c_class", None)
                st.rerun()

            students_df = query_df(
                "SELECT id, id_card, student_id, name, grade, college, major FROM students WHERE class_name = %s ORDER BY name",
                (c_class,),
            )
            
            # 查询该班学生在此学期的签名状态
            sig_df = query_df(
                """SELECT student_id, signature_data, signed_at 
                   FROM signatures 
                   WHERE semester_id = %s AND student_id IN (
                       SELECT id FROM students WHERE class_name = %s
                   )""",
                (sem_id, c_class),
            )
            sig_map = {}  # student_id -> {"signed_at": datetime, "data": base64}
            for _, sr in sig_df.iterrows():
                sig_map[int(sr["student_id"])] = {
                    "signed_at": str(sr["signed_at"])[:10] if sr["signed_at"] else "",
                    "data": sr["signature_data"],
                }
            textbooks_df = query_df(
                """SELECT m.name, m.publisher, m.price, t.id
                   FROM textbook_orders o
                   JOIN textbooks_master m ON o.textbook_id = m.id
                   LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                   WHERE o.semester_id = %s AND o.class_name = %s
                   ORDER BY m.name""",
                (sem_id, c_class),
            )

            if students_df.empty:
                st.info(f"班级「{c_class}」暂无学生")
                return
            if textbooks_df.empty:
                st.warning("该班级在当前学期暂无教材")
                return

            # 对没有 textbooks.id 的教材自动创建 textbooks 记录
            for _, tr in textbooks_df.iterrows():
                if tr["id"] is None or pd.isna(tr["id"]):
                    execute_sql(
                        "INSERT INTO textbooks (semester_id, grade, college, major, class_name, name, publisher, price, quantity, remark) VALUES (%s, '', '', '', %s, %s, %s, %s, 0, '[确认表·自动创建]')",
                        (
                            sem_id,
                            c_class,
                            tr["name"],
                            tr.get("publisher") or "",
                            tr["price"],
                        ),
                    )
            # 重新查询以获取新的 textbooks.id
            textbooks_df = query_df(
                """SELECT m.name, m.publisher, m.price, t.id
                   FROM textbook_orders o
                   JOIN textbooks_master m ON o.textbook_id = m.id
                   LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                   WHERE o.semester_id = %s AND o.class_name = %s
                   ORDER BY m.name""",
                (sem_id, c_class),
            )

            # 选教材
            st.markdown("##### 选择教材查看领取情况")
            tb_options = [(r["id"], r["name"]) for _, r in textbooks_df.iterrows()]
            sel_tb = st.selectbox(
                "教材",
                tb_options,
                format_func=lambda x: x[1],
                key="c_tb",
                label_visibility="collapsed",
            )

            # 该教材的发放情况
            dist_stu = query_df(
                "SELECT DISTINCT student_id FROM distributions WHERE textbook_id = %s",
                (sel_tb[0],),
            )
            dist_ids = set(int(r["student_id"]) for _, r in dist_stu.iterrows())

            # 统计
            total = len(students_df)
            got = len(dist_ids)
            not_got = total - got
            signed_count = len(sig_map)

            m1, m2, m3, m4 = st.columns(4)
            with m1:
                st.metric("班级人数", total)
            with m2:
                st.metric("已领", got)
            with m3:
                st.metric("未领", not_got)
            with m4:
                st.metric("✍️ 已签名", signed_count)

            # 学生列表（带补领+退书操作）—— 用 form 确保 checkbox 状态正确捕获
            st.markdown("##### 学生领取情况")
            st.caption("未领可「补领」，已领可「退书」，勾选后点击下方对应按钮")

            rows = []
            for i, (_, stu) in enumerate(students_df.iterrows(), 1):
                sid = int(stu["id"])
                has_book = sid in dist_ids
                sig_info = sig_map.get(sid)
                sig_status = f"✅ {sig_info['signed_at']}" if sig_info else "❌ 未签"
                rows.append(
                    {
                        "编号": i,
                        "_sid": sid,
                        "学号": stu["student_id"],
                        "姓名": stu["name"],
                        "领取状态": "已领" if has_book else "未领",
                        "学生签名": sig_status,
                        "补领": False,
                        "退书": False,
                    }
                )

            status_df = pd.DataFrame(rows)

            with st.form(f"confirm_book_form_{c_class}_{sel_tb[0]}"):
                edited_df = st.data_editor(
                    status_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "编号": st.column_config.NumberColumn("编号", disabled=True, width="small", alignment="center"),
                        "_sid": None,
                        "学号": st.column_config.TextColumn("学号", disabled=True, alignment="center"),
                        "姓名": st.column_config.TextColumn("姓名", disabled=True, alignment="center"),
                        "领取状态": st.column_config.TextColumn("领取状态", disabled=True, alignment="center"),
                        "学生签名": st.column_config.TextColumn("学生签名", disabled=True, alignment="center", help="学生在个人页面的签名确认状态"),
                        "补领": st.column_config.CheckboxColumn("补领", help="勾选=补发此书（仅未领）"),
                        "退书": st.column_config.CheckboxColumn("退书", help="勾选=退回此书（仅已领）"),
                    },
                    disabled=["编号", "学号", "姓名", "领取状态", "学生签名"],
                    column_order=["编号", "学号", "姓名", "领取状态", "学生签名", "补领", "退书"],
                    num_rows="fixed",
                )

                # 补领/退书操作区（在 form 内）
                to_supplement = edited_df[
                    (edited_df["领取状态"] == "未领") & (edited_df["补领"] == True)
                ]
                to_refund = edited_df[
                    (edited_df["领取状态"] == "已领") & (edited_df["退书"] == True)
                ]

                c_date, c_handler, c_btn1, c_btn2 = st.columns([1, 1, 1, 1])
                with c_date:
                    op_date = st.date_input("操作日期", value=date.today(), key="op_date_f")
                with c_handler:
                    op_handler = st.text_input("经手人", key="op_handler_f", placeholder="录入人")

                with c_btn1:
                    sup_label = (
                        f"补领 {len(to_supplement)} 人"
                        if len(to_supplement) > 0
                        else "补领"
                    )
                    sup_btn = st.form_submit_button(
                        sup_label, use_container_width=True, type="primary"
                    )
                with c_btn2:
                    ref_label = (
                        f"退书 {len(to_refund)} 人"
                        if len(to_refund) > 0
                        else "退书"
                    )
                    ref_btn = st.form_submit_button(
                        ref_label, use_container_width=True, type="secondary"
                    )

            # 处理 form 提交
            if sup_btn:
                if len(to_supplement) == 0:
                    st.warning("请先在表格中勾选需要「补领」的学生")
                else:
                    count = 0
                    tb_id = sel_tb[0]
                    for _, row in to_supplement.iterrows():
                        execute_sql(
                            "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, 1, %s, %s)",
                            (int(row["_sid"]), tb_id, op_date, op_handler or ""),
                        )
                        count += 1
                    st.success(f"成功补领 {count} 人")
                    st.rerun()

            if ref_btn:
                if len(to_refund) == 0:
                    st.warning("请先在表格中勾选需要「退书」的学生")
                else:
                    count = 0
                    for _, row in to_refund.iterrows():
                        execute_sql(
                            "DELETE FROM distributions WHERE student_id=%s AND textbook_id=%s",
                            (int(row["_sid"]), sel_tb[0]),
                        )
                        count += 1
                    st.success(f"成功退书 {count} 人")
                    st.rerun()

            # 提示信息（form 外）
            if not sup_btn and not ref_btn:
                sup_edited = edited_df[
                    (edited_df["领取状态"] == "未领") & (edited_df["补领"] == True)
                ]
                ref_edited = edited_df[
                    (edited_df["领取状态"] == "已领") & (edited_df["退书"] == True)
                ]
                if len(sup_edited) == 0 and len(ref_edited) == 0 and not_got > 0:
                    st.info(f"剩 {not_got} 人未领勾选「补领」；已领的可勾选「退书」")

            # 导出正式领书单
            st.markdown("##### 导出领书单")
            col_exp1, col_exp2, col_exp3, col_exp4 = st.columns([1.5, 1, 1, 1])
            with col_exp1:
                school_name = st.text_input(
                    "学校名称", value="湖南理工职业技术学院", key="c_school"
                )
            with col_exp2:
                handler_name = st.text_input("辅导员", value="", key="c_handler", placeholder="必填")
            with col_exp3:
                receiver_name = st.text_input(
                    "领书人", value="", key="c_receiver", placeholder="班级负责人"
                )
            with col_exp4:
                handler_phone = st.text_input(
                    "联系电话", value="", key="c_phone", placeholder="选填"
                )
            semester_label = c_semester[1].replace(" ", "")
            college_val = (
                students_df.iloc[0].get("college", "") if not students_df.empty else ""
            )

            def export_formal_booksheet():
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "领书单"
                nc = 10  # 固定10列（左右各5列: 序号、学号、姓名、领取状态、签字）

                # ── 重新统计（确保数据最新，不依赖闭包变量）──
                _dist_all = query_df(
                    "SELECT DISTINCT student_id FROM distributions WHERE textbook_id = %s",
                    (sel_tb[0],),
                )
                _got = len(_dist_all)

                # Row 1: 大标题
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                t = ws.cell(
                    row=1,
                    column=1,
                    value=f"{school_name} {semester_label}领书单",
                )
                t.font = Font(bold=True, size=18)
                t.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40

                # Row 2: 信息栏（分两行避免溢出）
                # Row 2a: 学院 + 班级 + 辅导员
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
                info_line1 = f"学院：{college_val}    班级：{c_class}    辅导员：{handler_name or '________'}    联系电话：{handler_phone or '________'}"
                inf = Font(bold=False, size=11)
                c2 = ws.cell(row=2, column=1, value=info_line1)
                c2.font = inf
                c2.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[2].height = 22

                # Row 2b: 领书人 + 统计数据
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=nc)
                info_line2 = f"领书人：{receiver_name or '________'}    班级人数：{total}人    领书人数：{_got}人"
                c3 = ws.cell(row=3, column=1, value=info_line2)
                c3.font = inf
                c3.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[3].height = 22

                # Row 4: 双栏表格表头
                left_headers = ["序号", "领书人", "学号", "领取", "签字"]
                right_headers = ["序号", "领书人", "学号", "领取", "签字"]
                half = math.ceil(total / 2)
                left_stu = students_df.iloc[:half].reset_index(drop=True)
                right_stu = students_df.iloc[half:].reset_index(drop=True)

                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                hfont = Font(bold=True, size=10)

                for i, h in enumerate(left_headers):
                    c = ws.cell(row=4, column=i + 1, value=h)
                    c.font = hfont
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                for i, h in enumerate(right_headers):
                    c = ws.cell(row=4, column=i + 6, value=h)
                    c.font = hfont
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")

                # 数据行 — OneCellAnchor 精确居中
                import base64
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
                from io import BytesIO as XLBytesIO
                from PIL import Image as PILImage

                dfont = Font(size=10)
                max_rows = max(len(left_stu), len(right_stu))
                SIG_W, SIG_H = 60, 30  # 签名缩放目标
                EMU = 9525
                COL_PX = 13 * 7         # 列宽 ≈91px
                ROW_PX = 55 * 4 // 3    # 行高 ≈73px

                for r_idx in range(max_rows):
                    row_num = 5 + r_idx
                    ws.row_dimensions[row_num].height = 55

                    if r_idx < len(left_stu):
                        stu = left_stu.iloc[r_idx]
                        sid = int(stu["id"])
                        has_book = sid in dist_ids
                        for c_idx, val in enumerate(
                            [
                                r_idx + 1,
                                stu["name"],
                                stu["student_id"],
                                "✓" if has_book else "",
                                "",
                            ],
                            1,
                        ):
                            cell = ws.cell(row=row_num, column=c_idx, value=val)
                            cell.font = dfont
                            cell.border = border
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        if sid in sig_map and sig_map[sid].get("data"):
                            try:
                                img_bytes = base64.b64decode(sig_map[sid]["data"])
                                pil_img = PILImage.open(XLBytesIO(img_bytes))
                                ratio = min(SIG_W / pil_img.width, SIG_H / pil_img.height)
                                w, h = int(pil_img.width * ratio), int(pil_img.height * ratio)
                                ox = max(0, (COL_PX - w) // 2) * EMU
                                oy = max(0, (ROW_PX - h) // 2) * EMU
                                sig_img = XLImage(XLBytesIO(img_bytes))
                                sig_img.width, sig_img.height = w, h
                                sig_img.anchor = OneCellAnchor(
                                    _from=AnchorMarker(col=4, row=row_num - 1, colOff=ox, rowOff=oy),
                                    ext=XDRPositiveSize2D(cx=w * EMU, cy=h * EMU),
                                )
                                ws.add_image(sig_img)
                            except:
                                pass

                    if r_idx < len(right_stu):
                        stu = right_stu.iloc[r_idx]
                        sid = int(stu["id"])
                        has_book = sid in dist_ids
                        for c_idx, val in enumerate(
                            [
                                half + r_idx + 1,
                                stu["name"],
                                stu["student_id"],
                                "✓" if has_book else "",
                                "",
                            ],
                            6,
                        ):
                            cell = ws.cell(row=row_num, column=c_idx, value=val)
                            cell.font = dfont
                            cell.border = border
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        if sid in sig_map and sig_map[sid].get("data"):
                            try:
                                img_bytes = base64.b64decode(sig_map[sid]["data"])
                                pil_img = PILImage.open(XLBytesIO(img_bytes))
                                ratio = min(SIG_W / pil_img.width, SIG_H / pil_img.height)
                                w, h = int(pil_img.width * ratio), int(pil_img.height * ratio)
                                ox = max(0, (COL_PX - w) // 2) * EMU
                                oy = max(0, (ROW_PX - h) // 2) * EMU
                                sig_img = XLImage(XLBytesIO(img_bytes))
                                sig_img.width, sig_img.height = w, h
                                sig_img.anchor = OneCellAnchor(
                                    _from=AnchorMarker(col=9, row=row_num - 1, colOff=ox, rowOff=oy),
                                    ext=XDRPositiveSize2D(cx=w * EMU, cy=h * EMU),
                                )
                                ws.add_image(sig_img)
                            except:
                                pass

                # 列宽（签名列 E/J 加宽以容纳图片）
                col_widths = [6, 10, 14, 8, 13, 6, 10, 14, 8, 13]
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = w

                wb.save(output)
                return output.getvalue()

            if st.button(
                "导出正式领书单",
                use_container_width=True,
                type="primary",
                key="c_formal_export",
            ):
                st.session_state.c_export_data = export_formal_booksheet()
                st.session_state.c_show_download = True
                st.rerun()

            if st.session_state.get("c_show_download"):
                st.download_button(
                    "点击下载",
                    data=st.session_state.c_export_data,
                    file_name=f"领书单_{c_class}_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key="c_dl_formal",
                )

            st.divider()
            st.caption("**退伍复学等新到学生**：先到「学生管理」新增，再到此处补领")

    # Tab 2: 发放情况汇总（含领书人明细）
    with tab2:
        st.markdown("#### 发放明细（含领书人信息）")

        with st.expander("🔍 筛选条件", expanded=True):
            col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1.5, 1.5, 1.5])
            with col_f1:
                r_semester = st.selectbox(
                "学期",
                [(0, "全部")] + [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                format_func=lambda x: x[1],
                key="cr_sem",
                index=1,
            )
        with col_f2:
            r_college = st.selectbox(
                "学院", ["全部"] + get_filtered_colleges(), key="cr_college"
            )
        with col_f3:
            if r_college != "全部":
                r_major_opts = ["全部"] + get_filtered_list(
                    "students", "major", "college = %s", (r_college,)
                )
            else:
                r_major_opts = ["全部"] + get_filtered_majors()
            r_major = st.selectbox("专业", r_major_opts, key="cr_major")
        with col_f4:
            r_class_where = "1=1"
            r_class_params = []
            if r_college != "全部":
                r_class_where += " AND college = %s"
                r_class_params.append(r_college)
            if r_major != "全部":
                r_class_where += " AND major = %s"
                r_class_params.append(r_major)
            r_class_opts = ["全部"] + (
                get_filtered_list(
                    "students", "class_name", r_class_where, tuple(r_class_params)
                )
                if r_class_params
                else class_names
            )
            r_class = st.selectbox("班级", r_class_opts, key="cr_class")

        # 获取发放明细（结算价格实时从 textbooks_master 计算）
        sql = f"""
            SELECT sem.name as semester_name, s.class_name, s.grade, s.college, s.major,
                   s.student_id, s.name as student_name,
                   t.name as textbook_name,
                   {PRICE_CALC} as calc_price,
                   d.quantity,
                   {PRICE_CALC} * d.quantity as subtotal,
                   d.distribute_date as 领书时间, d.handler as 经手人
            FROM distributions d
            JOIN students s ON d.student_id = s.id
            JOIN textbooks t ON d.textbook_id = t.id
            JOIN semesters sem ON t.semester_id = sem.id
            {PRICE_JOIN}
            WHERE 1=1
        """
        params = []
        if r_semester[0] > 0:
            sql += " AND t.semester_id = %s"
            params.append(r_semester[0])
        if r_college != "全部":
            sql += " AND s.college = %s"
            params.append(r_college)
        if r_major != "全部":
            sql += " AND s.major = %s"
            params.append(r_major)
        if r_class != "全部":
            sql += " AND s.class_name = %s"
            params.append(r_class)
        sql += " ORDER BY sem.id DESC, s.class_name, s.student_id, t.name"

        df = query_df(sql, tuple(params) if params else None)

        if not df.empty:
            df["subtotal"] = pd.to_numeric(df["subtotal"], errors="coerce").fillna(0)
            total_amount = df["subtotal"].sum()

            col_a, col_b = st.columns(2)
            with col_a:
                st.caption(f"共 **{len(df)}** 条发放记录")
            with col_b:
                st.metric("发放总金额", f"{total_amount:,.2f}")

            # 展示明细（含领书人信息，结算价=实洋>单价 折扣率）
            display_df = df.rename(
                columns={
                    "semester_name": "学期",
                    "class_name": "班级",
                    "grade": "年级",
                    "college": "学院",
                    "major": "专业",
                    "student_id": "学号",
                    "student_name": "领书人",
                    "textbook_name": "教材名称",
                    "calc_price": "结算价",
                    "quantity": "数量",
                    "subtotal": "小计",
                    "领书时间": "领书时间",
                    "经手人": "经手人",
                }
            )
            styled_dataframe(
                display_df[
                    [
                        "学期",
                        "年级",
                        "学院",
                        "专业",
                        "班级",
                        "学号",
                        "领书人",
                        "教材名称",
                        "结算价",
                        "数量",
                        "小计",
                        "领书时间",
                        "经手人",
                    ]
                ],
                hide_ids=True,
            )

            st.caption("💡 如需导出明细，请到「费用统计」页面操作")
