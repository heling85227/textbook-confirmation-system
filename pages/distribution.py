import streamlit as st
import pandas as pd
import io
import math
from datetime import datetime, date
from openpyxl.styles import Font, Border, Side

from database import query_df, execute_sql, PRICE_CALC, PRICE_JOIN
from components import show_header, excel_export, apply_excel_borders, styled_dataframe
from utils import (
    get_filtered_list, get_filtered_class_names, get_filtered_colleges,
    get_filtered_majors, get_filtered_grades, safe_int, safe_float, safe_str,
    safe_field, get_class_student_counts, read_excel_upload, write_import_log,
    get_current_academic_info, make_template_df
)


def distribution_management():
    show_header("📦 教材发放表", "按征订表带出教材清单，按班级打印领书单供学生签字确认")

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先添加学期")
        return

    class_names = get_filtered_class_names()
    colleges = get_filtered_colleges()
    majors = get_filtered_majors()

    tab1, tab2 = st.tabs(["📋 发书单", "📥 导入 Excel"])

    # ── Tab 1: 发书单 ──
    with tab1:
        with st.expander("🔍 筛选条件", expanded=True):
            col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1.5, 1.5, 1.5])
            with col_f1:
                f_sem = st.selectbox("学期", [(0, "全部")] + [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                                     format_func=lambda x: x[1], key="dl_sem", index=1)
            with col_f2:
                f_college = st.selectbox("学院", ["全部"] + colleges, key="dl_college")
            with col_f3:
                # 专业：基于所选学院级联过滤
                if f_college != "全部":
                    dl_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (f_college,))
                else:
                    dl_major_opts = ["全部"] + majors
                f_major = st.selectbox("专业", dl_major_opts, key="dl_major")
            with col_f4:
                # 班级：基于所选学院+专业级联过滤
                dl_class_where = "1=1"; dl_class_params = []
                if f_college != "全部":
                    dl_class_where += " AND college = %s"; dl_class_params.append(f_college)
                if f_major != "全部":
                    dl_class_where += " AND major = %s"; dl_class_params.append(f_major)
                dl_class_opts = ["全部"] + (get_filtered_list("students", "class_name", dl_class_where, tuple(dl_class_params)) if dl_class_params else get_filtered_class_names())
                f_class = st.selectbox("班级", dl_class_opts, key="dl_class")

        if f_sem[0] > 0:
            sem_id = f_sem[0]
            ay, semester_n = get_current_academic_info()
            sem_name = f_sem[1]
            sem_label = sem_name.replace(" ", "")

            # ── 录入发放记录 ──
            with st.expander("➕ 录入发放记录", expanded=False):
                st.caption("选择班级→学生→教材，手动录入一条发放记录")

                # 录入专用的班级选择
                entry_class_where = "1=1"; entry_class_params = []
                if f_college != "全部":
                    entry_class_where += " AND college = %s"; entry_class_params.append(f_college)
                if f_major != "全部":
                    entry_class_where += " AND major = %s"; entry_class_params.append(f_major)
                entry_class_opts = get_filtered_list("students", "class_name", entry_class_where, tuple(entry_class_params)) if entry_class_params else get_filtered_class_names()
                if not entry_class_opts:
                    entry_class_opts = class_names

                col_e1, col_e2, col_e3 = st.columns([1.5, 1.5, 1])
                with col_e1:
                    entry_class = st.selectbox("发放班级 *", entry_class_opts, key="entry_class")
                with col_e2:
                    if entry_class:
                        entry_stu = query_df(
                            "SELECT id, student_id, name FROM students WHERE class_name = %s ORDER BY student_id",
                            (entry_class,))
                        stu_opts = [(r["id"], f"{r['student_id']} {r['name']}") for _, r in entry_stu.iterrows()]
                        entry_student = st.selectbox("学生 *", stu_opts, format_func=lambda x: x[1], key="entry_student")
                    else:
                        entry_student = None
                with col_e3:
                    if entry_class:
                        # 查该班该学期的教材（从 textbook_orders + textbooks_master，确保有数据）
                        entry_tb = query_df("""
                            SELECT t.id, m.name, m.course_name, m.price
                            FROM textbook_orders o
                            JOIN textbooks_master m ON o.textbook_id = m.id
                            LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                            WHERE o.semester_id = %s AND o.class_name = %s
                            ORDER BY m.name
                        """, (sem_id, entry_class))
                        if entry_tb.empty:
                            entry_textbook = None
                            st.caption("⚠️ 该班暂无教材")
                        else:
                            # 确保 textbooks 表有记录（没有则自动创建），返回 textbooks.id
                            tb_opts = []
                            for _, r in entry_tb.iterrows():
                                tbid = r["id"]
                                if tbid is None or pd.isna(tbid):
                                    # 自动创建 textbooks 记录
                                    execute_sql(
                                        "INSERT INTO textbooks (semester_id, grade, college, major, class_name, name, price, course_name, quantity, remark) VALUES (%s, '', '', '', %s, %s, %s, %s, 0, '[逐条录入·自动创建]')",
                                        (sem_id, entry_class, r["name"], r["price"], r.get("course_name") or ""))
                                    nr = query_df("SELECT id FROM textbooks WHERE semester_id=%s AND class_name=%s AND name=%s",
                                        (sem_id, entry_class, r["name"]))
                                    tbid = int(nr.iloc[0]["id"]) if not nr.empty else None
                                if tbid:
                                    tb_opts.append((tbid, f"{r['name']}" + (f"({r['course_name']})" if r.get('course_name') else "") + f" ¥{r['price']:.2f}"))
                            entry_textbook = st.selectbox("教材 *", tb_opts, format_func=lambda x: x[1], key="entry_textbook") if tb_opts else None
                    else:
                        entry_textbook = None

                col_e4, col_e5, col_e6 = st.columns([1, 1.5, 1.5])
                with col_e4:
                    entry_qty = st.number_input("数量", min_value=1, value=1, step=1, key="entry_qty")
                with col_e5:
                    entry_date = st.date_input("发放日期", value=date.today(), key="entry_date")
                with col_e6:
                    entry_handler = st.text_input("经手人", key="entry_handler", placeholder="录入人")

                if st.button("➕ 确认录入", use_container_width=True, type="primary", key="entry_submit"):
                    if not entry_class:
                        st.warning("⚠️ 请选择发放班级")
                    elif entry_student is None:
                        st.warning("⚠️ 请选择学生")
                    elif entry_textbook is None:
                        st.warning("⚠️ 该班暂无教材，请先在「征订总表」中下发教材")
                    else:
                        tb_id = int(entry_textbook[0])
                        execute_sql(
                            "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, %s, %s, %s)",
                            (int(entry_student[0]), tb_id, entry_qty, entry_date, entry_handler or ""))
                        st.success(f"✅ 已录入发放：{entry_student[1]} ← 《{entry_textbook[1]}》 ×{entry_qty}")
                        st.rerun()

            # ── 批量发放录入（从征订表带出数据，支持多班级）──
            st.divider()
            st.markdown("#### 📝 批量发放录入")
            st.caption('从征订表带出所选班级全部学生的教材清单，勾选「已发放」后一键批量确认。支持多班同时操作。')

            batch_class_where = "1=1"; batch_class_params = []
            if f_college != "全部":
                batch_class_where += " AND college = %s"; batch_class_params.append(f_college)
            if f_major != "全部":
                batch_class_where += " AND major = %s"; batch_class_params.append(f_major)
            batch_class_opts = get_filtered_list("students", "class_name", batch_class_where, tuple(batch_class_params)) if batch_class_params else get_filtered_class_names()
            if not batch_class_opts:
                batch_class_opts = class_names

            cc1, cc2 = st.columns([4, 1])
            with cc1:
                batch_classes = st.multiselect("批量录入班级（可多选）*", batch_class_opts, key="batch_classes",
                    placeholder="选择一个或多个班级...")
            with cc2:
                def _select_all_classes():
                    st.session_state["batch_classes"] = batch_class_opts
                st.button("☑ 全选班级", key="bsel_all_classes", on_click=_select_all_classes, use_container_width=True)

            if batch_classes:
                # ── 按班级收集数据 ──
                all_rows = []
                all_existing_set = set()
                total_stu_count = 0
                unique_tb_names = set()
                has_empty_stu = False
                has_empty_tb = False

                for bc in batch_classes:
                    # 查该班学生
                    bc_stu = query_df(
                        "SELECT id, student_id, name FROM students WHERE class_name = %s ORDER BY student_id",
                        (bc,))
                    if bc_stu.empty:
                        has_empty_stu = True
                        continue

                    # 查该班该学期的教材
                    raw_tb = query_df("""
                        SELECT o.textbook_id as master_id, m.name, m.course_name, m.isbn, m.publisher, m.editor, m.price
                        FROM textbook_orders o
                        JOIN textbooks_master m ON o.textbook_id = m.id
                        WHERE o.semester_id = %s AND o.class_name = %s
                        ORDER BY m.name
                    """, (sem_id, bc))

                    if raw_tb.empty:
                        has_empty_tb = True
                        continue

                    # 确保 textbooks 表存在对应记录
                    tb_id_map = {}
                    for _, rt in raw_tb.iterrows():
                        mid = int(rt["master_id"])
                        tbid_row = query_df(
                            "SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s AND name = %s",
                            (sem_id, bc, rt["name"]))
                        if not tbid_row.empty:
                            tb_id_map[mid] = int(tbid_row.iloc[0]["id"])
                        else:
                            execute_sql(
                                """INSERT INTO textbooks
                                (semester_id, grade, college, major, class_name, name, isbn, publisher, editor, price, course_name, quantity, remark)
                                VALUES (%s, '', '', '', %s, %s, %s, %s, %s, %s, %s, 0, '[自动创建·来自征订数据]')""",
                                (sem_id, bc, rt["name"], rt.get("isbn") or "",
                                 rt.get("publisher") or "", rt.get("editor") or "",
                                 rt["price"], rt.get("course_name") or ""))
                            new_row = query_df(
                                "SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s AND name = %s",
                                (sem_id, bc, rt["name"]))
                            if not new_row.empty:
                                tb_id_map[mid] = int(new_row.iloc[0]["id"])

                    # 重建 batch_tb
                    batch_tb = []
                    for _, rt in raw_tb.iterrows():
                        mid = int(rt["master_id"])
                        if mid in tb_id_map:
                            batch_tb.append({
                                "textbook_id": tb_id_map[mid],
                                "name": rt["name"],
                                "course_name": rt.get("course_name") or "",
                                "price": rt["price"],
                            })
                            unique_tb_names.add(rt["name"])

                    # 查已有发放记录
                    existing = query_df(
                        "SELECT student_id, textbook_id FROM distributions WHERE textbook_id IN (SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s)",
                        (sem_id, bc))
                    for _, er in existing.iterrows():
                        all_existing_set.add((int(er["student_id"]), int(er["textbook_id"])))

                    # 构建学生×教材矩阵
                    for _, stu in bc_stu.iterrows():
                        sid = int(stu["id"])
                        for tb in batch_tb:
                            tid = tb["textbook_id"]
                            already = (sid, tid) in all_existing_set
                            all_rows.append({
                                "班级": bc,
                                "student_id": sid,
                                "textbook_id": tid,
                                "学号": stu["student_id"],
                                "姓名": stu["name"],
                                "教材名称": tb["name"],
                                "课程": tb["course_name"],
                                "单价": tb["price"],
                                "已发放": already,
                            })
                    total_stu_count += len(bc_stu)

                if not all_rows:
                    if has_empty_stu:
                        st.warning("⚠️ 部分班级暂无学生")
                    if has_empty_tb:
                        st.warning("⚠️ 部分班级暂无教材，请先在「征订总表」中下发教材")
                else:
                    batch_df = pd.DataFrame(all_rows)
                    st.caption(f"共 {len(batch_classes)} 个班、{total_stu_count} 人 × {len(unique_tb_names)} 种教材 = {len(batch_df)} 条记录，已发放 {len(all_existing_set)} 条")

                    # 全选/取消全选 按钮
                    sel_col1, sel_col2, sel_col3, sel_col4 = st.columns([1, 1, 2, 2])
                    with sel_col1:
                        if st.button("☑ 全选", key="bsel_all", use_container_width=True):
                            st.session_state["batch_select_mode"] = "all"
                            st.rerun()
                    with sel_col2:
                        if st.button("☐ 全不选", key="bsel_none", use_container_width=True):
                            st.session_state["batch_select_mode"] = "none"
                            st.rerun()
                    with sel_col3:
                        if st.button("🔄 恢复已发放", key="bsel_restore", use_container_width=True):
                            st.session_state.pop("batch_select_mode", None)
                            st.rerun()
                    # 根据 session_state 调整初始勾选
                    sel_mode = st.session_state.get("batch_select_mode", "")
                    if sel_mode == "all":
                        batch_df["已发放"] = True
                    elif sel_mode == "none":
                        batch_df["已发放"] = False

                    with st.form("batch_dist_form"):
                        edited_batch = st.data_editor(
                            batch_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(35 * len(batch_df) + 38, 600),
                            column_config={
                                "班级": st.column_config.TextColumn("班级", disabled=True, width="small", alignment="center"),
                                "student_id": None,
                                "textbook_id": None,
                                "学号": st.column_config.TextColumn("学号", disabled=True, width="small", alignment="center"),
                                "姓名": st.column_config.TextColumn("姓名", disabled=True, width="small", alignment="center"),
                                "教材名称": st.column_config.TextColumn("教材名称", disabled=True, alignment="center"),
                                "课程": st.column_config.TextColumn("课程", disabled=True, alignment="center"),
                                "单价": st.column_config.NumberColumn("单价", disabled=True, format="¥%.2f", width="small", alignment="center"),
                                "已发放": st.column_config.CheckboxColumn("已发放", help="勾选=该生已领此书"),
                            },
                            disabled=["班级", "学号", "姓名", "教材名称", "课程", "单价"],
                            num_rows="fixed"
                        )

                        col_bf1, col_bf2 = st.columns([1.5, 1])
                        with col_bf1:
                            batch_date = st.date_input("发放日期", value=date.today(), key="batch_date")
                        with col_bf2:
                            batch_handler = st.text_input("经手人", key="batch_handler", placeholder="录入人")

                        batch_btn = st.form_submit_button("✅ 批量确认发放", use_container_width=True, type="primary")

                    if batch_btn:
                        to_insert = edited_batch[edited_batch["已发放"] == True]
                        new_records = []
                        for _, row in to_insert.iterrows():
                            key = (int(row["student_id"]), int(row["textbook_id"]))
                            if key not in all_existing_set:
                                new_records.append(row)

                        if not new_records:
                            st.info("📭 没有新的发放记录需要保存（全部已存在）")
                        else:
                            count = 0
                            for row in new_records:
                                tb_id = int(row["textbook_id"])
                                execute_sql(
                                    "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, 1, %s, %s)",
                                    (int(row["student_id"]), tb_id, batch_date, batch_handler or ""))
                                count += 1
                            st.success(f"✅ 批量录入 {count} 条发放记录")
                            st.rerun()

            st.divider()

            # ── 发书单预览 ──
            st.markdown("#### 📋 发书单预览")
            st.caption("下方展示各班的教材配发清单，供核对用")
            tb_sql = """SELECT o.class_name, o.grade, o.college, o.major, o.quantity,
                               m.name, m.course_name, m.isbn, m.publisher, m.editor, m.price,
                               o.id as order_id
                        FROM textbook_orders o
                        JOIN textbooks_master m ON o.textbook_id = m.id
                        WHERE o.semester_id = %s"""
            tb_params = [sem_id]
            if f_college != "全部":
                tb_sql += " AND o.college = %s"; tb_params.append(f_college)
            if f_major != "全部":
                tb_sql += " AND o.major = %s"; tb_params.append(f_major)
            if f_class != "全部":
                tb_sql += " AND o.class_name = %s"; tb_params.append(f_class)
            tb_sql += " ORDER BY o.class_name, m.name"

            textbooks = query_df(tb_sql, tuple(tb_params))

            if textbooks.empty:
                st.info("📭 当前条件下暂无已下发的教材领用数据，请先在「征订总表」中执行「一键下发」")
            else:
                # 按班级分组
                class_books = {}
                for _, tb in textbooks.iterrows():
                    cn = tb["class_name"]
                    if cn not in class_books:
                        class_books[cn] = []
                    class_books[cn].append(tb)

                # 发书单预览
                st.caption("📋 **发书单** — 按学生列出教材及费用，供发书核对")
                for cn, tbs in class_books.items():
                    students = query_df(
                        "SELECT student_id, name FROM students WHERE class_name = %s ORDER BY student_id", (cn,))
                    rows = []
                    for i, (_, stu) in enumerate(students.iterrows(), 1):
                        book_names = "、".join([f"《{b['name']}》" + (f"({b['course_name']})" if b.get('course_name') else "") for b in tbs])
                        total = sum(b["price"] for b in tbs)
                        rows.append({"序号": i, "学号": stu["student_id"], "姓名": stu["name"],
                                     "领取教材": book_names, "教材数": len(tbs), "合计(元)": round(total, 2),
                                     "领书人签字": "", "联系电话": "", "领书时间": ""})
                    lsd = pd.DataFrame(rows)
                    total_fee = sum(r["合计(元)"] for r in rows)
                    with st.expander(f"📦 {cn}（{len(students)}人 × {len(tbs)}种教材 = ¥{total_fee:.2f}）", expanded=False):
                        styled_dataframe(lsd, hide_ids=True)

                # 导出 - 教材发放清单（正式格式）
                st.divider()
                with st.expander("📥 导出正式教材发放清单", expanded=False):
                    col_sl1, col_sl2 = st.columns([2, 1])
                    with col_sl1:
                        lq_school = st.text_input("学校名称", value="湖南理工职业技术学院", key="lq_school")
                    with col_sl2:
                        lq_phone = st.text_input("联系电话", value="", key="lq_phone", placeholder="选填")

                    def export_formal_bookreceipt():
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                        from openpyxl.worksheet.properties import PageSetupProperties
                        import io
                        output = io.BytesIO()
                        wb = Workbook()
                        default_ws = wb.active
                        wb.remove(default_ws)

                        cell_border = Border(
                            left=Side(style="medium"), right=Side(style="medium"),
                            top=Side(style="medium"), bottom=Side(style="medium")
                        )
                        blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                        hfont = Font(bold=True, size=10)
                        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        dfont = Font(size=10)

                        for cn, tbs in class_books.items():
                            ws = wb.create_sheet(title=str(cn)[:31])
                            nc = 9

                            cq = query_df("SELECT COUNT(*) as c FROM students WHERE class_name = %s", (cn,))
                            stu_count = int(cq.iloc[0]["c"]) if not cq.empty else 0
                            college_q = query_df("SELECT DISTINCT college FROM students WHERE class_name = %s", (cn,))
                            tb_college = college_q.iloc[0]["college"] if not college_q.empty else ""

                            for ci in range(1, nc + 1):
                                ws.cell(row=1, column=ci).border = cell_border
                            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                            t = ws.cell(row=1, column=1, value=f"{lq_school} {sem_label}教材发放清单")
                            t.font = Font(bold=True, size=14)
                            t.alignment = Alignment(horizontal="center", vertical="center")
                            for ci in range(1, nc + 1):
                                ws.cell(row=1, column=ci).border = cell_border
                            ws.row_dimensions[1].height = 30

                            for ci in range(1, nc + 1):
                                ws.cell(row=2, column=ci).border = cell_border
                            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
                            info_text = f"学院：{tb_college}    学期：{sem_name}    人数：{stu_count}人    电话：{lq_phone or '_____________'}"
                            ci2 = ws.cell(row=2, column=1, value=info_text)
                            ci2.font = Font(size=10)
                            ci2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            for ci in range(1, nc + 1):
                                ws.cell(row=2, column=ci).border = cell_border
                            ws.row_dimensions[2].height = 22

                            headers = ["班级", "序号", "课程", "教材名称", "主编", "出版社", "单价", "实发", "领书人"]
                            for i, h in enumerate(headers, 1):
                                c = ws.cell(row=3, column=i, value=h)
                                c.font = hfont; c.border = cell_border; c.fill = blue_fill
                                c.alignment = center
                            ws.row_dimensions[3].height = 22

                            for idx, tb in enumerate(tbs):
                                rn = 4 + idx
                                ws.row_dimensions[rn].height = 26
                                vals = [
                                    cn if idx == 0 else "",
                                    idx + 1,
                                    tb.get("course_name", ""),
                                    tb["name"],
                                    tb.get("editor", ""),
                                    tb.get("publisher", ""),
                                    round(float(tb.get("price", 0)), 2),
                                    "",
                                    ""
                                ]
                                for ci, v in enumerate(vals, 1):
                                    cell = ws.cell(row=rn, column=ci, value=v)
                                    cell.font = dfont; cell.border = cell_border
                                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                            if len(tbs) > 1:
                                ws.merge_cells(start_row=4, start_column=1, end_row=4 + len(tbs) - 1, end_column=1)
                            for rrow in range(4, 4 + len(tbs)):
                                for ccol in range(1, nc + 1):
                                    cell = ws.cell(row=rrow, column=ccol)
                                    cell.border = cell_border
                                    cell.alignment = center

                            widths = [10, 6, 14, 28, 10, 18, 8, 6, 10]
                            for i, w in enumerate(widths, 1):
                                ws.column_dimensions[chr(64 + i)].width = w

                            ws.page_setup.paperSize = 9
                            ws.page_setup.orientation = 'portrait'
                            ws.page_setup.fitToWidth = 1
                            ws.page_setup.fitToHeight = 0
                            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                            ws.page_margins.left = 0.4
                            ws.page_margins.right = 0.4
                            ws.page_margins.top = 0.5
                            ws.page_margins.bottom = 0.5
                            ws.page_margins.header = 0.3
                            ws.page_margins.footer = 0.3
                            ws.print_title_rows = '1:3'

                        wb.save(output)
                        return output.getvalue()

                    if st.button("📥 导出正式教材发放清单", use_container_width=True, type="primary", key="lq_formal_export"):
                        st.session_state.lq_export_data = export_formal_bookreceipt()
                        st.session_state.lq_show_download = True
                        st.rerun()

                    if st.session_state.get("lq_show_download"):
                        st.download_button(
                            "✅ 点击下载",
                            data=st.session_state.lq_export_data,
                            file_name=f"教材发放清单_{date.today()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, type="primary", key="lq_dl"
                        )
        else:
            st.info("👆 请先选择学期")

    # ── Tab 2: 导入 Excel（批量导入发放记录）──
    with tab2:
        st.markdown("#### 📥 从 Excel 批量导入发放记录")
        st.caption("Excel 列名建议：学号（或身份证号）、教材名称、领取数量、发放日期、经手人")

        # 下载模板
        with st.expander("📄 模板下载与说明", expanded=False):
            template_cols = ["学号", "身份证号", "教材名称", "领取数量", "发放日期", "经手人"]
            template_df = make_template_df(template_cols)
            template_bytes = excel_export(template_df, "发放记录导入模板")
            st.download_button(
                "📄 下载导入模板",
                data=template_bytes,
                file_name="发放记录导入模板.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="secondary"
            )

        col1, col2 = st.columns(2)
        with col1:
            i_semester = st.selectbox("学期 *", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                              format_func=lambda x: x[1], key="di_sem")
        with col2:
            i_class = st.selectbox("班级 *", ["请选择班级"] + class_names, key="di_class")

        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="dist_upload")
        if uploaded:
            try:
                raw_df = read_excel_upload(uploaded)
                st.info(f"📄 检测到 {len(raw_df)} 行")

                col_map = {}
                for col in raw_df.columns:
                    cl = col.strip().lower()
                    if "学号" in cl or "student_id" in cl: col_map["student_id"] = col
                    elif "身份证" in cl or "id_card" in cl: col_map["id_card"] = col
                    elif "教材" in cl or "书名" in cl or "textbook" in cl: col_map["textbook_name"] = col
                    elif "数量" in cl or "quantity" in cl: col_map["quantity"] = col
                    elif "日期" in cl or "date" in cl: col_map["date"] = col
                    elif "经手" in cl or "handler" in cl: col_map["handler"] = col

                if not ("student_id" in col_map or "id_card" in col_map) or "textbook_name" not in col_map:
                    st.error("❌ 缺少必要列：学号/身份证号、教材名称")
                else:
                    if st.button("✅ 确认导入", use_container_width=True, type="primary"):
                        mapped = raw_df.rename(columns={v: k for k, v in col_map.items()})
                        sem_id = i_semester[0]

                        # 预加载映射
                        all_students = query_df("SELECT id, student_id, id_card FROM students WHERE class_name = %s", (i_class,))
                        students_map = {}
                        for _, s in all_students.iterrows():
                            students_map[str(s["student_id"])] = s["id"]
                            students_map[str(s["id_card"])] = s["id"]

                        all_textbooks = query_df(
                            "SELECT m.id, m.name FROM textbook_orders o "
                            "JOIN textbooks_master m ON o.textbook_id = m.id "
                            "WHERE o.semester_id = %s AND o.class_name = %s", (sem_id, i_class))
                        textbooks_map = {t["name"]: t["id"] for _, t in all_textbooks.iterrows()}

                        success, errors, total_rows = 0, [], len(mapped)
                        progress_bar = st.progress(0, text="正在导入...")
                        for i, (_, row) in enumerate(mapped.iterrows()):
                            sid_val = row.get("student_id")
                            lookup = safe_str(sid_val) if pd.notna(sid_val) else safe_str(row.get("id_card"))
                            if not lookup:
                                errors.append(f"空学号/身份证号"); continue
                            sid = students_map.get(lookup)
                            if sid is None:
                                errors.append(f"学生「{lookup}」未找到"); continue
                            tname = safe_str(row.get("textbook_name"))
                            tid = textbooks_map.get(tname)
                            if tid is None:
                                errors.append(f"教材「{tname}」未找到"); continue
                            try:
                                q = safe_int(row.get("quantity", 1), 1)
                                d = row.get("date")
                                h = row.get("handler", "")
                                if pd.notna(d):
                                    try:
                                        d = pd.Timestamp(d).date()
                                    except Exception:
                                        d = date.today()
                                else:
                                    d = date.today()
                                execute_sql(
                                    "INSERT INTO distributions (student_id,textbook_id,quantity,distribute_date,handler) VALUES (%s,%s,%s,%s,%s)",
                                    (sid, tid, q, d, safe_str(h)))
                                success += 1
                            except Exception as e:
                                errors.append(f"{lookup}×{tname}: {str(e)[:100]}")
                            progress_bar.progress((i + 1) / total_rows, text=f"已处理 {i+1}/{total_rows}")
                        # 统一展示结果
                        mc_ok3, mc_err3 = st.columns(2)
                        mc_ok3.metric("✅ 导入成功", success)
                        mc_err3.metric("⚠️ 导入失败", len(errors))
                        if errors:
                            with st.expander(f"查看 {len(errors)} 条失败详情"):
                                for err in errors:
                                    st.caption(f"• {err}")
                        write_import_log("发放记录导入", uploaded.name, total_rows, success, errors)
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")
