"""
学生教材费用核对系统 v2.5
====================
- 管理员：学期管理、学生管理、教材征订、教材发放、领书确认、费用统计
- 学生端：凭身份证或学号查询个人教材费用
- 支持 MySQL（生产）/ SQLite（本地测试）
- v2.4：发放表/领书单导出共性字段移至表头；新增班级人数、征订总数表头信息
"""
import streamlit as st
import pymysql
import sqlite3
import pandas as pd
import configparser
import io
import os
import re
import bcrypt
from datetime import datetime, date
from contextlib import contextmanager
from openpyxl.styles import Font

# ═════════════════════════════════════════════════════════
# 全局配置
# ═════════════════════════════════════════════════════════

st.set_page_config(
    page_title="学生教材费用核对系统",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 自定义 CSS ──
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, 'Microsoft YaHei', sans-serif;
    }
    
    :root {
        --primary: #1e40af;
        --primary-light: #3b82f6;
        --primary-bg: #eff6ff;
        --success: #059669;
        --success-bg: #ecfdf5;
        --warning: #d97706;
        --warning-bg: #fffbeb;
        --danger: #dc2626;
        --danger-bg: #fef2f2;
        --gray-50: #f9fafb;
        --gray-100: #f3f4f6;
        --gray-200: #e5e7eb;
        --gray-600: #4b5563;
        --gray-700: #374151;
        --gray-800: #1f2937;
        --radius: 12px;
        --shadow: 0 1px 3px rgba(0,0,0,0.08), 0 1px 2px rgba(0,0,0,0.06);
        --shadow-md: 0 4px 12px rgba(0,0,0,0.08);
    }

    .app-header {
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 50%, #6366f1 100%);
        color: white;
        padding: 24px 32px;
        border-radius: var(--radius);
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(30,64,175,0.2);
    }
    .app-header h1 {
        color: white !important;
        font-size: 28px;
        font-weight: 700;
        margin: 0;
        padding: 0;
    }
    .app-header p {
        color: rgba(255,255,255,0.85);
        margin: 4px 0 0 0;
        font-size: 14px;
    }

    .info-card {
        background: white;
        border: 1px solid var(--gray-200);
        border-radius: var(--radius);
        padding: 20px 24px;
        margin-bottom: 16px;
        box-shadow: var(--shadow);
        transition: all 0.2s;
    }
    .info-card:hover {
        box-shadow: var(--shadow-md);
        transform: translateY(-1px);
    }
    .info-card h3 {
        font-size: 16px;
        font-weight: 600;
        color: var(--gray-800);
        margin: 0 0 4px 0;
    }
    .info-card p {
        font-size: 13px;
        color: var(--gray-600);
        margin: 0;
    }

    .student-card {
        background: linear-gradient(135deg, #eff6ff 0%, #f0fdf4 100%);
        border: 1px solid #bfdbfe;
        border-radius: var(--radius);
        padding: 20px;
        margin-bottom: 12px;
    }

    .stat-number {
        font-size: 32px;
        font-weight: 700;
        color: var(--primary);
        line-height: 1.2;
    }
    .stat-label {
        font-size: 13px;
        color: var(--gray-600);
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    [data-testid="stMetricValue"] {
        font-size: 28px !important;
        font-weight: 700 !important;
    }
    [data-testid="stMetricLabel"] {
        font-size: 13px !important;
        color: var(--gray-600) !important;
    }
    [data-testid="stMetricDelta"] {
        font-size: 14px !important;
    }

    .stButton > button {
        border-radius: 8px !important;
        font-weight: 500 !important;
        transition: all 0.2s !important;
        border: none !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
    }
    .stButton > button:active {
        transform: translateY(0);
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8fafc 0%, #eff6ff 100%);
        border-right: 1px solid var(--gray-200);
    }
    [data-testid="stSidebar"] h1 {
        font-size: 22px !important;
        color: var(--primary) !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: var(--gray-50);
        padding: 6px;
        border-radius: 10px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px !important;
        padding: 8px 16px !important;
        font-weight: 500 !important;
        transition: all 0.2s;
    }
    .stTabs [aria-selected="true"] {
        background: white !important;
        box-shadow: var(--shadow);
    }

    div[data-testid="stForm"] {
        border: 1px solid var(--gray-200);
        border-radius: var(--radius);
        padding: 24px;
        background: white;
        box-shadow: var(--shadow);
    }

    [data-testid="stDataFrame"] {
        border-radius: var(--radius) !important;
        overflow: hidden;
        border: 1px solid var(--gray-200) !important;
    }
    [data-testid="stDataFrame"] th {
        background: var(--primary-bg) !important;
        color: var(--primary) !important;
        font-weight: 600 !important;
        font-size: 13px !important;
    }

    .streamlit-expanderHeader {
        background: var(--gray-50);
        border-radius: 8px !important;
        font-weight: 600 !important;
    }

    hr {
        border-color: var(--gray-200) !important;
        margin: 24px 0 !important;
    }

    input, textarea, .stSelectbox > div > div {
        border-radius: 8px !important;
    }

    [data-testid="stFileUploader"] {
        border-radius: var(--radius) !important;
        border: 2px dashed var(--gray-200) !important;
        padding: 20px !important;
        transition: all 0.2s;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: var(--primary-light) !important;
        background: var(--primary-bg) !important;
    }

    .element-container:has([data-testid="stSuccess"]) {
        animation: slideIn 0.3s ease;
    }
    @keyframes slideIn {
        from { opacity: 0; transform: translateY(-10px); }
        to { opacity: 1; transform: translateY(0); }
    }

    @media print {
        [data-testid="stSidebar"],
        .stButton,
        .stDownloadButton,
        [data-testid="stFileUploader"],
        header,
        .stTabs [data-baseweb="tab-list"],
        .st-emotion-cache-ocsh0s {
            display: none !important;
        }
        .stTabs [role="tabpanel"] {
            display: block !important;
        }
        [data-testid="stVerticalBlock"] {
            gap: 0 !important;
        }
        body {
            font-size: 12px;
        }
        table {
            font-size: 11px;
        }
    }

    .user-badge {
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: var(--primary-bg);
        color: var(--primary);
        padding: 8px 16px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 14px;
        margin-bottom: 16px;
    }

    .class-tag {
        display: inline-block;
        background: var(--primary-bg);
        color: var(--primary);
        padding: 4px 12px;
        border-radius: 999px;
        font-size: 13px;
        font-weight: 500;
        margin: 2px;
    }

    /* 多选框样式 */
    .stMultiSelect [data-baseweb="tag"] {
        background: var(--primary-bg) !important;
        color: var(--primary) !important;
    }

    /* 模板下载按钮 */
    .template-btn {
        background: #f0fdf4 !important;
        color: #059669 !important;
        border: 1px solid #bbf7d0 !important;
    }

    /* ── data_editor 表格内的 checkbox 可见性修复 ── */
    [data-testid="stDataEditor"] input[type="checkbox"],
    [data-testid="stDataFrame"] input[type="checkbox"] {
        accent-color: #1677ff !important;
        width: 18px !important;
        height: 18px !important;
        opacity: 1 !important;
        visibility: visible !important;
    }
    [data-testid="stDataEditor"] input[type="checkbox"]:checked,
    [data-testid="stDataFrame"] input[type="checkbox"]:checked {
        accent-color: #1677ff !important;
    }
    /* checkbox 列单元格居中 */
    [data-testid="stDataEditor"] td:has(input[type="checkbox"]),
    [data-testid="stDataFrame"] td:has(input[type="checkbox"]) {
        text-align: center !important;
    }

    /* ── 红色删除按钮：全局 secondary 按钮变红，侧边栏另作恢复 ── */
    button[kind="secondary"] {
        background-color: #ef4444 !important;
        color: #fff !important;
        border: 1px solid #dc2626 !important;
    }
    button[kind="secondary"]:hover:not(:disabled) {
        background-color: #dc2626 !important;
        border-color: #b91c1c !important;
    }
    button[kind="secondary"]:disabled {
        background-color: #e5e7eb !important;
        color: #9ca3af !important;
        border-color: #d1d5db !important;
    }
    /* 侧边栏内的按钮恢复默认样式 */
    [data-testid="stSidebar"] button[kind="secondary"] {
        background-color: inherit !important;
        color: inherit !important;
        border: inherit !important;
    }

    /* ── 分页栏容器 ── */
    .pagination-bar {
        display: flex;
        align-items: center;
        justify-content: flex-end;
        gap: 12px;
        padding: 10px 0;
        font-size: 14px;
        color: #4b5563;
    }
    .pagination-bar .page-info {
        font-weight: 500;
        min-width: 80px;
        text-align: center;
    }
    .pagination-bar .total-info {
        color: #6b7280;
        font-size: 13px;
    }
    .pagination-bar button {
        min-width: 32px !important;
        height: 32px !important;
        padding: 0 8px !important;
    }
</style>
""", unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════
# 配置加载
# ═════════════════════════════════════════════════════════

def load_config():
    config = configparser.ConfigParser()
    config_path = os.path.join(os.path.dirname(__file__), "config.ini")
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    config.read(config_path, encoding="utf-8")
    return config

CONFIG = load_config()
DB_TYPE = CONFIG.get("database", "type", fallback="mysql")

DB_CONFIG = {
    "host": os.environ.get("DB_HOST", CONFIG.get("database", "host", fallback="localhost")),
    "port": int(os.environ.get("DB_PORT", CONFIG.get("database", "port", fallback="3306"))),
    "user": os.environ.get("DB_USER", CONFIG.get("database", "user", fallback="root")),
    "password": os.environ.get("DB_PASSWORD", CONFIG.get("database", "password", fallback="")),
    "database": os.environ.get("DB_NAME", CONFIG.get("database", "database", fallback="textbook_fee")),
    "charset": "utf8mb4",
}

ADMIN_PASSWORD_HASH = os.environ.get(
    "ADMIN_PASSWORD_HASH",
    "$2b$12$Tt41dNlGAIe8dGGt5ybUGu2OALT7E26IaBpIiQJtybKrtnNL5wv62"  # 默认: admin123
)

# ═════════════════════════════════════════════════════════
# 数据库层（双引擎：MySQL + SQLite）
# ═════════════════════════════════════════════════════════

# 价格计算 SQL 片段（实时从 textbooks_master 计算结算价）
PRICE_CALC = "COALESCE(tm.actual_price, tm.price*COALESCE(tm.discount_rate,1), t.price, 0)"
PRICE_JOIN = "LEFT JOIN textbooks_master tm ON tm.isbn = t.isbn"

def get_sqlite_path():
    return os.path.join(os.path.dirname(__file__), "textbook_data.db")

@contextmanager
def get_connection():
    """统一获取数据库连接"""
    if DB_TYPE == "sqlite":
        conn = sqlite3.connect(get_sqlite_path())
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            yield conn
        finally:
            conn.close()
    else:
        conn = pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)
        try:
            yield conn
        finally:
            conn.close()

def init_db():
    """初始化数据库表结构（兼容 MySQL 和 SQLite）"""
    if DB_TYPE == "sqlite":
        pass  # SQLite 直接连接，IF NOT EXISTS 自动处理建表/迁移
    else:
        try:
            tmp_conn = pymysql.connect(
                host=DB_CONFIG["host"], port=DB_CONFIG["port"],
                user=DB_CONFIG["user"], password=DB_CONFIG["password"],
                charset="utf8mb4"
            )
            tmp_cur = tmp_conn.cursor()
            tmp_cur.execute(f"CREATE DATABASE IF NOT EXISTS `{DB_CONFIG['database']}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
            tmp_conn.commit()
            tmp_cur.close()
            tmp_conn.close()
        except Exception:
            pass

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS semesters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(100) UNIQUE NOT NULL,
                academic_year VARCHAR(20),
                semester_name VARCHAR(50),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_card VARCHAR(18) UNIQUE NOT NULL,
                student_id VARCHAR(50) UNIQUE NOT NULL,
                name VARCHAR(50) NOT NULL,
                grade VARCHAR(20),
                college VARCHAR(100),
                major VARCHAR(100),
                class_name VARCHAR(100),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS textbooks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                semester_id INTEGER NOT NULL,
                grade VARCHAR(20),
                college VARCHAR(100),
                major VARCHAR(100),
                class_name VARCHAR(100),
                name VARCHAR(200) NOT NULL,
                publisher VARCHAR(200),
                editor VARCHAR(200),
                course_name VARCHAR(200),
                isbn VARCHAR(50),
                price DECIMAL(10,2) NOT NULL DEFAULT 0,
                quantity INTEGER DEFAULT 0,
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (semester_id) REFERENCES semesters(id) ON DELETE CASCADE
            )
        """)
        # 旧版textbooks表保留兼容，新版使用 textbooks_master + textbook_orders
        cur.execute("""
            CREATE TABLE IF NOT EXISTS textbooks_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(200) NOT NULL,
                isbn VARCHAR(50),
                publisher VARCHAR(200),
                editor VARCHAR(200),
                price DECIMAL(10,2) DEFAULT 0,
                course_name VARCHAR(200),
                publication_date VARCHAR(50),
                edition VARCHAR(100),
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS textbook_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                semester_id INTEGER NOT NULL,
                textbook_id INTEGER NOT NULL,
                grade VARCHAR(20),
                college VARCHAR(100),
                major VARCHAR(100),
                class_name VARCHAR(100),
                quantity INTEGER DEFAULT 0,
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (textbook_id) REFERENCES textbooks_master(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_exemptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                semester_id INTEGER NOT NULL,
                student_id INTEGER NOT NULL,
                is_exempt INTEGER NOT NULL DEFAULT 0,
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (semester_id) REFERENCES semesters(id) ON DELETE CASCADE,
                FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
                UNIQUE(semester_id, student_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS distributions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                textbook_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                distribute_date DATE,
                handler VARCHAR(50),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
                FOREIGN KEY (textbook_id) REFERENCES textbooks(id) ON DELETE CASCADE
            )
        """)
        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_students_grade ON students(grade)",
            "CREATE INDEX IF NOT EXISTS idx_students_college ON students(college)",
            "CREATE INDEX IF NOT EXISTS idx_students_major ON students(major)",
            "CREATE INDEX IF NOT EXISTS idx_students_class ON students(class_name)",
            "CREATE INDEX IF NOT EXISTS idx_students_id_card ON students(id_card)",
            "CREATE INDEX IF NOT EXISTS idx_students_student_id ON students(student_id)",
            "CREATE INDEX IF NOT EXISTS idx_textbooks_semester ON textbooks(semester_id)",
            "CREATE INDEX IF NOT EXISTS idx_textbooks_class ON textbooks(class_name)",
            "CREATE INDEX IF NOT EXISTS idx_textbooks_name ON textbooks(name)",
            "CREATE INDEX IF NOT EXISTS idx_exemptions_semester ON student_exemptions(semester_id)",
            "CREATE INDEX IF NOT EXISTS idx_exemptions_student ON student_exemptions(student_id)",
            "CREATE INDEX IF NOT EXISTS idx_distributions_student ON distributions(student_id)",
            "CREATE INDEX IF NOT EXISTS idx_distributions_textbook ON distributions(textbook_id)",
        ]
        for sql in indexes:
            try:
                cur.execute(sql)
            except Exception:
                pass
        conn.commit()

        # ── 数据库迁移：旧版 edition → editor ──
        try:
            cur.execute("SELECT editor FROM textbooks LIMIT 0")
        except Exception:
            try:
                cur.execute("ALTER TABLE textbooks ADD COLUMN editor VARCHAR(200)")
                conn.commit()
            except Exception:
                pass

        # ── 数据库迁移：旧版无 isbn 列（PRAGMA 方式更可靠）──
        try:
            cur.execute("PRAGMA table_info(textbooks)")
            tb_cols = {r[1] for r in cur.fetchall()}
            if "isbn" not in tb_cols:
                cur.execute("ALTER TABLE textbooks ADD COLUMN isbn VARCHAR(50)")
                conn.commit()
                print("Migration: added isbn column")
            if "course_name" not in tb_cols:
                cur.execute("ALTER TABLE textbooks ADD COLUMN course_name VARCHAR(200)")
                conn.commit()
                print("Migration: added course_name column")
            if "publication_date" not in tb_cols:
                cur.execute("ALTER TABLE textbooks ADD COLUMN publication_date VARCHAR(50)")
                conn.commit()
                print("Migration: added publication_date column")
        except Exception:
            pass

        # ── 数据库迁移：textbooks_master 缺少 publication_date ──
        try:
            cur.execute("PRAGMA table_info(textbooks_master)")
            master_cols = {r[1] for r in cur.fetchall()}
            if "publication_date" not in master_cols:
                cur.execute("ALTER TABLE textbooks_master ADD COLUMN publication_date VARCHAR(50)")
                conn.commit()
                print("Migration: added publication_date to textbooks_master")
            if "discount_rate" not in master_cols:
                cur.execute("ALTER TABLE textbooks_master ADD COLUMN discount_rate REAL DEFAULT NULL")
                conn.commit()
                print("Migration: added discount_rate to textbooks_master")
            if "actual_price" not in master_cols:
                cur.execute("ALTER TABLE textbooks_master ADD COLUMN actual_price REAL DEFAULT NULL")
                conn.commit()
                print("Migration: added actual_price to textbooks_master")
            if "is_national_standard" not in master_cols and "textbook_type" not in master_cols:
                cur.execute("ALTER TABLE textbooks_master ADD COLUMN textbook_type TEXT DEFAULT NULL")
                conn.commit()
                print("Migration: added textbook_type to textbooks_master")
            elif "is_national_standard" in master_cols and "textbook_type" not in master_cols:
                # 迁移旧字段：is_national_standard(0/1) → textbook_type TEXT
                cur.execute("ALTER TABLE textbooks_master ADD COLUMN textbook_type TEXT DEFAULT NULL")
                cur.execute("UPDATE textbooks_master SET textbook_type='国规教材' WHERE is_national_standard=1")
                conn.commit()
                print("Migration: migrated is_national_standard -> textbook_type")
            if "textbook_type" not in master_cols:
                # 再查一次确保新字段存在
                cur.execute("PRAGMA table_info(textbooks_master)")
                master_cols = {r[1] for r in cur.fetchall()}
        except Exception:
            pass

        # ── 数据库迁移：textbook_subscriptions 缺少 class_names ──
        try:
            cur.execute("PRAGMA table_info(textbook_subscriptions)")
            sub_cols = {r[1] for r in cur.fetchall()}
            if "class_names" not in sub_cols:
                cur.execute("ALTER TABLE textbook_subscriptions ADD COLUMN class_names TEXT")
                conn.commit()
                print("Migration: added class_names to textbook_subscriptions")
        except Exception:
            pass

        # ── 数据库迁移：旧版无 student_exemptions 表 ──
        try:
            cur.execute("SELECT COUNT(*) FROM student_exemptions")
        except Exception:
            cur.execute("""
                CREATE TABLE student_exemptions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semester_id INTEGER NOT NULL,
                    student_id INTEGER NOT NULL,
                    is_exempt INTEGER NOT NULL DEFAULT 0,
                    remark TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (semester_id) REFERENCES semesters(id) ON DELETE CASCADE,
                    FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
                    UNIQUE(semester_id, student_id)
                )
            """)
            conn.commit()

        # ── 数据库迁移：新增 textbook_subscriptions 征订总表 ──
        try:
            cur.execute("SELECT COUNT(*) FROM textbook_subscriptions")
        except Exception:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS textbook_subscriptions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    semester_id INTEGER NOT NULL,
                    textbook_id INTEGER,
                    book_name VARCHAR(200) NOT NULL,
                    isbn VARCHAR(50),
                    publisher VARCHAR(200),
                    editor VARCHAR(200),
                    price DECIMAL(10,2) DEFAULT 0,
                    course_name VARCHAR(200),
                    college VARCHAR(100),
                    major VARCHAR(100),
                    grade VARCHAR(20),
                    class_scope TEXT,
                    class_names TEXT,
                    total_qty INTEGER DEFAULT 0,
                    teacher_qty INTEGER DEFAULT 0,
                    status VARCHAR(20) DEFAULT 'pending',
                    remark TEXT,
                    source VARCHAR(50) DEFAULT 'manual',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    dispatched_at DATETIME,
                    FOREIGN KEY (semester_id) REFERENCES semesters(id) ON DELETE CASCADE
                )
            """)
            conn.commit()

# ═════════════════════════════════════════════════════════
# 工具函数
# ═════════════════════════════════════════════════════════

def get_current_academic_info() -> tuple:
    """自动识别当前学年和学期"""
    now = datetime.now()
    year, month = now.year, now.month
    if month >= 9:
        return f"{year}-{year+1}", "第一学期"
    elif month <= 2:
        return f"{year-1}-{year}", "第一学期"
    else:
        return f"{year-1}-{year}", "第二学期"

def query_df(sql: str, params: tuple = None) -> pd.DataFrame:
    """执行查询返回 DataFrame"""
    if DB_TYPE == "sqlite":
        sql = sql.replace("%s", "?")
    with get_connection() as conn:
        return pd.read_sql(sql, conn, params=params)

def execute_sql(sql: str, params: tuple = None) -> int:
    """执行写操作，返回 lastrowid"""
    if DB_TYPE == "sqlite":
        sql = sql.replace("%s", "?")
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(sql, params or ())
        conn.commit()
        return cur.lastrowid

def excel_export(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """DataFrame -> Excel 字节流"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()

def apply_excel_borders(ws, min_row, max_row, min_col, max_col):
    """给指定区域添加细边框"""
    from openpyxl.styles import Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

def excel_export_by_class(df: pd.DataFrame, class_col: str = "班级", file_prefix: str = "导出") -> bytes:
    """
    按班级分 sheet 导出 Excel。
    返回 bytes
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if class_col in df.columns:
            classes = sorted(df[class_col].dropna().unique())
            if len(classes) <= 1:
                df.to_excel(writer, index=False, sheet_name="全部班级")
            else:
                for cls in classes:
                    cls_df = df[df[class_col] == cls]
                    sheet_name = str(cls)[:31]
                    cls_df.to_excel(writer, index=False, sheet_name=sheet_name)
        else:
            df.to_excel(writer, index=False, sheet_name="数据")
    return output.getvalue()

def make_template_df(columns: list) -> pd.DataFrame:
    """生成空模板 DataFrame"""
    return pd.DataFrame(columns=columns)

MAX_UPLOAD_MB = 5

def read_excel_upload(uploaded_file) -> pd.DataFrame:
    """读取上传的 Excel，带安全校验"""
    size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        raise ValueError(f"文件过大（{size_mb:.1f}MB），上限 {MAX_UPLOAD_MB}MB")
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, engine="openpyxl")
    if df.empty:
        raise ValueError("文件为空或无有效数据")
    if len(df) > 5000:
        raise ValueError(f"行数过多（{len(df)}），上限 5000 行")
    return df

def get_filtered_list(table: str, column: str, where: str = "1=1", params: tuple = ()) -> list:
    """通用不重复值查询"""
    df = query_df(f"SELECT DISTINCT {column} FROM {table} WHERE {column} IS NOT NULL AND {column} != '' AND {where} ORDER BY {column}", params)
    return df[column].tolist() if not df.empty else []

def get_filtered_class_names():
    return get_filtered_list("students", "class_name")

def get_filtered_grades():
    return get_filtered_list("students", "grade")

def get_filtered_majors():
    return get_filtered_list("students", "major")

def get_filtered_colleges():
    return get_filtered_list("students", "college")

def normalize_grade(g: str) -> str:
    """统一年级格式为 \"202X级\"，兼容 \"202X\" / \"202X级\" 两种写法"""
    if not g:
        return g
    g = str(g).strip()
    if not g.endswith("级"):
        g = g + "级"
    return g

def get_class_student_counts(grade=None, college=None, major=None, class_names=None) -> pd.DataFrame:
    """查询指定条件下各班级的实际学生人数分布"""
    where = "1=1"; params = []
    if grade:
        # 年级格式兼容：学生表 grade 可能是 "2023" 或 "2023级"，统一比较
        g_raw = str(grade).rstrip("级")
        where += " AND (grade = %s OR grade = %s)"
        params.extend([g_raw + "级", g_raw])
    if college: where += " AND college = %s"; params.append(college)
    if major: where += " AND major = %s"; params.append(major)
    if class_names:
        ph = ",".join(["%s"] * len(class_names))
        where += f" AND class_name IN ({ph})"; params.extend(class_names)
    return query_df(
        f"SELECT class_name, COUNT(*) as student_count FROM students WHERE {where} GROUP BY class_name ORDER BY class_name",
        tuple(params)
    )

def split_qty_by_class(total_qty: int, grade=None, college=None, major=None, class_names=None):
    """将总数量按各班级实际人数比例分摊到每个班
    
    返回: [(class_name, student_count, allocated_qty), ...]
    如无法获取班级分布（无对应 students 数据），返回 [(None, 0, total_qty)]
    """
    df = get_class_student_counts(grade, college, major, class_names)
    if df.empty:
        return [(None, 0, total_qty)]
    total_students = df["student_count"].sum()
    result = []
    for _, row in df.iterrows():
        cn = row["class_name"]
        cnt = int(row["student_count"])
        # 按人数比例分摊，确保整数且总和 = total_qty
        alloc = round(total_qty * cnt / total_students)
        result.append((cn, cnt, alloc))
    # 四舍五入可能导致总和偏差，校正到第一位
    diff = total_qty - sum(r[2] for r in result)
    if diff != 0 and result:
        # 把差额补给数量最大的班
        max_idx = max(range(len(result)), key=lambda i: result[i][2])
        result[max_idx] = (result[max_idx][0], result[max_idx][1], result[max_idx][2] + diff)
    return result

def parse_major_grade_from_scope(class_scope: str, college: str = ""):
    """从班级范围说明中智能解析专业和年级

    支持的格式：
    - "会计1241-1244" → 从班级号解析年级（1241→24→2024级），查学生表匹配完整专业名
    - "机械25级"   → major="机械", grades=["2025级"]
    - "25市场营销级" → major="市场营销", grades=["2025级"]
    - "24级、25级学生" → grades=["2024级","2025级"]（多年级拆分为多条记录）
    - "教师用书"    → ("", [])
    返回: (major_full, [grades]) — grades 始终为列表
    """
    if not class_scope:
        return "", []
    scope = str(class_scope).strip()
    if "教师" in scope:
        return "", []

    major_short = ""
    grades = []

    # ── 优先级1：从「专业名+班级号」模式提取年级 ──
    # 如 "会计1241-1244", "风电1251-1255", "会计1241-会计1251"（跨年级）
    # 班级编号规则: XYYZ, 其中 YY = 2位年份 (23→2023级, 24→2024级, 25→2025级)
    class_num_m = re.findall(r'(\d{4})', scope)
    if class_num_m:
        years_from_class = set()
        for cn in class_num_m:
            if len(cn) == 4:
                yy = cn[1:3]  # 第2-3位是年份
                if yy.isdigit():
                    years_from_class.add(int(yy))
        if years_from_class:
            grades = sorted([f"20{yy}级" for yy in years_from_class])

        # 提取中文专业名（在第一个数字之前）
        mm = re.match(r'^([\u4e00-\u9fff]+)', scope)
        if mm:
            major_short = mm.group(1)
        return _resolve_major_name(major_short, college), grades

    # ── 优先级2: 匹配「XX级」模式（支持多个"XX级"）──
    # "机械25级" → ["2025级"]; "24级、25级学生" → ["2024级","2025级"]
    yy_matches = re.findall(r'(\d{2})\s*级', scope)
    if yy_matches:
        grades = sorted([f"20{yy}级" for yy in yy_matches])
        remaining = re.sub(r'\d{2}\s*级', '', scope).strip()
        mm = re.match(r'^([\u4e00-\u9fff]+)', remaining)
        if mm:
            major_short = mm.group(1)
        return _resolve_major_name(major_short, college), grades

    # ── 优先级3: 匹配「开头数字+中文+级」(如 25市场营销级) ──
    m = re.match(r'^(\d{2})\s*([\u4e00-\u9fff]+)\s*级', scope)
    if m:
        grades = ["20" + m.group(1) + "级"]
        major_short = m.group(2)
        return _resolve_major_name(major_short, college), grades

    # ── 兜底: 提取开头中文，年级从学生表查 ──
    mm = re.match(r'^([\u4e00-\u9fff]+)', scope)
    if mm:
        major_short = mm.group(1)
    return _resolve_major_name(major_short, college), grades


def _resolve_major_name(major_short: str, college: str = ""):
    """通过学生表将短专业名解析为完整专业名"""
    if not major_short:
        return ""

    conditions = ["major LIKE %s"]
    params = [f"%{major_short}%"]

    if college:
        conditions.append("college = %s")
        params.append(college)

    where = " AND ".join(conditions)
    students_df = query_df(
        f"SELECT DISTINCT major FROM students WHERE {where} LIMIT 3",
        tuple(params)
    )

    if not students_df.empty:
        return str(students_df.iloc[0]["major"])

    # 如果 major 模糊匹配没找到，尝试通过班级名反查（班级名用短名，如「风电1231」）
    conds2 = ["class_name LIKE %s"]
    params2 = [f"{major_short}%"]
    if college:
        conds2.append("college = %s")
        params2.append(college)
    students_df = query_df(
        f"SELECT DISTINCT major FROM students WHERE {' AND '.join(conds2)} LIMIT 3",
        tuple(params2)
    )
    if not students_df.empty:
        return str(students_df.iloc[0]["major"])

    return major_short


def safe_int(v, default=0):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default

def safe_float(v, default=0.0):
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def safe_str(v, default=""):
    """避免 NaN/None 被写成 'nan'/'None'；日期对象转为纯日期字符串"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    # 处理 pandas Timestamp / Python datetime → 仅保留日期部分
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # 去掉时间部分（如 "2024-04-01 00:00:00" → "2024-04-01"）
    if " 00:00:00" in s:
        s = s.replace(" 00:00:00", "")
    return s

def safe_field(v, default=""):
    """将字段值转为字符串，处理数字（2023.0→2023）、NaN、None"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

# ── 导入日志 ──
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")

def write_import_log(module: str, filename: str, total: int, success: int, errors: list):
    """记录导入操作到日志文件"""
    os.makedirs(LOG_DIR, exist_ok=True)
    log_file = os.path.join(LOG_DIR, "import.log")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    error_detail = "; ".join(errors[:20]) if errors else "无"
    line = (
        f"[{now}]\n"
        f"  模块: {module}\n"
        f"  文件: {filename}\n"
        f"  总计: {total} 条 | 成功: {success} 条 | 失败: {len(errors)} 条\n"
        f"  错误: {error_detail}\n"
        f"  {'─' * 60}\n"
    )
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line)

def read_import_logs(n=50):
    """读取最近的导入日志"""
    log_file = os.path.join(LOG_DIR, "import.log")
    if not os.path.exists(log_file):
        return ["暂无日志记录"]
    with open(log_file, "r", encoding="utf-8") as f:
        lines = f.readlines()
    # 按时间倒序返回最近 n 条（每条约6行）
    blocks = []
    block = []
    for line in reversed(lines):
        if line.startswith("["):
            if block:
                blocks.append("".join(reversed(block)))
                block = []
        block.append(line)
    if block:
        blocks.append("".join(reversed(block)))
    return blocks[:n]

# ═════════════════════════════════════════════════════════
# 生成测试数据
# ═════════════════════════════════════════════════════════

def generate_test_data():
    """
    生成完整的测试数据：
    - 3个学期（2024-2025 两个学期 + 2025-2026 第二学期）
    - 12名学生，分3个班级
    - 每个学期各班级若干教材
    - 每个学期都有发放记录
    """
    # 1. 创建学期
    semesters_data = [
        ("2024-2025 第一学期", "2024-2025", "第一学期"),
        ("2024-2025 第二学期", "2024-2025", "第二学期"),
        ("2025-2026 第二学期", "2025-2026", "第二学期"),
    ]
    sem_ids = {}
    for name, ay, sn in semesters_data:
        try:
            sid = execute_sql(
                "INSERT INTO semesters (name, academic_year, semester_name) VALUES (%s, %s, %s)",
                (name, ay, sn)
            )
            sem_ids[name] = sid
        except Exception:
            # 已存在则查询 id
            df = query_df("SELECT id FROM semesters WHERE name = %s", (name,))
            if not df.empty:
                sem_ids[name] = int(df.iloc[0]["id"])

    # 2. 创建学生（3个班级，各4人）
    students_data = [
        # 软件工程1班
        ("110101200501010001", "2024001", "张三",   "2024级", "计算机学院", "软件工程", "软件工程1班"),
        ("110101200501010002", "2024002", "李四",   "2024级", "计算机学院", "软件工程", "软件工程1班"),
        ("110101200501010003", "2024003", "王五",   "2024级", "计算机学院", "软件工程", "软件工程1班"),
        ("110101200501010004", "2024004", "赵六",   "2024级", "计算机学院", "软件工程", "软件工程1班"),
        # 计算机科学1班
        ("110101200502010001", "2024005", "钱七",   "2024级", "计算机学院", "计算机科学", "计算机科学1班"),
        ("110101200502010002", "2024006", "孙八",   "2024级", "计算机学院", "计算机科学", "计算机科学1班"),
        ("110101200502010003", "2024007", "周九",   "2024级", "计算机学院", "计算机科学", "计算机科学1班"),
        ("110101200502010004", "2024008", "吴十",   "2024级", "计算机学院", "计算机科学", "计算机科学1班"),
        # 应用数学1班
        ("110101200503010001", "2024009", "郑十一", "2024级", "数学学院",   "应用数学",   "应用数学1班"),
        ("110101200503010002", "2024010", "王十二", "2024级", "数学学院",   "应用数学",   "应用数学1班"),
        ("110101200503010003", "2024011", "李十三", "2024级", "数学学院",   "应用数学",   "应用数学1班"),
        ("110101200503010004", "2024012", "赵十四", "2024级", "数学学院",   "应用数学",   "应用数学1班"),
    ]
    for id_card, sid, name, grade, college, major, cls in students_data:
        try:
            execute_sql(
                """INSERT INTO students (id_card,student_id,name,grade,college,major,class_name)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (id_card, sid, name, grade, college, major, cls)
        )
        except Exception:
            pass  # 已存在则跳过

    # 3. 创建教材（每个学期，每个班级2-3本教材）
    # 先获取学期ID
    sem_df = query_df("SELECT id, name FROM semesters ORDER BY id")
    sem_id_map = {row["name"]: int(row["id"]) for _, row in sem_df.iterrows()}

    textbooks_data = [
        # 2024-2025 第一学期（软件工程1班4人：征订量有的多有少）
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "Python程序设计",       "清华大学出版社", "张三丰", "978-7-302-00001-1", 59.00, 5, ""),   # 订5本→够发
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "数据结构与算法",     "清华大学出版社", "李思",   "978-7-302-00002-8", 45.00, 3, ""),   # 订3本→不够！缺口1
        # 2024-2025 第一学期（计算机科学1班4人）
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "计算机学院", "计算机科学", "计算机科学1班", "计算机网络",      "电子工业出版社", "王武",   "978-7-121-00003-5", 42.00, 4, ""),   # 订4本→刚好
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "计算机学院", "计算机科学", "计算机科学1班", "操作系统原理",    "机械工业出版社", "赵柳",   "978-7-111-00004-2", 55.00, 6, ""),   # 订6本→有富余
        # 2024-2025 第一学期（应用数学1班4人）
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "数学学院",   "应用数学",   "应用数学1班",   "高等数学（上）",  "高等教育出版社", "陈七",   "978-7-04-00005-9", 38.00, 4, ""),
        (sem_id_map.get("2024-2025 第一学期"), "2024级", "数学学院",   "应用数学",   "应用数学1班",   "线性代数",       "高等教育出版社", "刘八",   "978-7-04-00006-6", 32.00, 2, ""),   # 订2本→不够
        # 2024-2025 第二学期
        (sem_id_map.get("2024-2025 第二学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "Python程序设计（下）", "清华大学出版社", "张三丰", "978-7-302-00007-3", 62.00, 4, ""),
        (sem_id_map.get("2024-2025 第二学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "数据库系统概论",    "高等教育出版社", "李思",   "978-7-04-00008-0", 48.00, 4, ""),
        (sem_id_map.get("2024-2025 第二学期"), "2024级", "计算机学院", "计算机科学", "计算机科学1班", "计算机组成原理",   "清华大学出版社", "王武",   "978-7-302-00009-7", 52.00, 4, ""),
        (sem_id_map.get("2024-2025 第二学期"), "2024级", "数学学院",   "应用数学",   "应用数学1班",   "高等数学（下）",   "高等教育出版社", "陈七",   "978-7-04-00010-3", 41.00, 4, ""),
        (sem_id_map.get("2024-2025 第二学期"), "2024级", "数学学院",   "应用数学",   "应用数学1班",   "概率论与数理统计", "高等教育出版社", "刘八",   "978-7-04-00011-0", 35.00, 4, ""),
        # 2025-2026 第二学期
        (sem_id_map.get("2025-2026 第二学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "软件工程导论",     "清华大学出版社", "周九",   "978-7-302-00012-7", 50.00, 4, ""),
        (sem_id_map.get("2025-2026 第二学期"), "2024级", "计算机学院", "软件工程",   "软件工程1班", "Web前端开发",       "电子工业出版社", "吴十",   "978-7-121-00013-4", 46.00, 4, ""),
        (sem_id_map.get("2025-2026 第二学期"), "2024级", "计算机学院", "计算机科学", "计算机科学1班", "编译原理",         "机械工业出版社", "赵柳",   "978-7-111-00014-1", 54.00, 4, ""),
        (sem_id_map.get("2025-2026 第二学期"), "2024级", "计算机学院", "计算机科学", "计算机科学1班", "人工智能导论",     "清华大学出版社", "李思",   "978-7-302-00015-8", 49.00, 4, ""),
        (sem_id_map.get("2025-2026 第二学期"), "2024级", "数学学院",   "应用数学",   "应用数学1班",   "数学建模",         "高等教育出版社", "陈七",   "978-7-04-00016-5", 37.00, 4, ""),
    ]
    for t in textbooks_data:
        try:
            execute_sql(
                """INSERT INTO textbooks (semester_id,grade,college,major,class_name,name,course_name,publisher,editor,isbn,price,quantity,remark)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                t
            )
        except Exception:
            pass

    # 4. 创建发放记录（每个学生每个学期领取对应班级的教材各1本）
    # 获取学生和教材映射
    all_students = query_df("SELECT id, student_id, class_name FROM students ORDER BY class_name, student_id")
    all_textbooks = query_df("SELECT id, semester_id, class_name FROM textbooks ORDER BY semester_id, class_name")

    # 构建：semester_id -> class_name -> [textbook_id, ...]
    sem_class_books = {}
    for _, t in all_textbooks.iterrows():
        sid = int(t["semester_id"])
        cls = t["class_name"]
        sem_class_books.setdefault(sid, {}).setdefault(cls, []).append(int(t["id"]))

    # 构建：class_name -> [student_id, ...]
    class_students = {}
    for _, s in all_students.iterrows():
        cls = s["class_name"]
        class_students.setdefault(cls, []).append((int(s["id"]), s["student_id"]))

    # 每个学生在每个学期领取本班所有教材各1本
    distribute_dates = {
        "2024-2025 第一学期": "2024-09-10",
        "2024-2025 第二学期": "2025-02-20",
        "2025-2026 第二学期": "2026-02-25",
    }
    for sid, class_map in sem_class_books.items():
        sem_name = query_df("SELECT name FROM semesters WHERE id = %s", (sid,))["name"].iloc[0]
        d_date = distribute_dates.get(sem_name, "2026-05-01")
        for cls, book_ids in class_map.items():
            for stu_id, stu_sid in class_students.get(cls, []):
                for bid in book_ids:
                    try:
                        execute_sql(
                            "INSERT INTO distributions (student_id,textbook_id,quantity,distribute_date,handler) VALUES (%s,%s,1,%s,'管理员')",
                            (stu_id, bid, d_date)
                        )
                    except Exception:
                        pass

    return len(students_data), len(textbooks_data), len(sem_class_books)

# ═════════════════════════════════════════════════════════
# 共享组件
# ═════════════════════════════════════════════════════════

def show_header(title, subtitle=None):
    """统一页面头部"""
    html = f'<div class="app-header"><h1>{title}</h1>'
    if subtitle:
        html += f'<p>{subtitle}</p>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

def styled_metric(label, value, delta=None, color="#1e40af"):
    """自定义美化 metric"""
    delta_html = f'<span style="color:#059669;font-size:14px;">↑ {delta}</span>' if delta else ''
    st.markdown(f"""
    <div class="info-card" style="text-align:center;">
        <div class="stat-number" style="color:{color};">{value}</div>
        <div class="stat-label">{label}</div>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════
# 登录页面
# ═════════════════════════════════════════════════════════

def login_page():
    st.markdown("""
    <div style="text-align:center; padding: 60px 0 20px 0;">
        <div style="font-size:64px; margin-bottom:12px;">📚</div>
        <h1 style="color:#1e40af; font-size:36px; font-weight:800; margin:0;">学生教材费用核对系统</h1>
        <p style="color:#6b7280; font-size:16px; margin-top:8px;">Textbook Fee Management System</p>
    </div>
    """, unsafe_allow_html=True)

    ay, sem = get_current_academic_info()
    st.markdown(f"""
    <div style="text-align:center; margin-bottom:32px;">
        <span style="background:#eff6ff; color:#1e40af; padding:6px 20px; border-radius:20px; font-weight:600;">
            📅 当前：{ay} 学年 {sem}
        </span>
    </div>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 1.5, 1])
    with col_c:
        tab1, tab2 = st.tabs(["🔑 管理员登录", "🎓 学生查询"])

        with tab1:
            with st.form("admin_login", clear_on_submit=False):
                st.markdown("### 管理员入口")
                pwd = st.text_input("管理员密码", type="password", placeholder="请输入管理员密码")
                submitted = st.form_submit_button("🔓 登录系统", use_container_width=True)
                if submitted:
                    if not ADMIN_PASSWORD_HASH:
                        st.error("⚠️ 未配置管理员密码（ADMIN_PASSWORD_HASH 环境变量）")
                    elif bcrypt.checkpw(pwd.encode(), ADMIN_PASSWORD_HASH.encode()):
                        st.session_state.role = "admin"
                        st.session_state.user = "管理员"
                        st.session_state.page = "semester"
                        st.rerun()
                    else:
                        st.error("❌ 密码错误，请重试")

        with tab2:
            with st.form("student_login", clear_on_submit=False):
                st.markdown("### 学生查询入口")
                login_id = st.text_input("请输入学号或身份证号", placeholder="学号 或 身份证号")
                submitted = st.form_submit_button("🔍 查询费用", use_container_width=True)
                if submitted:
                    if login_id.strip():
                        df = query_df(
                            "SELECT * FROM students WHERE student_id = %s OR id_card = %s",
                            (login_id.strip(), login_id.strip())
                        )
                        if not df.empty:
                            student = df.iloc[0].to_dict()
                            st.session_state.role = "student"
                            st.session_state.user = student
                            st.session_state.page = "student_query"
                            st.rerun()
                        else:
                            st.error("❌ 未找到该学生信息，请检查学号或身份证号")
                    else:
                        st.warning("请输入学号或身份证号")

    st.markdown("""
    <div style="text-align:center; color:#9ca3af; font-size:12px; margin-top:48px;">
        v2.1 &nbsp;|&nbsp; 支持 MySQL / SQLite &nbsp;|&nbsp; 按班级分页打印
    </div>
    """, unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════
# 管理员侧边栏
# ═════════════════════════════════════════════════════════

def admin_sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="text-align:center; padding:16px 0 8px 0;">
            <div style="font-size:40px;">📚</div>
            <h3 style="color:#1e40af; margin:8px 0 0 0; font-weight:700;">教材费用系统</h3>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="user-badge">👤 管理员模式</div>', unsafe_allow_html=True)

        ay, sem = get_current_academic_info()
        st.caption(f"🏫 {ay} {sem}")

        st.divider()

        pages = [
            ("📅", "学期管理", "semester", "管理学年学期"),
            ("👨‍🎓", "学生管理", "students", "导入/编辑学生信息"),
            ("📖", "教材表管理", "textbook_master", "管理教材库（名称/ISBN/出版社/主编/单价/课程）"),
            ("📋", "征订总表", "subscriptions", "导入原始征订数据，一键下发到班级"),
            ("📖", "教材征订表", "textbooks", "按班级查看/编辑征订明细"),
            ("📦", "教材发放表", "distribution", "录入发放记录"),
            ("✅", "领书确认表", "confirmation", "学生领书确认（免领标记）"),
            ("📊", "费用统计", "statistics", "多维度费用汇总"),
            ("📋", "系统日志", "logs", "查看导入/操作记录"),
        ]

        for emoji, label, key, desc in pages:
            active = st.session_state.get("page") == key
            btn_type = "primary" if active else "secondary"
            if st.button(f"{emoji} {label}", use_container_width=True, type=btn_type,
                         help=desc, key=f"nav_{key}"):
                st.session_state.page = key
                st.rerun()

        st.divider()
        if st.button("🚪 退出登录", use_container_width=True):
            for k in ["role", "user", "page"]:
                st.session_state.pop(k, None)
            st.rerun()

def student_sidebar():
    with st.sidebar:
        student = st.session_state.user
        st.markdown(f"""
        <div style="text-align:center; padding:16px 0 8px 0;">
            <div style="font-size:48px;">🎓</div>
            <h3 style="color:#1e40af; margin:8px 0 0 0;">费用查询</h3>
        </div>
        """, unsafe_allow_html=True)

        st.info(f"**{student['name']}**\n学号：{student['student_id']}\n班级：{student.get('class_name', '-')}")

        st.divider()
        if st.button("🚪 退出登录", use_container_width=True):
            for k in ["role", "user", "page"]:
                st.session_state.pop(k, None)
            st.rerun()

# ═════════════════════════════════════════════════════════
# 1. 学期管理
# ═════════════════════════════════════════════════════════

def semester_management():
    show_header("📅 学期管理", "管理学年和学期，系统会自动识别当前学期")

    ay, sem = get_current_academic_info()

    col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
    with col1:
        st.info(f"🔍 系统自动识别当前：**{ay} 学年 {sem}**")
    with col2:
        if st.button("➕ 新增学期", use_container_width=True, type="primary"):
            st.session_state.show_semester_form = True
    with col3:
        if st.button("🎲 生成测试数据", use_container_width=True, type="secondary"):
            st.session_state.show_gen_confirm = True
    with col4:
        if st.button("🔄 刷新", use_container_width=True):
            st.rerun()

    # 生成测试数据确认
    if st.session_state.get("show_gen_confirm"):
        st.warning("⚠️ 将生成完整的测试数据（学期、学生、教材、发放记录），已有同名数据会自动跳过。确定要生成吗？")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("✅ 确认生成", use_container_width=True, type="primary"):
                with st.spinner("正在生成测试数据..."):
                    n_stu, n_tb, n_sem = generate_test_data()
                st.success(f"✅ 测试数据生成完成！学生 {n_stu} 条，教材 {n_tb} 条，涉及 {n_sem} 个学期")
                st.session_state.show_gen_confirm = False
                st.rerun()
        with c2:
            if st.button("取消", use_container_width=True):
                st.session_state.show_gen_confirm = False
                st.rerun()

    if st.session_state.get("show_semester_form"):
        with st.form("semester_form_inner"):
            st.markdown("#### ✨ 新增学期")
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                name = st.text_input("学期名称", placeholder=f"如：{ay} {sem}")
            with c2:
                ac_year = st.text_input("学年", value=ay)
            with c3:
                sem_name = st.selectbox("学期", ["第一学期", "第二学期"],
                                        index=1 if sem == "第二学期" else 0)

            col_a, col_b = st.columns(2)
            with col_a:
                save_btn = st.form_submit_button("💾 保存", use_container_width=True)
            with col_b:
                cancel_btn = st.form_submit_button("取消", use_container_width=True)

            if save_btn:
                full_name = name or f"{ac_year} {sem_name}"
                try:
                    execute_sql(
                        "INSERT INTO semesters (name, academic_year, semester_name) VALUES (%s, %s, %s)",
                        (full_name, ac_year, sem_name)
                    )
                    st.success(f"✅ 已添加学期：{full_name}")
                    st.session_state.show_semester_form = False
                    st.rerun()
                except Exception as e:
                    if "UNIQUE" in str(e) or "unique" in str(e).lower():
                        st.error("❌ 该学期已存在")
                    else:
                        st.error(f"保存失败：{e}")
            if cancel_btn:
                st.session_state.show_semester_form = False
                st.rerun()

    df = query_df("SELECT * FROM semesters ORDER BY id DESC")
    if not df.empty:
        st.markdown("#### 📋 已有学期")
        for _, row in df.iterrows():
            with st.container():
                col1, col2, col3, col4 = st.columns([3, 1, 1, 0.8])
                with col1:
                    st.write(f"**{row['name']}**")
                with col2:
                    st.caption(f"📅 {row.get('academic_year', '-')}")
                with col3:
                    st.caption(f"📝 {row.get('semester_name', '-')}")
                with col4:
                    if st.button("🗑️", key=f"del_sem_{row['id']}", help="删除此学期"):
                        execute_sql("DELETE FROM semesters WHERE id = %s", (row['id'],))
                        st.warning(f"已删除学期「{row['name']}」")
                        st.rerun()
        st.divider()
    else:
        st.warning("⚠️ 暂无学期数据，请先点击「新增学期」添加")

# ═════════════════════════════════════════════════════════
# 2. 学生管理
# ═════════════════════════════════════════════════════════

def student_management():
    show_header("👨‍🎓 学生管理", "管理学生基础信息，支持 Excel 批量导入导出")

    tab1, tab2, tab3 = st.tabs(["📋 学生列表", "➕ 新增/编辑", "📥 导入 Excel"])

    # ── Tab 1: 列表 ──
    with tab1:
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            f_grade = st.selectbox("年级", ["全部"] + get_filtered_grades(), key="s_grade")
        with col2:
            # 学院：基于所选年级级联过滤
            if f_grade != "全部":
                college_filter = get_filtered_list("students", "college", "grade = %s", (f_grade,))
            else:
                college_filter = get_filtered_colleges()
            f_college = st.selectbox("学院", ["全部"] + college_filter, key="s_college")
        with col3:
            # 专业：基于所选年级+学院级联过滤
            s_major_where = "1=1"; s_major_params = []
            if f_grade != "全部":
                s_major_where += " AND grade = %s"; s_major_params.append(f_grade)
            if f_college != "全部":
                s_major_where += " AND college = %s"; s_major_params.append(f_college)
            major_options = get_filtered_list("students", "major", s_major_where, tuple(s_major_params)) if s_major_params else get_filtered_majors()
            f_major = st.selectbox("专业", ["全部"] + major_options, key="s_major")
        with col4:
            # 班级：基于所选年级+学院+专业级联过滤
            f_where = "1=1"
            f_params = []
            if f_grade != "全部":
                f_where += " AND grade = %s"; f_params.append(f_grade)
            if f_college != "全部":
                f_where += " AND college = %s"; f_params.append(f_college)
            if f_major != "全部":
                f_where += " AND major = %s"; f_params.append(f_major)
            class_options = get_filtered_list("students", "class_name", f_where, tuple(f_params))
            f_class = st.selectbox("班级", ["全部"] + class_options, key="s_class")
        with col5:
            search = st.text_input("🔍 搜索", key="s_search", placeholder="姓名/学号/身份证...")

        sql = "SELECT id, id_card, student_id, name, grade, college, major, class_name FROM students WHERE 1=1"
        params = []
        for val, col in [(f_grade, "grade"), (f_college, "college"), (f_major, "major"), (f_class, "class_name")]:
            if val != "全部":
                sql += f" AND {col} = %s"; params.append(val)
        if search:
            sql += " AND (name LIKE %s OR student_id LIKE %s OR id_card LIKE %s)"
            like = f"%{search}%"; params.extend([like, like, like])
        sql += " ORDER BY grade, class_name, name"

        df = query_df(sql, tuple(params) if params else None)
        total = len(df)

        # ── 分页（从 session_state 读取，控件移到表格下方）──
        import math
        page_size = st.session_state.get("s_ps", 50)
        page = st.session_state.get("s_pg", 1)
        total_pages = max(1, math.ceil(total / page_size))
        # 页码保护：删除后当前页可能越界，自动跳到最后页
        if page > total_pages:
            page = total_pages
            st.session_state.s_pg = total_pages

        # 导出按钮保持在顶部右侧
        _, exp_col = st.columns([4, 1])
        with exp_col:
            excel_data = excel_export_by_class(
                df.rename(columns={
                    "id_card": "身份证号", "student_id": "学号", "name": "姓名",
                    "grade": "年级", "college": "学院", "major": "专业", "class_name": "班级"
                }),
                class_col="班级", file_prefix="学生名单"
            ) if not df.empty else b""
            st.download_button("📥 导出（按班级分页）", data=excel_data,
                               file_name=f"学生名单_{date.today()}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, disabled=df.empty)

        st.caption(f"共 **{total}** 名学生 ｜ 第 {page}/{total_pages} 页")

        if not df.empty:
            # 当前页数据
            start_idx = (page - 1) * page_size
            end_idx = min(start_idx + page_size, total)
            page_df = df.iloc[start_idx:end_idx].copy()

            # 保存原始数据用于对比
            edit_cols = ["name", "id_card", "grade", "college", "major", "class_name"]
            orig_key = "_orig"
            page_df[orig_key] = page_df[edit_cols].to_dict("records")

            # 全选状态：如果之前勾选了全选，默认全勾上
            select_all_init = st.session_state.get(f"s_selall_{page}", False)
            page_df["选择"] = select_all_init
            display_df = page_df.rename(columns={
                "id_card": "身份证号", "student_id": "学号", "name": "姓名",
                "grade": "年级", "college": "学院", "major": "专业", "class_name": "班级"
            })

            edited = st.data_editor(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_order=["选择", "学号", "姓名", "身份证号", "年级", "学院", "专业", "班级"],
                column_config={
                    "选择": st.column_config.CheckboxColumn("选择"),
                    "学号": st.column_config.TextColumn("学号", disabled=True, alignment="center"),
                    "姓名": st.column_config.TextColumn("姓名", alignment="center"),
                    "身份证号": st.column_config.TextColumn("身份证号", alignment="center"),
                    "年级": st.column_config.TextColumn("年级", alignment="center"),
                    "学院": st.column_config.TextColumn("学院", alignment="center"),
                    "专业": st.column_config.TextColumn("专业", alignment="center"),
                    "班级": st.column_config.TextColumn("班级", alignment="center"),
                },
                disabled=["学号"],
                key=f"stu_editor_{page}"
            )

            # ═══ 布局：上行(全选+分页) + 下行(红删蓝存) ═══
            # 检测修改
            changes = []
            for i in range(len(edited)):
                orig = page_df.iloc[i][orig_key]
                for col in edit_cols:
                    cn = {"name":"姓名","id_card":"身份证号","grade":"年级","college":"学院","major":"专业","class_name":"班级"}[col]
                    new_val = edited.iloc[i].get(cn)
                    old_val = orig.get(col)
                    if str(new_val or "") != str(old_val or ""):
                        changes.append((page_df.iloc[i]["id"], col, new_val))

            # 处理按钮触发的操作（用 session_state 代替 query_params，避免 form 提交导致会话丢失）
            pending = st.session_state.get("pending_action", "")
            if pending == "s_save":
                if changes:
                    for sid, col, val in changes:
                        execute_sql(f"UPDATE students SET {col}=%s WHERE id=%s", (val, sid))
                    st.toast(f"✅ 已保存 {len(changes)} 处修改", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()
            elif pending == "s_del":
                select_all = st.session_state.get(f"s_selall_{page}", False)
                if select_all:
                    del_ids = tuple(page_df["id"].tolist())
                else:
                    selected = edited[edited["选择"] == True].index if "选择" in edited.columns else []
                    del_ids = tuple(page_df.iloc[selected]["id"].tolist()) if len(selected) > 0 else ()
                if del_ids:
                    if len(del_ids) == 1:
                        execute_sql("DELETE FROM students WHERE id = %s", (del_ids[0],))
                    else:
                        placeholders = ",".join(["%s"] * len(del_ids))
                        execute_sql(f"DELETE FROM students WHERE id IN ({placeholders})", del_ids)
                    st.toast(f"✅ 已删除 {len(del_ids)} 名学生", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()

            # ── 上行：全选 + 分页 ──
            r1_sel, r1_info, r1_ps, r1_prev, r1_num, r1_next = st.columns([1.5, 2, 1, 0.7, 0.7, 0.7])
            with r1_sel:
                select_all = st.checkbox("全选本页", key=f"s_selall_{page}",
                    help="勾选后删除操作将应用于本页全部记录")
            with r1_info:
                st.caption(f"共 **{total}** 名学生")
            with r1_ps:
                st.selectbox("每页", [50, 100, 200, 500], key="s_ps",
                             on_change=lambda: st.session_state.update({"s_pg": 1}),
                             label_visibility="collapsed")
            with r1_prev:
                if st.button("◀", key="s_pp", disabled=(page <= 1), use_container_width=True):
                    st.session_state.s_pg = page - 1; st.rerun()
            with r1_num:
                st.markdown(f"<div style='text-align:center;padding-top:5px;font-weight:500'>{page}/{total_pages}</div>", unsafe_allow_html=True)
            with r1_next:
                if st.button("▶", key="s_np", disabled=(page >= total_pages), use_container_width=True):
                    st.session_state.s_pg = page + 1; st.rerun()

            # ── 下行：删除(红) + 保存(蓝) ──
            del_count = len(page_df) if select_all else (int(edited["选择"].sum()) if "选择" in edited.columns else 0)
            del_disabled = "disabled" if del_count == 0 else ""
            save_disabled = "disabled" if not changes else ""
            del_btn_label = f"🗑️ 删除（{del_count}人）"
            save_btn_label = f"💾 保存修改（{len(changes)}处）" if changes else "💾 保存修改"

            r2_del, r2_save = st.columns([1, 1])
            with r2_del:
                st.button(del_btn_label, key=f"s_del_btn_{page}", type="secondary",
                          disabled=(del_count == 0), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "s_del"}))
            with r2_save:
                st.button(save_btn_label, key=f"s_save_btn_{page}", type="primary",
                          disabled=(not changes), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "s_save"}))

    # ── Tab 2: 新增/编辑 ──
    with tab2:
        st.markdown("#### ➕ 新增 / 编辑学生")

        edit_id = st.number_input("编辑学生 ID（留空为新增）", min_value=0, value=0, step=1, key="edit_sid")
        defaults = {}
        if edit_id > 0:
            existing = query_df("SELECT * FROM students WHERE id = %s", (edit_id,))
            if not existing.empty:
                defaults = existing.iloc[0].to_dict()
                st.info(f"正在编辑：**{defaults.get('name', '')}** ({defaults.get('student_id', '')})")

        with st.form("student_form"):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input("姓名 *", value=defaults.get("name", ""), placeholder="必填")
                id_card = st.text_input("身份证号 *", value=defaults.get("id_card", ""),
                                        max_chars=18, placeholder="必填，18位")
                student_id = st.text_input("学号 *", value=defaults.get("student_id", ""), placeholder="必填")
            with c2:
                grade = st.text_input("年级", value=defaults.get("grade", ""), placeholder="如：2024级")
                college = st.text_input("学院", value=defaults.get("college", ""))
                major = st.text_input("专业", value=defaults.get("major", ""))
            class_name = st.text_input("班级", value=defaults.get("class_name", ""), placeholder="如：软件工程1班")

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                submitted = st.form_submit_button("💾 保存", use_container_width=True)
            with col_btn2:
                delete_btn = st.form_submit_button("🗑️ 删除", use_container_width=True) if edit_id > 0 else False

            if submitted:
                if not all([name, id_card, student_id]):
                    st.error("❌ 姓名、身份证号、学号为必填项")
                else:
                    try:
                        if edit_id > 0:
                            execute_sql(
                                """UPDATE students SET name=%s,id_card=%s,student_id=%s,
                                   grade=%s,college=%s,major=%s,class_name=%s WHERE id=%s""",
                                (name, id_card, student_id, grade, college, major, class_name, edit_id))
                            st.success("✅ 已更新")
                        else:
                            execute_sql(
                                """INSERT INTO students (name,id_card,student_id,grade,college,major,class_name)
                                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                                (name, id_card, student_id, grade, college, major, class_name))
                            st.success("✅ 已添加")
                        st.rerun()
                    except Exception as e:
                        if "UNIQUE" in str(e) or "unique" in str(e).lower():
                            st.error("❌ 身份证号或学号已存在，不能重复")
                        else:
                            st.error(f"保存失败：{e}")

            if delete_btn:
                execute_sql("DELETE FROM students WHERE id = %s", (edit_id,))
                st.success("✅ 已删除")

    # ── Tab 3: 导入 Excel ──
    with tab3:
        st.markdown("#### 📥 从 Excel 导入学生")

        # 下载模板
        template_cols = ["身份证号", "学号", "姓名", "年级", "学院", "专业", "班级"]
        template_df = make_template_df(template_cols)
        template_bytes = excel_export(template_df, "学生导入模板")
        st.download_button(
            "📄 下载导入模板",
            data=template_bytes,
            file_name="学生导入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="secondary"
        )
        st.caption("Excel 表头建议包含：身份证号、学号、姓名、年级、学院、专业、班级")

        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="stu_upload")
        if uploaded:
            try:
                raw_df = read_excel_upload(uploaded)
                st.info(f"📄 检测到 {len(raw_df)} 行，列名：{', '.join(raw_df.columns[:8])}")

                col_map = {}
                for col in raw_df.columns:
                    cl = col.strip().lower()
                    if "身份证" in cl or "id_card" in cl: col_map["id_card"] = col
                    elif "学号" in cl or "student_id" in cl or "编号" in cl: col_map["student_id"] = col
                    elif "姓名" in cl or "name" in cl: col_map["name"] = col
                    elif "年级" in cl or "grade" in cl: col_map["grade"] = col
                    elif "学院" in cl or "college" in cl: col_map["college"] = col
                    elif "专业" in cl or "major" in cl: col_map["major"] = col
                    elif "班级" in cl or "class" in cl: col_map["class_name"] = col

                if not all(k in col_map for k in ["id_card", "student_id", "name"]):
                    st.error("❌ 缺少必要列：身份证号、学号、姓名")
                    st.json(col_map)
                else:
                    preview = raw_df.head(5)
                    st.dataframe(preview.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

                    if st.button("✅ 确认导入", use_container_width=True, type="primary"):
                        mapped = raw_df.rename(columns={v: k for k, v in col_map.items()})
                        wanted = [c for c in ["id_card","student_id","name","grade","college","major","class_name"] if c in mapped.columns]
                        mapped = mapped[wanted].where(pd.notnull(mapped), None)

                        # 处理数字型字段（2023 → "2023"，避免变成 "2023.0"）
                        for text_col in ["grade", "college", "major", "class_name"]:
                            if text_col in mapped.columns:
                                mapped[text_col] = mapped[text_col].apply(
                                    lambda x: safe_field(x) if pd.notna(x) else None
                                )

                        success_count = 0
                        update_count = 0
                        errors = []
                        progress = st.progress(0)
                        total_rows = len(mapped)
                        for i, (_, row) in enumerate(mapped.iterrows()):
                            try:
                                # 使用 INSERT OR REPLACE 处理重复：学生已存在则更新全部字段
                                execute_sql(
                                    """INSERT OR REPLACE INTO students
                                       (id, id_card, student_id, name, grade, college, major, class_name)
                                       VALUES (
                                           (SELECT id FROM students WHERE student_id=%s),
                                           %s, %s, %s, %s, %s, %s, %s
                                       )""",
                                    (safe_str(row["student_id"]),
                                     safe_str(row["id_card"]), safe_str(row["student_id"]), safe_str(row["name"]),
                                     row.get("grade"), row.get("college"), row.get("major"), row.get("class_name")))
                                success_count += 1
                            except Exception as e:
                                errors.append(f"{row.get('name','?')}: {str(e)[:80]}")
                            progress.progress((i + 1) / total_rows)

                        # 写日志
                        write_import_log(
                            module="学生管理",
                            filename=uploaded.name,
                            total=total_rows,
                            success=success_count,
                            errors=errors
                        )

                        # 展示结果
                        sm_ok, sm_err = st.columns(2)
                        sm_ok.metric("✅ 导入成功", success_count)
                        sm_err.metric("⚠️ 导入失败", len(errors))
                        if errors:
                            with st.expander(f"查看 {len(errors)} 条失败详情"):
                                for err in errors:
                                    st.caption(f"• {err}")
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")

# ═════════════════════════════════════════════════════════
# 3. 征订总表（原始征订数据管理 + 一键下发）
# ═════════════════════════════════════════════════════════

def subscription_management():
    """征订总表：导入教务处原始征订数据（专业年级粒度），一键下发拆分到班级"""
    show_header("📋 征订总表", "管理原始征订数据（专业/年级粒度），自动按实际班级人数分摊后下发")

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先在「学期管理」中添加学期")
        return

    semester_options = ["全部"] + [f"{r['id']}|{r['name']}" for _, r in semesters.iterrows()]
    master_books = query_df("SELECT id, name, isbn, publisher, editor, price, course_name FROM textbooks_master ORDER BY name")
    master_options = [(0, "➕ 新增教材...")] + [(r["id"], r["name"]) for _, r in master_books.iterrows()]

    tab1, tab2, tab3, tab4 = st.tabs(["📋 征订总览", "➕ 手动新增", "📥 导入 Excel", "🚀 一键下发"])

    # ── Tab1：征订总览 ──
    with tab1:
        st.markdown("#### 📋 征订总览")
        st.caption("这里列出教务处原始征订数据（专业/年级粒度），支持在线编辑和批量删除。勾选后可在 Tab4 一键下发。")
        col1, col2, col3 = st.columns(3)
        with col1:
            f_sub_sem = st.selectbox("学期", semester_options, key="sub_sem_filter")
        with col2:
            sub_college_opts = ["全部"] + get_filtered_colleges()
            f_sub_college = st.selectbox("学院", sub_college_opts, key="sub_college_filter")
        with col3:
            status_opts = ["全部", "待下发", "已下发"]
            f_sub_status = st.selectbox("状态", status_opts, key="sub_status_filter")

        sub_sql = """SELECT s.id, sm.name as semester_name, s.book_name, s.isbn,
                            s.college, s.major, s.grade, s.class_scope,
                            s.total_qty, s.teacher_qty, s.price,
                            s.status, s.remark, s.source, s.created_at, s.dispatched_at
                     FROM textbook_subscriptions s
                     JOIN semesters sm ON s.semester_id = sm.id
                     WHERE 1=1"""
        sub_params = []
        if f_sub_sem != "全部":
            sub_sql += " AND s.semester_id = %s"; sub_params.append(int(f_sub_sem.split("|")[0]))
        if f_sub_college != "全部":
            sub_sql += " AND s.college = %s"; sub_params.append(f_sub_college)
        if f_sub_status != "全部":
            status_map = {"待下发": "pending", "已下发": "dispatched"}
            sub_sql += " AND s.status = %s"; sub_params.append(status_map[f_sub_status])
        sub_sql += " ORDER BY s.semester_id DESC, s.college, s.major, s.grade, s.book_name"

        sub_df = query_df(sub_sql, tuple(sub_params) if sub_params else None)
        total = len(sub_df)

        if not sub_df.empty:
            # 统计
            pending_cnt = (sub_df["status"] == "pending").sum()
            dispatched_cnt = (sub_df["status"] == "dispatched").sum()
            mc1, mc2, mc3 = st.columns(3)
            mc1.metric("📄 合计记录", total)
            mc2.metric("⏳ 待下发", pending_cnt)
            mc3.metric("✅ 已下发", dispatched_cnt)

            # ── 分页（从 session_state 读取，控件移到表格下方）──
            import math
            sub_ps = st.session_state.get("sub_ps", 50)
            sub_page = st.session_state.get("sub_page", 1)
            sub_pages = max(1, math.ceil(total / sub_ps))
            # 页码保护
            if sub_page > sub_pages:
                sub_page = sub_pages
                st.session_state.sub_page = sub_pages

            st.caption(f"共 **{total}** 条 ｜ 第 {sub_page}/{sub_pages} 页")

            # 当前页
            s_start = (sub_page - 1) * sub_ps
            s_end = min(s_start + sub_ps, total)
            page_df = sub_df.iloc[s_start:s_end].copy()

            # 原始备份用于对比修改（先用 NaN 安全的字典）
            edit_cols = ["college", "major", "grade", "class_scope", "book_name", "isbn", "price", "total_qty", "teacher_qty", "remark"]
            orig_key = "_sub_orig"
            page_df[orig_key] = page_df[edit_cols].apply(
                lambda row: {col: ("" if pd.isna(row[col]) and col != "price" and col != "total_qty" and col != "teacher_qty" else row[col]) for col in edit_cols},
                axis=1
            ).tolist()

            # ── 构建展示用 DataFrame，确保类型正确 ──
            # 全选初始化
            sel_all_init = st.session_state.get(f"sub_selall_{sub_page}", False)

            # 状态转中文
            status_lbl = {"pending": "⏳ 待下发", "dispatched": "✅ 已下发"}
            status_display = page_df["status"].map(status_lbl).fillna(page_df["status"])

            # 班级范围（处理 None/空）
            class_scope_display = page_df["class_scope"].fillna("全部班级").replace("", "全部班级")

            # 下发时间
            dispatched_display = page_df["dispatched_at"].fillna("—")

            # 数值列——强制转 float 并填 0，避免 NaN 导致 NumberColumn 异常
            price_vals = pd.to_numeric(page_df["price"], errors="coerce").fillna(0.0)
            total_qty_vals = pd.to_numeric(page_df["total_qty"], errors="coerce").fillna(0)
            teacher_qty_vals = pd.to_numeric(page_df["teacher_qty"], errors="coerce").fillna(0)

            # 文本列——填空字符串
            text_fill = lambda s: s.fillna("").astype(str)

            display_df = pd.DataFrame({
                "选择": sel_all_init,
                "学期": page_df["semester_name"].fillna("").values,
                "学院": text_fill(page_df["college"]),
                "专业": text_fill(page_df["major"]),
                "年级": text_fill(page_df["grade"]),
                "班级范围": text_fill(class_scope_display),
                "教材名称": text_fill(page_df["book_name"]),
                "书号": text_fill(page_df["isbn"]),
                "单价": price_vals,
                "征订总量": total_qty_vals.astype(int),
                "教师用书": teacher_qty_vals.astype(int),
                "状态": status_display.values,
                "备注": text_fill(page_df["remark"]),
                "下发时间": dispatched_display.values,
            })

            edited = st.data_editor(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_order=["选择", "学期", "学院", "专业", "年级", "班级范围",
                             "教材名称", "书号", "单价", "征订总量", "教师用书", "状态", "备注", "下发时间"],
                column_config={
                    "选择": st.column_config.CheckboxColumn("选择"),
                    "学期": st.column_config.TextColumn("学期", disabled=True, alignment="center"),
                    "学院": st.column_config.TextColumn("学院", alignment="center"),
                    "专业": st.column_config.TextColumn("专业", alignment="center"),
                    "年级": st.column_config.TextColumn("年级", alignment="center"),
                    "班级范围": st.column_config.TextColumn("班级范围", alignment="center"),
                    "教材名称": st.column_config.TextColumn("教材名称", alignment="center"),
                    "书号": st.column_config.TextColumn("书号", alignment="center"),
                    "单价": st.column_config.NumberColumn("单价", format="¥%.2f", alignment="center"),
                    "征订总量": st.column_config.NumberColumn("征订总量", min_value=0, alignment="center"),
                    "教师用书": st.column_config.NumberColumn("教师用书", min_value=0, alignment="center"),
                    "状态": st.column_config.TextColumn("状态", disabled=True, alignment="center"),
                    "备注": st.column_config.TextColumn("备注", alignment="center"),
                    "下发时间": st.column_config.TextColumn("下发时间", disabled=True, alignment="center"),
                },
                key=f"sub_editor_{sub_page}"
            )

            # ═══ 布局：上行(全选+分页) + 下行(红删蓝存) ═══
            # 检测修改（DB列名 → 展示列名映射）
            col_map = {
                "college": "学院", "major": "专业", "grade": "年级",
                "class_scope": "班级范围", "book_name": "教材名称", "isbn": "书号",
                "price": "单价", "total_qty": "征订总量", "teacher_qty": "教师用书", "remark": "备注"
            }
            changes_sub = []
            for i in range(len(edited)):
                orig = page_df.iloc[i][orig_key]
                for db_col, disp_col in col_map.items():
                    new_val = edited.iloc[i].get(disp_col)
                    old_val = orig.get(db_col)
                    if db_col in ("price", "total_qty", "teacher_qty"):
                        try:
                            nv = float(new_val) if not pd.isna(new_val) else 0.0
                        except (ValueError, TypeError):
                            nv = 0.0
                        try:
                            ov = float(old_val) if not pd.isna(old_val) else 0.0
                        except (ValueError, TypeError):
                            ov = 0.0
                        if abs(nv - ov) > 1e-9:
                            changes_sub.append((int(page_df.iloc[i]["id"]), db_col, nv))
                    else:
                        nv_str = str(new_val or "").strip()
                        ov_str = str(old_val or "").strip()
                        if nv_str != ov_str:
                            changes_sub.append((int(page_df.iloc[i]["id"]), db_col, nv_str if db_col != "class_scope" else (new_val if new_val and str(new_val) != "全部班级" else "")))

            # 处理按钮触发的操作（用 session_state 代替 query_params）
            pending = st.session_state.get("pending_action", "")
            if pending == "sub_save":
                if changes_sub:
                    for sid, col, val in changes_sub:
                        execute_sql(f"UPDATE textbook_subscriptions SET {col}=%s WHERE id=%s", (val, sid))
                    st.toast(f"✅ 已保存 {len(changes_sub)} 处修改", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()
            elif pending == "sub_del":
                select_all = st.session_state.get(f"sub_selall_{sub_page}", False)
                if select_all:
                    del_ids_sub = tuple(page_df["id"].tolist())
                else:
                    selected = edited[edited["选择"] == True].index if "选择" in edited.columns else []
                    del_ids_sub = tuple(page_df.iloc[selected]["id"].tolist()) if len(selected) > 0 else ()
                if del_ids_sub:
                    if len(del_ids_sub) == 1:
                        execute_sql("DELETE FROM textbook_subscriptions WHERE id = %s", (del_ids_sub[0],))
                    else:
                        ph_sub = ",".join(["%s"] * len(del_ids_sub))
                        execute_sql(f"DELETE FROM textbook_subscriptions WHERE id IN ({ph_sub})", del_ids_sub)
                    st.toast(f"✅ 已删除 {len(del_ids_sub)} 条", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()

            # ── 上行：全选 + 分页 ──
            r1_sel, r1_info, r1_ps, r1_prev, r1_num, r1_next = st.columns([1.5, 2, 1, 0.7, 0.7, 0.7])
            with r1_sel:
                select_all = st.checkbox("全选本页", key=f"sub_selall_{sub_page}",
                    help="勾选后删除操作将应用于本页全部记录")
            with r1_info:
                st.caption(f"共 **{total}** 条")
            with r1_ps:
                st.selectbox("每页", [50, 100, 200], key="sub_ps",
                             on_change=lambda: st.session_state.update({"sub_page": 1}),
                             label_visibility="collapsed")
            with r1_prev:
                if st.button("◀", key="sub_pp", disabled=(sub_page <= 1), use_container_width=True):
                    st.session_state.sub_page = sub_page - 1; st.rerun()
            with r1_num:
                st.markdown(f"<div style='text-align:center;padding-top:5px;font-weight:500'>{sub_page}/{sub_pages}</div>", unsafe_allow_html=True)
            with r1_next:
                if st.button("▶", key="sub_np", disabled=(sub_page >= sub_pages), use_container_width=True):
                    st.session_state.sub_page = sub_page + 1; st.rerun()

            # ── 下行：删除(红) + 保存(蓝) ──
            del_count = len(page_df) if select_all else (int(edited["选择"].sum()) if "选择" in edited.columns else 0)
            del_disabled = "disabled" if del_count == 0 else ""
            save_disabled = "disabled" if not changes_sub else ""
            del_btn_label = f"🗑️ 删除（{del_count}条）"
            save_btn_label = f"💾 保存修改（{len(changes_sub)}处）" if changes_sub else "💾 保存修改"

            r2_del, r2_save = st.columns([1, 1])
            with r2_del:
                st.button(del_btn_label, key=f"sub_del_btn_{sub_page}", type="secondary",
                          disabled=(del_count == 0), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "sub_del"}))
            with r2_save:
                st.button(save_btn_label, key=f"sub_save_btn_{sub_page}", type="primary",
                          disabled=(not changes_sub), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "sub_save"}))
        else:
            st.info("暂无征订数据，请通过「手动新增」或「导入 Excel」添加")

    # ── Tab2：手动新增 ──
    with tab2:
        st.markdown("#### ➕ 手动新增征订记录")
        st.caption("按「专业+年级」粒度登记。征订数量按各班学生人数自动计算（每人1本），下发时每班分配其学生人数。")

        sem_id2 = st.selectbox("学期*", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
            format_func=lambda x: x[1], key="sub_sem2")

        # 教材选择（支持多选）
        actual_books = [(r["id"], r["name"]) for _, r in master_books.iterrows()]
        if not actual_books:
            st.warning("⚠️ 教材库为空，请先在「教材表管理」中添加教材")
            return
        picked_books = st.multiselect(
            "选择已有教材（可多选）",
            options=actual_books,
            format_func=lambda x: x[1],
            key="sub_book2",
            placeholder="点击选择一本或多本教材..."
        )

        if picked_books:
            st.caption(f"📚 已选 **{len(picked_books)}** 本教材：")
            # 展示所选教材摘要
            summary_rows = []
            for pid, pname in picked_books:
                r = master_books[master_books["id"] == pid]
                if not r.empty:
                    summary_rows.append({
                        "教材名称": pname,
                        "书号": str(r.iloc[0].get("isbn", "") or ""),
                        "单价(元)": float(r.iloc[0].get("price", 0)),
                    })
            st.dataframe(pd.DataFrame(summary_rows).style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

        st.divider()
        st.caption("🎯 征订范围（按顺序选择，级联过滤）")

        # ── ① 年级 ──
        all_grades2 = get_filtered_grades()
        sub_grade = st.selectbox("① 年级", ["（不限）"] + all_grades2, key="sub_grade2",
            format_func=lambda x: f"📘 {x}" if x != "（不限）" else x)

        # ── ② 学院（级联年级）──
        if sub_grade != "（不限）":
            all_colleges2 = get_filtered_list("students", "college", "grade=%s OR grade=%s",
                (sub_grade.rstrip("级"), sub_grade.rstrip("级") + "级"))
        else:
            all_colleges2 = get_filtered_colleges()
        sub_college = st.selectbox("② 学院", ["（不限）"] + all_colleges2, key="sub_college2",
            format_func=lambda x: f"🏫 {x}" if x != "（不限）" else x)

        # ── ③ 专业（级联年级+学院）──
        major_where_parts = []
        major_params_parts = []
        if sub_grade != "（不限）":
            major_where_parts.append("(grade=%s OR grade=%s)")
            major_params_parts.extend([sub_grade.rstrip("级"), sub_grade.rstrip("级") + "级"])
        if sub_college != "（不限）":
            major_where_parts.append("college=%s")
            major_params_parts.append(sub_college)
        if major_where_parts:
            all_majors2 = get_filtered_list("students", "major",
                " AND ".join(major_where_parts), tuple(major_params_parts))
        else:
            all_majors2 = get_filtered_majors()
        sub_major = st.selectbox("③ 专业", ["（不限）"] + all_majors2, key="sub_major2",
            format_func=lambda x: f"📚 {x}" if x != "（不限）" else x)

        # ── ④ 班级范围（自动计算，无需手动填写）──
        _preview_grade   = normalize_grade(sub_grade) if sub_grade != "（不限）" else None
        _preview_college = None if sub_college == "（不限）" else sub_college
        _preview_major   = None if sub_major  == "（不限）" else sub_major

        auto_scope = ""
        sub_class_names = []
        if _preview_grade or _preview_college or _preview_major:
            preview_df = get_class_student_counts(_preview_grade, _preview_college, _preview_major)
            if not preview_df.empty:
                matched_classes = preview_df["class_name"].tolist()
                total_stu = int(preview_df["student_count"].sum())
                # 自动生成班级范围说明
                parts = []
                if _preview_grade: parts.append(_preview_grade)
                if _preview_college: parts.append(_preview_college)
                if _preview_major: parts.append(_preview_major)
                auto_scope = " ".join(parts) + " 全部班级"
                st.success(f"📋 ④ 班级范围（自动匹配）：**{auto_scope}** — {len(matched_classes)}个班级，{total_stu}名学生")
                # 班级详情
                with st.expander(f"查看匹配的 {len(matched_classes)} 个班级"):
                    class_preview_rows = []
                    for _, pr in preview_df.iterrows():
                        class_preview_rows.append({
                            "班级": pr["class_name"],
                            "学生人数": int(pr["student_count"])
                        })
                    st.dataframe(pd.DataFrame(class_preview_rows).style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

                # 可选：进一步缩小班级范围（多选）
                st.caption("如需精确到特定班级，可从下方多选（留空=匹配全部）")
                sub_class_names = st.multiselect(
                    "精确班级（可多选）",
                    options=matched_classes,
                    key="sub_class_names2",
                    placeholder="留空 = 匹配上述全部班级"
                )
                if sub_class_names:
                    auto_scope = "、".join(sub_class_names)
                    st.info(f"✂️ 仅下发到：**{auto_scope}**")
            else:
                st.warning("⚠️ ④ 当前条件下未找到任何班级，请检查学生数据中是否有对应年级的数据")
                preview_df = pd.DataFrame()  # 确保变量存在
        else:
            st.info("💡 未选择任何条件 = 匹配学生表中所有班级（请谨慎使用）")
            auto_scope = "全部班级"
            preview_df = pd.DataFrame()  # 确保变量存在

        # ── 按班级人数自动计算数量（每人1本）──
        st.caption("📦 征订数量（按班级人数自动计算，每人1本）")
        if not preview_df.empty:
            sub_total_qty = int(preview_df["student_count"].sum())
            q_c1, q_c2 = st.columns(2)
            with q_c1:
                st.metric("征订总量（学生用书）", f"{sub_total_qty} 本",
                         delta=f"{len(preview_df)}个班级" if not sub_class_names else f"{len(sub_class_names)}个精确班级")
            with q_c2:
                sub_teacher_qty = st.number_input("教师用书数量", min_value=0, value=0, step=1, key="sub_teacher_qty2")
            # 每班按人数分配预览
            if sub_class_names:
                filtered_class = preview_df[preview_df["class_name"].isin(sub_class_names)]
            else:
                filtered_class = preview_df
            prows = [{"班级": pr["class_name"], "学生人数": int(pr["student_count"]), "分配数量": int(pr["student_count"])}
                     for _, pr in filtered_class.iterrows()]
            if prows:
                st.caption("📦 每班分配预览（每人1本）：")
                st.dataframe(pd.DataFrame(prows).style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)
        else:
            sub_total_qty = 0
            sub_teacher_qty = 0
            st.warning("⚠️ 请先选择年级/学院/专业，系统将自动按班级人数计算征订数量")

        sub_remark = st.text_input("备注", key="sub_remark2")

        if st.button("💾 保存到征订总表", use_container_width=True, type="primary", key="sub_save2"):
            if not picked_books:
                st.error("❌ 请至少选择一本教材")
                st.stop()

            saved_count = 0
            for pid, pname in picked_books:
                r = master_books[master_books["id"] == pid]
                if r.empty:
                    continue
                b = r.iloc[0]
                execute_sql("""INSERT INTO textbook_subscriptions
                    (semester_id, textbook_id, book_name, isbn, publisher, editor, price, course_name,
                     college, major, grade, class_scope, class_names, total_qty, teacher_qty, status, remark, source)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,'manual')""",
                    (sem_id2[0], pid, b["name"],
                     str(b.get("isbn", "") or ""), str(b.get("publisher", "") or ""),
                     str(b.get("editor", "") or ""), float(b.get("price", 0)),
                     str(b.get("course_name", "") or ""),
                     _preview_college or "", _preview_major or "", _preview_grade or "",
                     auto_scope, ",".join(sub_class_names) if sub_class_names else "",
                     sub_total_qty, sub_teacher_qty, sub_remark))
                saved_count += 1

            st.success(f"✅ 已保存 **{saved_count}** 条征订记录到总表")
            st.rerun()

    # ── Tab3：导入 Excel ──
    with tab3:
        st.markdown("#### 📥 从 Excel 导入原始征订数据")
        st.caption("支持教务处格式：专业年级（如「电商2023」）、征订数量、教师用书数量等列。")

        # 模板下载
        tmpl_cols = ["学期", "学院", "专业", "年级", "班级范围说明", "教材名称", "书号(ISBN)", "出版社", "主编", "单价(元)", "征订总量", "教师用书数量", "备注"]
        tmpl_df = make_template_df(tmpl_cols)
        st.download_button("📄 下载征订总表导入模板", data=excel_export(tmpl_df, "征订总表导入模板"),
            file_name="征订总表导入模板.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="secondary")

        uploaded3 = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="sub_upload3")
        if uploaded3:
            try:
                raw3 = read_excel_upload(uploaded3)
                st.info(f"📄 检测到 {len(raw3)} 行数据")

                # 列映射
                cmap3 = {}
                for col in raw3.columns:
                    cl = col.strip().lower()
                    if "教材" in cl and "名" in cl: cmap3["book_name"] = col
                    elif "学期" in cl: cmap3["semester_name"] = col
                    elif "专业" in cl and "年级" in cl: cmap3["major_grade"] = col  # 合并列
                    elif "年级" in cl: cmap3["grade"] = col
                    elif "学院" in cl: cmap3["college"] = col
                    elif "专业" in cl: cmap3["major"] = col
                    elif "班级" in cl and "范围" in cl: cmap3["class_scope"] = col
                    elif "班级" in cl: cmap3["class_scope"] = col
                    elif "教师" in cl and ("用书" in cl or "数量" in cl): cmap3["teacher_qty"] = col
                    elif "单价" in cl: cmap3["price"] = col
                    elif ("数量" in cl or "合计" in cl or "总量" in cl) and "教师" not in cl: cmap3["total_qty"] = col
                    elif "书号" in cl or "isbn" in cl: cmap3["isbn"] = col
                    elif "出版社" in cl: cmap3["publisher"] = col
                    elif "主编" in cl or "作者" in cl: cmap3["editor"] = col
                    elif "备注" in cl or "说明" in cl: cmap3["remark"] = col

                if "book_name" not in cmap3 or "semester_name" not in cmap3:
                    st.error("❌ 缺少必要列：教材名称、学期")
                else:
                    # 预览
                    st.markdown("**列映射预览：**")
                    mapped_preview = {v: k for k, v in cmap3.items()}
                    preview_rows = [{"Excel列名": k, "识别为": v} for k, v in mapped_preview.items()]
                    st.dataframe(pd.DataFrame(preview_rows).style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

                    has_major_grade = "major_grade" in cmap3
                    if has_major_grade:
                        st.info("📌 检测到「专业年级」合并列，将自动拆解专业名和年级")

                    target_sem = st.selectbox("确认导入到学期", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                        format_func=lambda x: x[1], key="sub_import_sem3")

                    if st.button("✅ 确认导入到征订总表", use_container_width=True, type="primary", key="sub_import3_btn"):
                        sem_map3 = {r["name"]: r["id"] for _, r in semesters.iterrows()}
                        success3, errors3, total3 = 0, [], len(raw3)
                        progress3 = st.progress(0, text="正在导入...")

                        # ── 预处理：前向填充学期和学院（兼容合并单元格模板）──
                        last_sem, last_college = "", ""
                        for i in range(len(raw3)):
                            sv = safe_str(raw3.iloc[i].get(cmap3.get("semester_name", ""), ""))
                            cv = safe_str(raw3.iloc[i].get(cmap3.get("college", ""), ""))
                            if sv:
                                last_sem = sv
                            elif last_sem:
                                raw3.at[raw3.index[i], cmap3["semester_name"]] = last_sem
                            if cv:
                                last_college = cv
                            elif last_college and "college" in cmap3:
                                raw3.at[raw3.index[i], cmap3["college"]] = last_college

                        # 统计智能解析命中数
                        auto_parse_count = 0

                        for i, (_, row3) in enumerate(raw3.iterrows()):
                            try:
                                book_name3 = safe_str(row3.get(cmap3.get("book_name", ""), ""))
                                if not book_name3: continue
                                # 学期
                                sem_name3 = safe_str(row3.get(cmap3.get("semester_name", ""), ""))
                                sid3 = sem_map3.get(sem_name3, target_sem[0])
                                # 专业年级拆解
                                mg_val3 = safe_str(row3.get(cmap3.get("major_grade", ""), "")) if has_major_grade else ""
                                grade3  = safe_str(row3.get(cmap3.get("grade", ""), ""))
                                major3  = safe_str(row3.get(cmap3.get("major", ""), ""))
                                if has_major_grade and mg_val3 and not (grade3 and major3):
                                    mg_parts3 = re.split(r'(\d{4})', mg_val3, maxsplit=1)
                                    if len(mg_parts3) >= 2:
                                        major3 = major3 or mg_parts3[0].strip()
                                        grade3 = grade3 or mg_parts3[1].strip()
                                    else:
                                        major3 = major3 or mg_val3
                                college3     = safe_str(row3.get(cmap3.get("college", ""), ""))

                                # ── 智能解析：从班级范围说明中提取专业/年级 ──
                                class_scope3 = safe_str(row3.get(cmap3.get("class_scope", ""), ""))
                                parsed_grades = []  # 从 scope 解析出的年级列表
                                if (not major3 or not grade3) and class_scope3:
                                    parsed_major, parsed_grades = parse_major_grade_from_scope(class_scope3, college3)
                                    if parsed_major and not major3:
                                        major3 = parsed_major
                                        auto_parse_count += 1
                                    if parsed_grades and not grade3:
                                        grade3 = parsed_grades[0]  # 先用第一个年级做后续判断
                                # 确定最终要写入的年级列表（优先用模板显式年级，否则用解析结果）
                                final_grades = [grade3] if grade3 else []
                                if parsed_grades and not grade3:
                                    final_grades = parsed_grades
                                # 如果没有年级，给个空列表（兜底）
                                if not final_grades:
                                    final_grades = [grade3]
                                total_qty3   = safe_int(row3.get(cmap3.get("total_qty", ""), 0))
                                teacher_qty3 = safe_int(row3.get(cmap3.get("teacher_qty", ""), 0))
                                price3       = safe_float(row3.get(cmap3.get("price", ""), 0))
                                isbn3        = safe_str(row3.get(cmap3.get("isbn", ""), ""))
                                pub3         = safe_str(row3.get(cmap3.get("publisher", ""), ""))
                                ed3          = safe_str(row3.get(cmap3.get("editor", ""), ""))
                                remark3      = safe_str(row3.get(cmap3.get("remark", ""), ""))
                                # 教材主表
                                old3 = query_df("SELECT id FROM textbooks_master WHERE name=%s", (book_name3,))
                                if old3.empty:
                                    execute_sql("INSERT INTO textbooks_master (name,isbn,publisher,editor,price) VALUES (%s,%s,%s,%s,%s)",
                                        (book_name3, isbn3, pub3, ed3, price3))
                                    _c3 = sqlite3.connect(get_sqlite_path())
                                    _cc3 = _c3.cursor(); _cc3.execute("SELECT MAX(id) FROM textbooks_master"); tid3 = _cc3.fetchone()[0]; _c3.close()
                                else:
                                    tid3 = old3.iloc[0]["id"]
                                # 写入征订总表（多年级则拆分为多条记录，每条保留原始总量）
                                for g3 in final_grades:
                                    execute_sql("""INSERT INTO textbook_subscriptions
                                        (semester_id,textbook_id,book_name,isbn,publisher,editor,price,
                                         college,major,grade,class_scope,total_qty,teacher_qty,status,remark,source)
                                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,'import')""",
                                        (sid3, tid3, book_name3, isbn3, pub3, ed3, price3,
                                         college3, major3, g3, class_scope3, total_qty3, teacher_qty3, remark3))
                                    success3 += 1
                            except Exception as e3:
                                errors3.append(f"{safe_str(row3.get(cmap3.get('book_name',''),''))}: {str(e3)[:100]}")
                            progress3.progress((i + 1) / total3, text=f"已处理 {i+1}/{total3}")
                        # 统一展示结果（对齐学生管理）
                        mc_ok, mc_err = st.columns(2)
                        mc_ok.metric("✅ 导入成功", success3)
                        mc_err.metric("⚠️ 导入失败", len(errors3))
                        if auto_parse_count > 0:
                            st.info(f"🔍 智能解析：从「班级范围说明」自动提取了 **{auto_parse_count}** 条的专业/年级信息")
                        if errors3:
                            with st.expander(f"查看 {len(errors3)} 条失败详情"):
                                for e3 in errors3:
                                    st.caption(f"• {e3}")
                        write_import_log("征订总表导入", uploaded3.name, total3, success3, errors3)
                        st.rerun()
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")

    # ── Tab4：一键下发 ──
    with tab4:
        st.markdown("#### 🚀 一键下发：将征订总表数据拆分到班级征订明细")
        st.caption("""
        **下发逻辑：**  
        - 系统根据「学院 + 专业 + 年级」条件，在 students 表中匹配实际班级  
        - 按各班级实际学生人数下发（每人1本），班级学生人数即为该班征订数量  
        - 写入「教材征订表（按班级明细）」，供后续发放使用  
        - 下发后状态变为「已下发」，可重复下发（会覆盖同条件的已有征订）
        """)

        # 选择要下发的记录
        pending_df = query_df("""
            SELECT ts.id, sm.name as semester_name, ts.book_name, ts.college, ts.major, ts.grade,
                   ts.class_scope, ts.total_qty, ts.teacher_qty, ts.status
            FROM textbook_subscriptions ts
            JOIN semesters sm ON ts.semester_id = sm.id
            WHERE ts.status = 'pending'
            ORDER BY ts.semester_id DESC, ts.college, ts.major, ts.grade
        """)

        if pending_df.empty:
            st.info("✅ 暂无待下发的征订记录")
        else:
            st.markdown(f"共有 **{len(pending_df)}** 条待下发记录：")
            status_show = pending_df.copy()
            status_show["班级范围"] = status_show["class_scope"].fillna("全部班级").replace("", "全部班级")
            st.dataframe(status_show[["id","semester_name","college","major","grade","班级范围","book_name","total_qty","teacher_qty"]].rename(columns={
                "semester_name":"学期","college":"学院","major":"专业","grade":"年级",
                "book_name":"教材名称","total_qty":"征订总量","teacher_qty":"教师用书"
            }).style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

            # 下发选项 — checkbox 多选 + 全选/取消全选
            st.divider()
            st.caption("🎯 选择要下发的记录（勾选后点击底部按钮下发）")

            # 全选 / 取消全选 开关
            select_all_key = "dispatch_select_all"
            if select_all_key not in st.session_state:
                st.session_state[select_all_key] = False

            cb_col1, cb_col2 = st.columns([1, 5])
            with cb_col1:
                if st.button("☑️ 全选" if not st.session_state[select_all_key] else "☐ 取消全选",
                             use_container_width=True, key="dispatch_toggle_all"):
                    st.session_state[select_all_key] = not st.session_state[select_all_key]
                    st.rerun()
            with cb_col2:
                select_count = len(pending_df) if st.session_state[select_all_key] else 0
                st.caption(f"已选 **{select_count}** / {len(pending_df)} 条" if st.session_state[select_all_key] else "点击「全选」或逐条勾选")

            # 逐条 checkbox
            selected_ids = []
            for _, prow in pending_df.iterrows():
                pid = int(prow["id"])
                cb_key = f"dispatch_cb_{pid}"
                # 根据全选状态决定默认值
                default_val = st.session_state[select_all_key]
                if st.checkbox(
                    f"#{pid} 【{prow['semester_name']}】{prow['college']} {prow['major']} {prow['grade']}级 — {prow['book_name']}（总量{prow['total_qty']}，教师{prow['teacher_qty']}）",
                    value=st.session_state.get(cb_key, default_val),
                    key=cb_key
                ):
                    selected_ids.append(pid)

            if selected_ids:
                st.caption(f"✅ 已选中 **{len(selected_ids)}** 条记录")
            else:
                st.warning("⚠️ 请至少勾选一条记录")

            conflict_mode = st.radio("遇到同条件已有征订时", ["跳过（保留原有）", "覆盖（删除原有后重新写入）"], horizontal=True, key="conflict_mode")

            dispatch_disabled = len(selected_ids) == 0
            if st.button("🚀 确认下发所选记录", use_container_width=True, type="primary",
                         key="dispatch_btn", disabled=dispatch_disabled):
                to_dispatch = pending_df[pending_df["id"].isin(selected_ids)]

                # 获取完整信息
                dispatch_detail = query_df("""
                    SELECT ts.*, sm.id as sem_id_val
                    FROM textbook_subscriptions ts
                    JOIN semesters sm ON ts.semester_id = sm.id
                """ + " WHERE ts.id IN (" + ",".join(["%s"]*len(to_dispatch)) + ")",
                    tuple(to_dispatch["id"].tolist()))

                total_created = 0
                total_skipped = 0
                total_errors  = []
                total_records = len(dispatch_detail)
                progress = st.progress(0, text="正在下发...")
                st.info(f"🔄 正在处理 **{total_records}** 条征订记录...")

                for idx, (_, sub_row) in enumerate(dispatch_detail.iterrows()):
                    try:
                        sub_id    = sub_row["id"]
                        sem_id_v  = sub_row["semester_id"]
                        textbook_id = sub_row["textbook_id"]
                        book_name = sub_row["book_name"]

                        # 如果 textbook_id 为空，尝试通过书名查找教材主表
                        if not textbook_id:
                            bk_lookup = query_df("SELECT id FROM textbooks_master WHERE name=%s", (book_name,))
                            if not bk_lookup.empty:
                                textbook_id = int(bk_lookup.iloc[0]["id"])
                                # 回填 textbook_id
                                execute_sql("UPDATE textbook_subscriptions SET textbook_id=%s WHERE id=%s", (textbook_id, sub_row["id"]))
                            else:
                                total_errors.append(f"「{book_name}」：教材库中不存在，请先在教材征订表中添加该教材")
                                progress.progress((idx + 1) / total_records, text=f"处理中 {idx+1}/{total_records}")
                                continue
                        textbook_id = int(textbook_id)
                        college_v = sub_row["college"] or None
                        major_v   = sub_row["major"] or None
                        grade_v   = normalize_grade(sub_row["grade"]) if sub_row.get("grade") else None
                        total_qty_v   = int(sub_row["total_qty"] or 0)
                        teacher_qty_v = int(sub_row["teacher_qty"] or 0)
                        remark_v  = sub_row.get("remark", "") or ""

                        # 读取 class_names（Tab2 中手动精确选择的班级）
                        class_names_str = sub_row.get("class_names", "") or ""
                        class_names_list = [c.strip() for c in class_names_str.split(",") if c.strip()] if class_names_str else None

                        # 按班级人数下发（每人1本），每班分配 = 该班学生人数
                        class_df = get_class_student_counts(grade_v, college_v, major_v, class_names=class_names_list)
                        splits_v = [(row["class_name"], int(row["student_count"]), int(row["student_count"]))
                                     for _, row in class_df.iterrows()] if not class_df.empty else [(None, 0, 0)]

                        if len(splits_v) == 1 and splits_v[0][0] is None:
                            # 没找到匹配的班级
                            total_errors.append(f"「{book_name}」({college_v} {major_v} {grade_v}级)：未找到匹配班级，跳过")
                            progress.progress((idx + 1) / total_records, text=f"处理中 {idx+1}/{total_records}")
                            continue

                        for cn, sc, alloc in splits_v:
                            if alloc <= 0:
                                continue
                            # 检查是否已有同条件的征订记录
                            existing = query_df("""
                                SELECT id FROM textbook_orders
                                WHERE semester_id=%s AND textbook_id=%s AND class_name=%s
                            """, (sem_id_v, textbook_id, cn))

                            if not existing.empty:
                                if conflict_mode == "跳过（保留原有）":
                                    total_skipped += 1
                                    continue
                                else:
                                    # 覆盖：删除原有
                                    ph_del = ",".join(["%s"]*len(existing))
                                    execute_sql(f"DELETE FROM textbook_orders WHERE id IN ({ph_del})",
                                        tuple(existing["id"].tolist()))

                            execute_sql("""INSERT INTO textbook_orders
                                (semester_id, textbook_id, grade, college, major, class_name, quantity, remark)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                                (sem_id_v, textbook_id, grade_v or "", college_v or "", major_v or "",
                                 cn, alloc, (remark_v or "")))
                            total_created += 1

                        # 更新状态为已下发
                        execute_sql("UPDATE textbook_subscriptions SET status='dispatched', dispatched_at=CURRENT_TIMESTAMP WHERE id=%s", (sub_id,))

                    except Exception as e_d:
                        total_errors.append(f"ID={sub_row['id']} {sub_row['book_name']}: {str(e_d)[:120]}")
                    progress.progress((idx + 1) / total_records, text=f"处理中 {idx+1}/{total_records}")

                progress.empty()
                st.divider()
                st.markdown("### 📊 下发结果")
                mc1, mc2, mc3 = st.columns(3)
                mc1.metric("✅ 新增班级征订", total_created)
                mc2.metric("⏭️ 跳过（已有）", total_skipped)
                mc3.metric("⚠️ 失败/异常", len(total_errors))
                if total_errors:
                    with st.expander(f"查看 {len(total_errors)} 条失败详情"):
                        for e_d in total_errors:
                            st.caption(f"• {e_d}")
                if total_created > 0:
                    st.success(f"🎉 下发完成！新增 **{total_created}** 条班级征订记录到教材征订表")
                else:
                    st.warning("⚠️ 没有新增任何班级征订记录，请检查征订条件和班级数据")
                # 不自动 rerun，让用户查看结果后手动刷新


# ═════════════════════════════════════════════════════════
# 3.5. 教材表管理（独立管理 textbooks_master）
# ═════════════════════════════════════════════════════════

def textbook_master_management():
    """教材表管理 — 维护教材主表（textbooks_master），供征订时选用"""
    import sqlite3, math
    show_header("📖 教材表管理", "维护教材库，征订时从库中选择")

    tab1, tab2, tab3 = st.tabs(["📋 教材列表", "➕ 新增教材", "📥 导入 Excel"])

    # ── 引用次数统计 ──
    usage_sql = "SELECT textbook_id, COUNT(*) as cnt FROM textbook_orders GROUP BY textbook_id UNION ALL SELECT textbook_id, COUNT(*) as cnt FROM textbook_subscriptions WHERE textbook_id IS NOT NULL GROUP BY textbook_id"
    usage_df = query_df(usage_sql)
    usage_map = {}
    if not usage_df.empty:
        for _, ur in usage_df.iterrows():
            tid = safe_int(ur["textbook_id"])
            if tid <= 0:
                continue
            usage_map[tid] = usage_map.get(tid, 0) + safe_int(ur["cnt"])

    # ═══════════════ Tab1: 教材列表 ═══════════════
    with tab1:
        # ── 默认折扣率设置 ──
        default_dr = st.number_input("🔢 导入默认折扣率（如 0.76 = 76折，1.0 = 原价）",
            min_value=0.0, max_value=1.0, value=st.session_state.get("master_default_dr", 0.76),
            step=0.01, format="%.2f", key="master_default_dr",
            help="导入 Excel 时所有教材将自动应用此折扣率，之后可在列表中逐个修改")

        # ── 批量操作折扣率 ──
        st.divider()
        st.markdown("##### 🔧 批量操作")
        b_col1, b_col2 = st.columns([1, 1])
        with b_col1:
            st.number_input("批量设置折扣率", min_value=0.0, max_value=1.0,
                            value=st.session_state.get("master_default_dr", 0.76),
                            step=0.01, format="%.2f", key="batch_dr_widget",
                            help="勾选教材后，点击右侧按钮批量修改折扣率")
        with b_col2:
            st.markdown("<div style='padding-top:28px;'></div>", unsafe_allow_html=True)
            st.button(f"🔄 批量修改折扣率", key="batch_dr_btn",
                      use_container_width=True, type="secondary",
                      on_click=lambda: st.session_state.update({"pending_action": "batch_dr"}))
        st.divider()

        search_key = st.text_input("🔍 搜索教材（名称/ISBN/出版社）", key="master_search", placeholder="输入关键词搜索...")

        # ── 高级筛选 ──
        with st.expander("🔎 高级筛选", expanded=False):
            af_col1, af_col2, af_col3, af_col4 = st.columns([1, 1, 1, 1])
            all_masters = query_df(
                "SELECT DISTINCT publisher, editor, discount_rate, textbook_type FROM textbooks_master"
            )
            pub_opts = sorted(all_masters["publisher"].dropna().astype(str).unique().tolist()) if not all_masters.empty else []
            editor_opts = sorted(all_masters["editor"].dropna().astype(str).unique().tolist()) if not all_masters.empty else []
            type_opts = sorted([t for t in all_masters["textbook_type"].dropna().astype(str).unique().tolist() if t]) if not all_masters.empty else []
            with af_col1:
                filter_pub = st.multiselect("出版社", pub_opts, key="filter_pub", placeholder="全部出版社")
            with af_col2:
                filter_editor = st.multiselect("主编", editor_opts, key="filter_editor", placeholder="全部主编")
            with af_col3:
                filter_dr_mode = st.selectbox("折扣率", ["全部", "有折扣(<1.0)", "无折扣(=1.0)"], key="filter_dr_mode")
                if filter_dr_mode == "全部":
                    filter_dr_range = st.slider("折扣率范围", 0.0, 1.0, (0.0, 1.0), key="filter_dr_slider")
            with af_col4:
                filter_type = st.multiselect("教材类型", type_opts, key="filter_type", placeholder="全部类型")

        # ── 构建查询 ──
        base_sql = "FROM textbooks_master"
        where_parts = []
        params = []

        if search_key.strip():
            where_parts.append("(name LIKE %s OR COALESCE(isbn,'') LIKE %s OR COALESCE(publisher,'') LIKE %s)")
            like_val = f"%{search_key.strip()}%"
            params.extend([like_val, like_val, like_val])

        if filter_pub:
            ph = ",".join(["%s"] * len(filter_pub))
            where_parts.append(f"publisher IN ({ph})")
            params.extend(filter_pub)

        if filter_editor:
            ph = ",".join(["%s"] * len(filter_editor))
            where_parts.append(f"editor IN ({ph})")
            params.extend(filter_editor)

        if filter_dr_mode == "有折扣(<1.0)":
            where_parts.append("COALESCE(discount_rate, 1.0) < 1.0")
        elif filter_dr_mode == "无折扣(=1.0)":
            where_parts.append("COALESCE(discount_rate, 1.0) = 1.0")

        if "filter_dr_slider" in st.session_state and filter_dr_mode == "全部":
            dr_lo, dr_hi = st.session_state.filter_dr_slider
            where_parts.append("COALESCE(discount_rate, 1.0) BETWEEN %s AND %s")
            params.extend([dr_lo, dr_hi])

        if filter_type:
            ph = ",".join(["%s"] * len(filter_type))
            where_parts.append(f"COALESCE(textbook_type,'') IN ({ph})")
            params.extend(filter_type)

        where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
        count_sql = f"SELECT COUNT(*) as cnt {base_sql} {where_sql}"
        master_df = query_df(count_sql, tuple(params) if params else None)
        total_count = int(master_df.iloc[0]["cnt"]) if not master_df.empty else 0

        master_sql = f"SELECT id, name, isbn, publisher, editor, price, publication_date, course_name, discount_rate, actual_price, textbook_type {base_sql} {where_sql} ORDER BY name"
        master_df = query_df(master_sql, tuple(params) if params else None)

        if not master_df.empty:
            st.caption(f"共 **{len(master_df)}** 本教材，引用 **{sum(usage_map.values())}** 次")

            page_size = st.session_state.get("master_ps", 50)
            total_pages = max(1, math.ceil(len(master_df) / page_size))
            mp = st.session_state.get("master_page", 1)
            if mp > total_pages:
                mp = total_pages
                st.session_state.master_page = total_pages

            st.caption(f"共 **{len(master_df)}** 本教材 ｜ 第 {mp}/{total_pages} 页")

            start = (mp - 1) * page_size
            end = min(start + page_size, len(master_df))
            page_df = master_df.iloc[start:end].copy()

            master_selall = st.session_state.get(f"master_selall_{mp}", False)
            rows = []
            real_ids = []  # 索引→真实数据库ID 的映射
            for _, mr in page_df.iterrows():
                mid = int(mr["id"])
                real_ids.append(mid)
                usg = usage_map.get(mid, 0)
                dr = mr.get("discount_rate")
                if dr is None:
                    dr_val = 1.0
                else:
                    dr_val = float(dr)
                ap = mr.get("actual_price")
                ap_val = float(ap) if ap is not None else None
                tb_type = str(mr.get("textbook_type", "") or "")
                rows.append({
                    "选择": master_selall,
                    "课程名称": str(mr.get("course_name", "") or ""),
                    "教材名称": mr["name"],
                    "书号": str(mr.get("isbn", "") or ""),
                    "主编": str(mr.get("editor", "") or ""),
                    "出版社": str(mr.get("publisher", "") or ""),
                    "出版日期": safe_str(mr.get("publication_date", "")),
                    "单价(元)": float(mr.get("price", 0) or 0),
                    "折扣率": dr_val,
                    "实洋(元)": ap_val if ap_val is not None else float(mr.get("price", 0) or 0) * dr_val,
                    "教材类型": tb_type,
                    "引用次数": usg,
                })

            display_df = pd.DataFrame(rows)
            edited = st.data_editor(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "选择": st.column_config.CheckboxColumn("选择"),
                    "课程名称": st.column_config.TextColumn("课程名称", alignment="center"),
                    "教材名称": st.column_config.TextColumn("教材名称", alignment="center"),
                    "书号": st.column_config.TextColumn("书号", alignment="center"),
                    "主编": st.column_config.TextColumn("主编", alignment="center"),
                    "出版社": st.column_config.TextColumn("出版社", alignment="center"),
                    "出版日期": st.column_config.TextColumn("出版日期", alignment="center"),
                    "单价(元)": st.column_config.NumberColumn("单价(元)", format="¥%.2f", alignment="center"),
                    "折扣率": st.column_config.NumberColumn("折扣率", format="%.2f", min_value=0.0, max_value=1.0, help="如 0.76=76折，1.0=原价", alignment="center"),
                    "实洋(元)": st.column_config.NumberColumn("实洋(元)", format="¥%.2f", help="学生结算价格，留空则=单价×折扣率", alignment="center"),
                    "教材类型": st.column_config.TextColumn("教材类型", help="如：国规教材、公共基础课、专业课、实验实训等", alignment="center"),
                    "引用次数": st.column_config.NumberColumn("引用次数", disabled=True, alignment="center"),
                },
                disabled=["引用次数"],
                key=f"master_editor_{mp}"
            )

            # ═══ 布局：上行(全选+分页) + 下行(红删蓝存) ═══
            # 预检测是否有修改
            has_changes = False
            for i, row in edited.iterrows():
                eid = real_ids[i]
                o = display_df.iloc[i]
                if (str(row["教材名称"]) != str(o["教材名称"]) or
                    str(row["书号"]) != str(o["书号"]) or
                    str(row["出版社"]) != str(o["出版社"]) or
                    str(row["主编"]) != str(o["主编"]) or
                    str(row["出版日期"]) != str(o["出版日期"]) or
                    float(row["单价(元)"]) != float(o["单价(元)"]) or
                    str(row["课程名称"]) != str(o["课程名称"]) or
                    abs(float(row["折扣率"]) - float(o["折扣率"])) > 1e-9 or
                    abs((float(row.get("实洋(元)", 0) or 0)) - (float(o.get("实洋(元)", 0) or 0))) > 1e-9 or
                    str(row.get("教材类型", "") or "") != str(o.get("教材类型", "") or "")):
                    has_changes = True
                    break

            # 处理按钮触发的操作（统一处理所有 pending_action）
            pending = st.session_state.get("pending_action", "")
            if pending == "master_save":
                if has_changes:
                    changes = 0
                    for i, row in edited.iterrows():
                        eid = real_ids[i]
                        o = display_df.iloc[i]
                        new_name = str(row["教材名称"]).strip()
                        if not new_name: continue
                        # 实洋价格变化检测
                        ap_new = row.get("实洋(元)")
                        ap_old = o.get("实洋(元)")
                        ap_diff = False
                        if (ap_new is None) != (ap_old is None):
                            ap_diff = True
                        elif ap_new is not None:
                            try:
                                if abs(float(ap_new) - float(ap_old)) > 1e-9:
                                    ap_diff = True
                            except (ValueError, TypeError):
                                ap_diff = True

                        if (str(row["教材名称"]) != str(o["教材名称"]) or
                            str(row["书号"]) != str(o["书号"]) or
                            str(row["出版社"]) != str(o["出版社"]) or
                            str(row["主编"]) != str(o["主编"]) or
                            str(row["出版日期"]) != str(o["出版日期"]) or
                            abs(float(row["单价(元)"]) - float(o["单价(元)"])) > 1e-9 or
                            str(row["课程名称"]) != str(o["课程名称"]) or
                            abs(float(row["折扣率"]) - float(o["折扣率"])) > 1e-9 or
                            ap_diff or
                            str(row.get("教材类型", "") or "") != str(o.get("教材类型", "") or "")):
                            # actual_price: None → NULL; 0 → 0.0; other → value
                            ap_val = row.get("实洋(元)")
                            ap_sql = None if ap_val is None else float(ap_val)
                            execute_sql(
                                "UPDATE textbooks_master SET name=%s, isbn=%s, publisher=%s, editor=%s, publication_date=%s, price=%s, course_name=%s, discount_rate=%s, actual_price=%s, textbook_type=%s WHERE id=%s",
                                (new_name, str(row["书号"]).strip(), str(row["出版社"]).strip(),
                                 str(row["主编"]).strip(), str(row["出版日期"]).strip(),
                                 float(row["单价(元)"]), str(row["课程名称"]).strip(),
                                 float(row["折扣率"]),
                                 ap_sql,
                                 str(row.get("教材类型", "") or "").strip() or None,
                                 eid))
                            changes += 1
                    if changes > 0:
                        st.toast(f"✅ 已保存 {changes} 条修改", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()
            elif pending == "master_del":
                sl_checked = (edited["选择"].sum() if "选择" in edited.columns else 0)
                select_all = st.session_state.get(f"master_selall_{mp}", False)
                if select_all:
                    del_targets = page_df
                elif sl_checked > 0:
                    del_targets = page_df.iloc[edited[edited["选择"] == True].index]
                else:
                    del_targets = pd.DataFrame()
                del_ids = []
                del_names = []
                blocked = []
                for _, drow in del_targets.iterrows():
                    eid = int(drow["id"])
                    if usage_map.get(eid, 0) == 0:
                        del_ids.append(eid)
                        del_names.append(str(drow["name"]))
                    else:
                        blocked.append(f"「{drow['name']}」(被{usage_map.get(eid,0)}条引用)")
                if blocked:
                    for b in blocked:
                        st.warning(f"{b}，无法删除")
                if del_ids:
                    placeholders = ",".join(["%s"] * len(del_ids))
                    execute_sql(f"DELETE FROM textbooks_master WHERE id IN ({placeholders})", tuple(del_ids))
                    st.toast(f"✅ 已删除 {len(del_ids)} 本", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()
            elif pending == "batch_dr":
                sel_ids = [real_ids[i] for i in range(len(edited)) if edited.iloc[i]["选择"]]
                if sel_ids:
                    dr_val = st.session_state.get("batch_dr_widget", 0.76)
                    placeholders = ",".join(["%s"] * len(sel_ids))
                    execute_sql(f"UPDATE textbooks_master SET discount_rate=%s WHERE id IN ({placeholders})",
                                (dr_val,) + tuple(sel_ids))
                    st.toast(f"✅ 已批量修改 {len(sel_ids)} 本教材折扣率为 {dr_val:.2f}", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()

            # ── 上行：全选 + 分页 ──
            r1_sel, r1_info, r1_ps, r1_prev, r1_num, r1_next = st.columns([1.5, 2, 1, 0.7, 0.7, 0.7])
            with r1_sel:
                select_all = st.checkbox("全选本页", key=f"master_selall_{mp}",
                    help="勾选后删除操作将应用于本页全部记录")
            with r1_info:
                st.caption(f"共 **{len(master_df)}** 本")
            with r1_ps:
                st.selectbox("每页", [50, 100, 200], key="master_ps",
                             on_change=lambda: st.session_state.update({"master_page": 1}),
                             label_visibility="collapsed")
            with r1_prev:
                if st.button("◀", key="master_prev", disabled=(mp <= 1), use_container_width=True):
                    st.session_state["master_page"] = max(1, mp - 1); st.rerun()
            with r1_num:
                st.markdown(f"<div style='text-align:center;padding-top:5px;font-weight:500'>{mp}/{total_pages}</div>", unsafe_allow_html=True)
            with r1_next:
                if st.button("▶", key="master_next", disabled=(mp >= total_pages), use_container_width=True):
                    st.session_state["master_page"] = min(total_pages, mp + 1); st.rerun()

            # ── 下行：删除(红) + 保存(蓝) ──
            sl_checked = (edited["选择"].sum() if not select_all and "选择" in edited.columns else 0)
            del_count = len(page_df) if select_all else int(sl_checked)
            del_disabled = "disabled" if del_count == 0 else ""
            save_disabled = "disabled" if not has_changes else ""
            del_btn_label = f"🗑️ 删除（{del_count}本）"
            # 统计确切的修改数量
            chg_count = 0
            if has_changes:
                for i, row in edited.iterrows():
                    eid = real_ids[i]
                    o = display_df.iloc[i]
                    if (str(row["教材名称"]) != str(o["教材名称"]) or
                        str(row["书号"]) != str(o["书号"]) or
                        str(row["出版社"]) != str(o["出版社"]) or
                        str(row["主编"]) != str(o["主编"]) or
                        str(row["出版日期"]) != str(o["出版日期"]) or
                        float(row["单价(元)"]) != float(o["单价(元)"]) or
                        str(row["课程名称"]) != str(o["课程名称"]) or
                        abs(float(row["折扣率"]) - float(o["折扣率"])) > 1e-9 or
                        abs((float(row.get("实洋(元)", 0) or 0)) - (float(o.get("实洋(元)", 0) or 0))) > 1e-9 or
                        str(row.get("教材类型", "") or "") != str(o.get("教材类型", "") or "")):
                        chg_count += 1
            save_btn_label = f"💾 保存修改（{chg_count}处）" if has_changes else "💾 保存修改"

            r2_del, r2_save = st.columns([1, 1])
            with r2_del:
                st.button(del_btn_label, key=f"master_del_btn_{mp}", type="secondary",
                          disabled=(del_count == 0), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "master_del"}))
            with r2_save:
                st.button(save_btn_label, key=f"master_save_btn_{mp}", type="primary",
                          disabled=(not has_changes), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "master_save"}))
        else:
            st.info("📭 教材库为空，请通过「新增教材」或「导入 Excel」添加")

    # ═══════════════ Tab2: 新增教材 ═══════════════
    with tab2:
        st.markdown("#### ➕ 手动新增教材")

        with st.form("master_add_form", clear_on_submit=True):
            ac1, ac2 = st.columns(2)
            with ac1:
                new_name = st.text_input("教材名称 *", placeholder="必填，如：高等数学（第七版）", key="new_master_name")
                new_isbn = st.text_input("书号(ISBN)", placeholder="978-7-04-059109-9", key="new_master_isbn")
                new_publisher = st.text_input("出版社", placeholder="高等教育出版社", key="new_master_publisher")
                new_editor = st.text_input("主编", placeholder="同济大学数学科学学院", key="new_master_editor")
                new_course = st.text_input("课程名称", placeholder="如：高等数学A", key="new_master_course")
            with ac2:
                new_price = st.number_input("单价(元)", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="new_master_price")
                new_dr = st.number_input("折扣率", min_value=0.0, max_value=1.0,
                    value=st.session_state.get("master_default_dr", 0.76),
                    step=0.01, format="%.2f", key="new_master_dr",
                    help="如 0.76=76折，1.0=原价")
                new_ap = st.number_input("实洋(元)", min_value=0.0, value=None, step=0.01, format="%.2f",
                    key="new_master_ap", placeholder="留空=自动计算",
                    help="学生结算价格，留空则按 单价×折扣率 自动计算")
                new_pub_date = st.text_input("出版日期", placeholder="如：2023年08月", key="new_master_pubdate")
                new_type = st.text_input("教材类型", placeholder="如：国规教材、专业课、公共基础课", key="new_master_type")

            submitted = st.form_submit_button("✅ 确认添加", type="primary", use_container_width=True)

        if submitted:
            if not new_name.strip():
                st.error("❌ 教材名称为必填项")
            else:
                # 查重
                exist_check = query_df("SELECT id FROM textbooks_master WHERE name=%s", (new_name.strip(),))
                if not exist_check.empty:
                    st.error(f"❌ 教材「{new_name.strip()}」已存在（ID:{int(exist_check.iloc[0]['id'])}），请使用其他名称或前往列表编辑")
                else:
                    execute_sql(
                        "INSERT INTO textbooks_master (name, isbn, publisher, editor, price, publication_date, course_name, discount_rate, actual_price, textbook_type) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (new_name.strip(),
                         new_isbn.strip() or None,
                         new_publisher.strip() or None,
                         new_editor.strip() or None,
                         new_price,
                         new_pub_date.strip() or None,
                         new_course.strip() or None,
                         new_dr,
                         new_ap if new_ap is not None and new_ap > 0 else None,
                         new_type.strip() or None))
                    st.toast(f"✅ 教材「{new_name.strip()}」已添加", icon="✅")
                    st.rerun()

    # ═══════════════ Tab3: 导入 Excel ═══════════════
    with tab3:
        st.markdown("#### 📥 从 Excel 导入教材")

        # 下载模板
        template_cols = ["课程名称", "教材名称", "书号(ISBN)", "主编", "出版社", "出版日期", "单价(元)", "折扣率", "实洋(元)", "教材类型"]
        template_df = make_template_df(template_cols)
        template_bytes = excel_export(template_df, "教材表导入模板")
        st.download_button(
            "📄 下载导入模板",
            data=template_bytes,
            file_name="教材表导入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="secondary"
        )
        st.caption("Excel 表头建议包含：课程名称、教材名称、书号(ISBN)、主编、出版社、出版日期、单价(元)、折扣率、实洋(元)、教材类型")

        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="master_import_upload")
        if uploaded:
            try:
                raw_df = read_excel_upload(uploaded)
                st.info(f"📄 检测到 {len(raw_df)} 行，列名：{', '.join(raw_df.columns[:10])}")

                # 列映射
                col_map = {}
                for col in raw_df.columns:
                    cl = col.strip()
                    cl_lower = cl.lower().replace(" ", "").replace(" ", "")
                    # 先匹配课程名称（必须放在"名称"之前，否则"课程名称"会被误映射到name）
                    if "课程名称" in cl or "课程" in cl or "科目" in cl or "course" in cl_lower:
                        col_map["course_name"] = cl
                    elif "教材名称" in cl or "书名" in cl or "name" in cl_lower or ("名称" in cl and "课程" not in cl and "折扣" not in cl):
                        col_map["name"] = cl
                    elif "isbn" in cl_lower or "书号" in cl:
                        col_map["isbn"] = cl
                    elif "出版社" in cl or "press" in cl_lower or "publisher" in cl_lower:
                        col_map["publisher"] = cl
                    elif "主编" in cl or "作者" in cl or "编者" in cl or "editor" in cl_lower or "author" in cl_lower:
                        col_map["editor"] = cl
                    elif "单价" in cl or "价格" in cl or "定价" in cl or "price" in cl_lower:
                        col_map["price"] = cl
                    elif "出版日期" in cl or "出版时间" in cl or "出版年" in cl or "pub_date" in cl_lower or "publication_date" in cl_lower or "publish_date" in cl_lower:
                        col_map["publication_date"] = cl
                    elif "实洋" in cl or "实洋" in cl or "actual_price" in cl_lower:
                        col_map["actual_price"] = cl
                    elif "教材类型" in cl or "类型" in cl or "type" in cl_lower:
                        col_map["textbook_type"] = cl
                    elif "国规" in cl or "规划教材" in cl or "national" in cl_lower:
                        col_map["textbook_type"] = cl  # 旧字段也映射到 textbook_type
                    elif "折扣" in cl or "discount" in cl_lower:
                        col_map["discount_rate"] = cl

                if "name" not in col_map:
                    st.error("❌ 缺少必要列：教材名称")
                    st.json(col_map)
                else:
                    # ── 列映射（做完再预览，出版日期也会格式化） ──
                    mapped = raw_df.rename(columns={v: k for k, v in col_map.items()})
                    wanted = [c for c in ["name", "isbn", "publisher", "editor", "price", "publication_date", "course_name", "discount_rate", "actual_price", "textbook_type"] if c in mapped.columns]
                    mapped = mapped[wanted].where(pd.notnull(mapped), None)

                    # ── 出版日期统一转 "YYYY年MM月"（Excel 序列号 / datetime 均支持） ──
                    if "publication_date" in mapped.columns:
                        from datetime import datetime as _dt, timedelta as _td
                        def _fmt_pubdate(v):
                            if v is None or (isinstance(v, float) and pd.isna(v)):
                                return ""
                            if isinstance(v, (int, float)) and v > 1000:
                                try:
                                    d = _dt(1899, 12, 30) + _td(days=int(v))
                                    return f"{d.year}年{d.month:02d}月"
                                except:
                                    pass
                            if hasattr(v, "strftime"):
                                return f"{v.year}年{v.month:02d}月"
                            s = str(v).strip()
                            # 已经是 "年/月" 或 "年-月" 格式，统一为 "年 月"
                            import re as _re
                            m = _re.match(r"(\d{4})[年/\-](\d{1,2})", s)
                            if m:
                                return f"{m.group(1)}年{int(m.group(2)):02d}月"
                            return s
                        mapped["publication_date"] = mapped["publication_date"].apply(_fmt_pubdate)

                    # ── 折扣率：Excel 有则用，无则用页面默认值 ──
                    default_dr = st.session_state.get("master_default_dr", 0.76)
                    if "discount_rate" not in mapped.columns:
                        mapped["discount_rate"] = default_dr
                    else:
                        mapped["discount_rate"] = mapped["discount_rate"].apply(
                            lambda v: float(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else default_dr)

                    total_rows = len(mapped)

                    # 预览（日期已格式化，不再是序列号）
                    preview = mapped.head(5)
                    st.dataframe(preview.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

                    # 收集已存在的教材名称（一次性查，不在循环里逐条查）
                    book_names = [safe_str(r.get("name", "")) for _, r in mapped.iterrows() if safe_str(r.get("name", ""))]
                    unique_names = list(set(book_names))
                    if unique_names:
                        placeholders = ",".join(["%s"] * len(unique_names))
                        exist_df = query_df(f"SELECT name, id FROM textbooks_master WHERE name IN ({placeholders})", tuple(unique_names))
                        exist_set = set(exist_df["name"].tolist()) if not exist_df.empty else set()
                        exist_detail = {}
                        for _, er in exist_df.iterrows():
                            exist_detail[er["name"]] = int(er["id"])
                    else:
                        exist_set = set()
                        exist_detail = {}

                    # 预扫描统计（也处理 Excel 内部重复）
                    seen_names = set()
                    pre_new = 0
                    pre_update = 0
                    pre_skip = 0
                    pre_dup = 0  # Excel 内部重名
                    pre_update_names = []
                    pre_dup_names = []
                    for _, row in mapped.iterrows():
                        bn = safe_str(row.get("name", ""))
                        if not bn:
                            pre_skip += 1
                            continue
                        if bn in seen_names:
                            pre_dup += 1
                            pre_dup_names.append(bn)
                            continue
                        seen_names.add(bn)
                        if bn in exist_set:
                            pre_update += 1
                            pre_update_names.append(bn)
                        else:
                            pre_new += 1

                    # 预览面板
                    st.markdown("---")
                    st.markdown("#### 📊 导入预览")
                    c_a, c_b, c_c, c_d = st.columns(4)
                    c_a.metric("✅ 将新增", pre_new)
                    c_b.metric("🔄 将更新", pre_update)
                    c_c.metric("⏭️ 跳过（空名）", pre_skip)
                    c_d.metric("⚠️ 重名（去重）", pre_dup)
                    if pre_update > 0 and pre_update_names:
                        with st.expander(f"📋 查看将被更新的 {pre_update} 本教材（数据库中已存在）"):
                            for nm in sorted(pre_update_names):
                                st.caption(f"• {nm}")
                    if pre_dup > 0 and pre_dup_names:
                        with st.expander(f"⚠️ Excel 内部重复 {pre_dup} 行（仅保留首次出现）"):
                            for nm in sorted(set(pre_dup_names)):
                                st.caption(f"• {nm}")

                    if st.button("✅ 确认导入", use_container_width=True, type="primary", key="master_import_btn"):
                        # 去重：名称相同则覆盖（UPDATE），否则 INSERT
                        # Excel 内部也去重（仅保留首次出现的书名）
                        success_count = 0
                        update_count = 0
                        skip_count = 0
                        dup_skip = 0
                        errors = []
                        imported_names = set()
                        progress = st.progress(0, text="正在导入...")
                        for i, (_, row) in enumerate(mapped.iterrows()):
                            try:
                                book_name = safe_str(row.get("name", ""))
                                if not book_name:
                                    skip_count += 1
                                    continue
                                # Excel 内部去重：同名只保留首次出现
                                if book_name in imported_names:
                                    dup_skip += 1
                                    continue
                                imported_names.add(book_name)
                                # 查重（用预扫描结果加速）
                                if book_name in exist_detail:
                                    eid = exist_detail[book_name]
                                    execute_sql(
                                        "UPDATE textbooks_master SET isbn=%s, publisher=%s, editor=%s, price=%s, publication_date=%s, course_name=%s, discount_rate=%s, actual_price=%s, textbook_type=%s WHERE id=%s",
                                        (safe_str(row.get("isbn")), safe_str(row.get("publisher")),
                                         safe_str(row.get("editor")), safe_float(row.get("price"), 0),
                                         safe_str(row.get("publication_date")),
                                         safe_str(row.get("course_name")),
                                         float(row["discount_rate"]) if "discount_rate" in row.index else default_dr,
                                         safe_float(row.get("actual_price")) if row.get("actual_price") is not None and not (isinstance(row.get("actual_price"), float) and pd.isna(row.get("actual_price"))) else None,
                                         safe_str(row.get("textbook_type")) or None,
                                         eid))
                                    update_count += 1
                                else:
                                    execute_sql(
                                        "INSERT INTO textbooks_master (name, isbn, publisher, editor, price, publication_date, course_name, discount_rate, actual_price, textbook_type) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                        (book_name, safe_str(row.get("isbn")), safe_str(row.get("publisher")),
                                         safe_str(row.get("editor")), safe_float(row.get("price"), 0),
                                         safe_str(row.get("publication_date")),
                                         safe_str(row.get("course_name")),
                                         float(row["discount_rate"]) if "discount_rate" in row.index else default_dr,
                                         safe_float(row.get("actual_price")) if row.get("actual_price") is not None and not (isinstance(row.get("actual_price"), float) and pd.isna(row.get("actual_price"))) else None,
                                         safe_str(row.get("textbook_type")) or None))
                                    success_count += 1
                            except Exception as e:
                                errors.append(f"{row.get('name','?')}: {str(e)[:80]}")
                            progress.progress((i + 1) / total_rows, text=f"已处理 {i+1}/{total_rows}")

                        # 写日志
                        write_import_log(
                            module="教材表管理",
                            filename=uploaded.name,
                            total=total_rows,
                            success=success_count,
                            errors=errors
                        )

                        # 展示结果
                        mc_ok, mc_upd, mc_err, mc_dp = st.columns(4)
                        mc_ok.metric("✅ 新增", success_count)
                        mc_upd.metric("🔄 更新", update_count)
                        mc_err.metric("⚠️ 失败", len(errors))
                        mc_dp.metric("🔁 去重跳过", dup_skip)
                        if errors:
                            with st.expander(f"查看 {len(errors)} 条失败详情"):
                                for err in errors:
                                    st.caption(f"• {err}")
                        if skip_count > 0:
                            st.caption(f"⏭️ 跳过 {skip_count} 行（教材名称为空）")
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")


# ═════════════════════════════════════════════════════════
# 4. 教材征订表（按班级明细）
# ═════════════════════════════════════════════════════════

def textbook_management():
    """教材征订表管理（双表架构：textbooks_master + textbook_orders）"""
    import sqlite3
    show_header("📖 教材征订表", "按班级管理征订明细（教材请前往「教材表管理」维护）")

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先在「学期管理」中添加学期")
        return

    semester_options = ["全部"] + [f"{r['id']}|{r['name']}" for _, r in semesters.iterrows()]
    master_books = query_df("SELECT id, name, isbn, publisher, editor, price, course_name FROM textbooks_master ORDER BY name")
    master_options = [(0, "➕ 新增教材...")] + [(r["id"], r["name"]) for _, r in master_books.iterrows()]

    tab1, tab2, tab3, tab4 = st.tabs(["📋 征订列表", "➕ 新增/编辑征订", "📥 导入 Excel", "📚 教材批量分配"])

    with tab1:
        col1, col2, col3, col4 = st.columns(4)
        with col1: f_sem = st.selectbox("学期", semester_options, key="t_sem")
        with col2: f_tcollege = st.selectbox("学院", ["全部"] + get_filtered_colleges(), key="t_college")
        with col3:
            # 专业：基于所选学院级联过滤
            if f_tcollege != "全部":
                t_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (f_tcollege,))
            else:
                t_major_opts = ["全部"] + get_filtered_majors()
            f_tmajor = st.selectbox("专业", t_major_opts, key="t_major")
        with col4:
            # 班级：基于所选学院+专业级联过滤
            t_class_where = "1=1"; t_class_params = []
            if f_tcollege != "全部":
                t_class_where += " AND college = %s"; t_class_params.append(f_tcollege)
            if f_tmajor != "全部":
                t_class_where += " AND major = %s"; t_class_params.append(f_tmajor)
            t_class_opts = ["全部"] + (get_filtered_list("students", "class_name", t_class_where, tuple(t_class_params)) if t_class_params else get_filtered_class_names())
            f_tclass = st.selectbox("班级", t_class_opts, key="t_class")
        f_tsearch = st.text_input("🔍 搜索（教材名）", key="t_search", placeholder="输入教材名称...")

        sql = """SELECT o.*, tm.name, tm.isbn, tm.publisher, tm.editor, tm.price, tm.course_name,
                        s.name as semester_name
                 FROM textbook_orders o
                 JOIN textbooks_master tm ON o.textbook_id = tm.id
                 JOIN semesters s ON o.semester_id = s.id WHERE 1=1"""
        params = []
        if f_sem != "全部": sql += " AND o.semester_id = %s"; params.append(int(f_sem.split("|")[0]))
        if f_tcollege != "全部": sql += " AND o.college = %s"; params.append(f_tcollege)
        if f_tmajor != "全部": sql += " AND o.major = %s"; params.append(f_tmajor)
        if f_tclass != "全部": sql += " AND o.class_name = %s"; params.append(f_tclass)
        if f_tsearch: sql += " AND tm.name LIKE %s"; params.append(f"%{f_tsearch}%")
        sql += " ORDER BY o.semester_id DESC, o.college, o.major, o.class_name, tm.name"

        df = query_df(sql, tuple(params) if params else None)
        if not df.empty:
            df["小计"] = pd.to_numeric(df["price"], errors="coerce").fillna(0) * pd.to_numeric(df["quantity"], errors="coerce").fillna(0)
            total_amount = df["小计"].sum()

            # 分页（从 session_state 读取，控件移到表格下方）
            import math
            t_ps = st.session_state.get("t_ps", 30)
            t_pg = st.session_state.get("t_pg", 1)
            tp = max(1, math.ceil(len(df) / t_ps))
            # 页码保护：删除后当前页可能越界，自动跳到最后一页
            if t_pg > tp:
                t_pg = tp
                st.session_state.t_pg = tp

            start = (t_pg - 1) * t_ps
            end = min(start + t_ps, len(df))
            page_df = df.iloc[start:end].copy()

            st.caption(f"共 **{len(df)}** 条征订记录 ｜ 第 {t_pg}/{tp} 页 ｜ 💰 总计 ¥{total_amount:,.2f}")

            cols = ["id","semester_name","college","major","class_name","grade",
                    "name","course_name","isbn","publisher","editor","price","quantity","remark"]
            rename_map = {"semester_name":"学期","college":"学院","major":"专业","class_name":"班级","grade":"年级",
                "name":"教材名称","course_name":"课程","isbn":"书号","publisher":"出版社","editor":"主编",
                "price":"单价(元)","quantity":"征订数量","remark":"备注"}

            display_df = page_df[cols].rename(columns=rename_map).copy()
            t_selall = st.session_state.get(f"t_sel_{t_pg}", False)
            display_df["选择"] = t_selall
            ed_k = f"tb_ed_{t_pg}"
            edited = st.data_editor(
                display_df,
                use_container_width=True, hide_index=True,
                column_order=["选择","学期","学院","专业","班级","年级","教材名称","课程","书号","出版社","主编","单价(元)","征订数量","备注"],
                column_config={
                    "选择": st.column_config.CheckboxColumn("选择"),
                    "学期": st.column_config.TextColumn("学期", disabled=True, alignment="center"),
                    "学院": st.column_config.TextColumn("学院", disabled=True, alignment="center"),
                    "专业": st.column_config.TextColumn("专业", disabled=True, alignment="center"),
                    "班级": st.column_config.TextColumn("班级", disabled=True, alignment="center"),
                    "年级": st.column_config.TextColumn("年级", disabled=True, alignment="center"),
                    "教材名称": st.column_config.TextColumn("教材名称", disabled=True, alignment="center"),
                    "课程": st.column_config.TextColumn("课程", disabled=True, alignment="center"),
                    "书号": st.column_config.TextColumn("书号", disabled=True, alignment="center"),
                    "出版社": st.column_config.TextColumn("出版社", disabled=True, alignment="center"),
                    "主编": st.column_config.TextColumn("主编", disabled=True, alignment="center"),
                    "单价(元)": st.column_config.NumberColumn("单价(元)", format="¥%.2f", disabled=True, alignment="center"),
                    "征订数量": st.column_config.NumberColumn("征订数量", min_value=0, step=1, alignment="center"),
                    "备注": st.column_config.TextColumn("备注", alignment="center"),
                },
                disabled=["学期","学院","专业","班级","年级","教材名称","课程","书号","出版社","主编"],
                key=ed_k
            )

            # ═══ 布局：上行(全选+分页) + 下行(红删蓝存) ═══
            # 检测数量+备注修改
            t_changes = []
            for i in range(len(edited)):
                rid = page_df.iloc[i]["id"]
                old_q = int(page_df.iloc[i].get("quantity", 0))
                new_q = int(edited.iloc[i].get("征订数量", 0))
                old_r = str(page_df.iloc[i].get("remark", "") or "")
                new_r = str(edited.iloc[i].get("备注", "") or "")
                if old_q != new_q:
                    t_changes.append((rid, "quantity", new_q))
                if old_r != new_r:
                    t_changes.append((rid, "remark", new_r))

            # 处理按钮触发的操作（用 session_state 代替 query_params）
            pending = st.session_state.get("pending_action", "")
            if pending == "t_save":
                if t_changes:
                    for rid, col, val in t_changes:
                        execute_sql(f"UPDATE textbook_orders SET {col}=%s WHERE id=%s", (val, rid))
                    st.toast(f"✅ 已保存 {len(t_changes)} 处修改", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()
            elif pending == "t_del":
                t_sel_all = st.session_state.get(f"t_sel_{t_pg}", False)
                if t_sel_all:
                    del_ids = tuple(page_df["id"].tolist())
                else:
                    selected = edited[edited["选择"] == True].index if "选择" in edited.columns else []
                    del_ids = tuple(page_df.iloc[selected]["id"].tolist()) if len(selected) > 0 else ()
                if del_ids:
                    if len(del_ids) == 1:
                        execute_sql("DELETE FROM textbook_orders WHERE id=%s", (del_ids[0],))
                    else:
                        ph = ",".join(["%s"] * len(del_ids))
                        execute_sql(f"DELETE FROM textbook_orders WHERE id IN ({ph})", del_ids)
                    st.toast(f"✅ 已删除 {len(del_ids)} 条", icon="✅")
                st.session_state.pending_action = ""
                st.rerun()

            # ── 上行：全选 + 分页 ──
            r1_sel, r1_info, r1_ps, r1_prev, r1_num, r1_next = st.columns([1.5, 2, 1, 0.7, 0.7, 0.7])
            with r1_sel:
                t_sel_all = st.checkbox("全选本页", key=f"t_sel_{t_pg}",
                    help="勾选后删除操作将应用于本页全部记录")
            with r1_info:
                st.caption(f"共 **{len(df)}** 条 ｜ ¥{total_amount:,.2f}")
            with r1_ps:
                st.selectbox("每页", [30, 50, 100], key="t_ps",
                             on_change=lambda: st.session_state.update({"t_pg": 1}),
                             label_visibility="collapsed")
            with r1_prev:
                if st.button("◀", key="t_pp", disabled=(t_pg <= 1), use_container_width=True):
                    st.session_state.t_pg = t_pg - 1; st.rerun()
            with r1_num:
                st.markdown(f"<div style='text-align:center;padding-top:5px;font-weight:500'>{t_pg}/{tp}</div>", unsafe_allow_html=True)
            with r1_next:
                if st.button("▶", key="t_np", disabled=(t_pg >= tp), use_container_width=True):
                    st.session_state.t_pg = t_pg + 1; st.rerun()

            # ── 下行：删除(红) + 保存(蓝) ──
            del_count = len(page_df) if t_sel_all else (int(edited["选择"].sum()) if "选择" in edited.columns else 0)
            del_disabled = "disabled" if del_count == 0 else ""
            save_disabled = "disabled" if not t_changes else ""
            del_btn_label = f"🗑️ 删除（{del_count}条）"
            save_btn_label = f"💾 保存修改（{len(t_changes)}处）" if t_changes else "💾 保存修改"

            r2_del, r2_save = st.columns([1, 1])
            with r2_del:
                st.button(del_btn_label, key=f"t_del_btn_{t_pg}", type="secondary",
                          disabled=(del_count == 0), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "t_del"}))
            with r2_save:
                st.button(save_btn_label, key=f"t_save_btn_{t_pg}", type="primary",
                          disabled=(not t_changes), use_container_width=True,
                          on_click=lambda: st.session_state.update({"pending_action": "t_save"}))

        else:
            st.info("暂无征订数据")

    with tab2:
        st.markdown("#### ➕ 新增征订 / 管理教材库")
        edit_oid = st.number_input("编辑征订ID（留空为新增）", min_value=0, value=0, step=1, key="edit_oid")
        defaults = {}
        if edit_oid > 0:
            row = query_df("SELECT o.*, tm.name as book_name FROM textbook_orders o JOIN textbooks_master tm ON o.textbook_id=tm.id WHERE o.id=%s", (edit_oid,))
            if not row.empty: defaults = row.iloc[0].to_dict()

        # 教材选择放在表单外面，触发页面重绘
        st.caption("🔹 第一步：选择教材（从教材库选择，或选「新增教材」手动输入）")
        default_master_idx = 0 if edit_oid == 0 else next((i for i, o in enumerate(master_options) if o[0] == defaults.get("textbook_id", 0)), 0)
        picked = st.selectbox("选择已有教材", master_options,
            format_func=lambda x: x[1], index=default_master_idx, key="sel_master_out")
        is_new_book = (picked[0] == 0)

        # ── 联动：选中教材时自动填充信息（仅在选择变化时更新）──
        prev_book_id2 = st.session_state.get("_bk_prev_book_id", -1)
        if picked[0] != prev_book_id2:
            st.session_state["_bk_prev_book_id"] = picked[0]
            if not is_new_book:
                bk_info = master_books[master_books["id"] == picked[0]]
                if not bk_info.empty:
                    r = bk_info.iloc[0]
                    st.session_state["bk_name"] = r["name"]
                    st.session_state["bk_isbn"] = str(r.get("isbn", "") or "")
                    st.session_state["bk_pub"] = str(r.get("publisher", "") or "")
                    st.session_state["bk_ed"] = str(r.get("editor", "") or "")
                    st.session_state["bk_price"] = float(r.get("price", 0))
                    st.session_state["bk_course"] = str(r.get("course_name", "") or "")
            else:
                for k in ["bk_name", "bk_isbn", "bk_pub", "bk_ed", "bk_course"]:
                    st.session_state[k] = ""
                st.session_state["bk_price"] = 0.0

        st.caption("🔹 第二步：教材信息（从教材库自动带出，也可修改）")
        cols_book = st.columns(3)
        with cols_book[0]:
            bk_name = st.text_input("教材名称*", placeholder="必填", key="bk_name")
            bk_isbn = st.text_input("书号(ISBN)", key="bk_isbn")
        with cols_book[1]:
            bk_publisher = st.text_input("出版社", key="bk_pub")
            bk_editor = st.text_input("主编", key="bk_ed")
        with cols_book[2]:
            bk_price = st.number_input("单价(元)*", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="bk_price")
            bk_course = st.text_input("课程", key="bk_course")

        st.caption("🔹 第三步：选择征订范围")
        sem_id = st.selectbox("学期*", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
            format_func=lambda x: x[1], index=next((i for i, s in enumerate([(r["id"], r["name"]) for _, r in semesters.iterrows()]) if s[0] == defaults.get("semester_id")), 0))
        sel_grades = st.multiselect("年级（可多选）", options=get_filtered_grades(), default=[defaults.get("grade")] if defaults.get("grade") else [])

        # 学院：基于所选年级级联过滤
        if sel_grades:
            grade_where = " OR ".join(["grade = %s"] * len(sel_grades))
            college_opts = get_filtered_list("students", "college", grade_where, tuple(sel_grades))
        else:
            college_opts = get_filtered_colleges()
        sel_colleges = st.multiselect("学院（可多选）", options=college_opts, default=[defaults.get("college")] if defaults.get("college") else [])

        # 专业：基于所选年级+学院级联过滤
        major_where = "1=1"; major_params = []
        if sel_grades:
            major_where += " AND (" + " OR ".join(["grade = %s"] * len(sel_grades)) + ")"
            major_params.extend(sel_grades)
        if sel_colleges:
            major_where += " AND (" + " OR ".join(["college = %s"] * len(sel_colleges)) + ")"
            major_params.extend(sel_colleges)
        major_opts = get_filtered_list("students", "major", major_where, tuple(major_params)) if major_params else get_filtered_majors()
        sel_majors = st.multiselect("专业（可多选）", options=major_opts, default=[defaults.get("major")] if defaults.get("major") else [])

        # 班级：基于所选年级+学院+专业级联过滤
        class_where = "1=1"; class_params = []
        if sel_grades:
            class_where += " AND (" + " OR ".join(["grade = %s"] * len(sel_grades)) + ")"
            class_params.extend(sel_grades)
        if sel_colleges:
            class_where += " AND (" + " OR ".join(["college = %s"] * len(sel_colleges)) + ")"
            class_params.extend(sel_colleges)
        if sel_majors:
            class_where += " AND (" + " OR ".join(["major = %s"] * len(sel_majors)) + ")"
            class_params.extend(sel_majors)
        class_opts = get_filtered_list("students", "class_name", class_where, tuple(class_params)) if class_params else get_filtered_class_names()

        sel_classes = st.multiselect("班级（可多选）", options=class_opts, default=[defaults.get("class_name")] if defaults.get("class_name") else [])

        # 显示各班级实际人数，支持按人数自动计算征订数量
        use_auto_qty = st.checkbox("📊 按各班级实际人数自动计算征订数量", value=False, key="auto_qty")
        if use_auto_qty and sel_classes:
            # 获取所选班级的实际人数分布
            auto_where = "1=1"; auto_params = []
            if sel_grades:
                auto_where += " AND (" + " OR ".join(["grade = %s"] * len(sel_grades)) + ")"
                auto_params.extend(sel_grades)
            if sel_colleges:
                auto_where += " AND (" + " OR ".join(["college = %s"] * len(sel_colleges)) + ")"
                auto_params.extend(sel_colleges)
            if sel_majors:
                auto_where += " AND (" + " OR ".join(["major = %s"] * len(sel_majors)) + ")"
                auto_params.extend(sel_majors)
            cls = list(set(sel_classes))
            ph = ",".join(["%s"] * len(cls))
            auto_where += f" AND class_name IN ({ph})"; auto_params.extend(cls)
            class_dist = get_filtered_list("students", "class_name", auto_where, tuple(auto_params)) if auto_params else sel_classes
            # 获取各班的实际学生数
            cnt_df = get_class_student_counts(class_names=sel_classes)
            if not cnt_df.empty:
                total_actual_students = int(cnt_df["student_count"].sum())
                auto_teacher = st.number_input("👨‍🏫 教师用书数量", min_value=0, value=0, step=1, key="auto_teacher")
                total_for_all = total_actual_students + auto_teacher
                # 按班级人数分配（每人1本）
                splits = [(row["class_name"], int(row["student_count"]), int(row["student_count"]))
                          for _, row in cnt_df.iterrows()]
                st.info(f"👥 所选 **{len(splits)}** 个班共 **{total_actual_students}** 名学生，合计征订 **{total_for_all}** 本")
                # 显示各班明细
                summary_rows = []
                for cn, sc, alloc in splits:
                    summary_rows.append({"班级": cn, "学生人数": sc, "分配数量": alloc})
                summary_df = pd.DataFrame(summary_rows)
                st.dataframe(summary_df.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True,
                             column_config={"分配数量": st.column_config.NumberColumn("分配数量")})
                o_qty = st.number_input("征订数量（每班）", min_value=0, value=total_for_all, step=1,
                                         help="系统自动按各班人数计算（每人1本）")
            else:
                o_qty = st.number_input("征订数量", min_value=0, value=int(defaults.get("quantity", 0)), step=1)
        else:
            o_qty = st.number_input("征订数量", min_value=0, value=int(defaults.get("quantity", 0)), step=1)
        o_remark = st.text_input("备注", value=defaults.get("remark", ""))

        combos = max(len(sel_grades) or 1, len(sel_colleges) or 1, len(sel_majors) or 1, len(sel_classes) or 1)
        if edit_oid == 0 and combos > 1:
            st.info(f"📌 将新增 **{combos}** 条征订记录")

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1: submitted = st.button("💾 保存", use_container_width=True, type="primary")
        with col_btn2: delete_btn = st.button("🗑️ 删除", use_container_width=True) if edit_oid > 0 else False

        if submitted:
            if is_new_book and not bk_name:
                st.error("❌ 新增教材时，教材名称为必填")
                st.stop()
            # 1) 处理教材主表
            if is_new_book:
                execute_sql("INSERT INTO textbooks_master (name,isbn,publisher,editor,price,course_name) VALUES (%s,%s,%s,%s,%s,%s)",
                    (bk_name, bk_isbn, bk_publisher, bk_editor, bk_price, bk_course))
                conn = sqlite3.connect(get_sqlite_path())
                cur = conn.cursor()
                cur.execute("SELECT MAX(id) as mid FROM textbooks_master")
                mid = cur.fetchone()[0]
                conn.close()
            else:
                mid = picked[0]

            if edit_oid > 0:
                execute_sql("UPDATE textbook_orders SET semester_id=%s,grade=%s,college=%s,major=%s,class_name=%s,quantity=%s,remark=%s WHERE id=%s",
                    (sem_id[0], (sel_grades or [None])[0], (sel_colleges or [None])[0], (sel_majors or [None])[0], (sel_classes or [None])[0], o_qty, o_remark, edit_oid))
                st.success("✅ 征订已更新")
            else:
                # 如果启用了自动计算，构建班级→数量的映射表（使用表单内的变量）
                auto_qty_map = {}
                if use_auto_qty and sel_classes and 'cnt_df' in dir() and not cnt_df.empty:
                    splits_local = [(row["class_name"], int(row["student_count"]), int(row["student_count"]))
                                     for _, row in cnt_df.iterrows()]
                    for cn, sc, alloc in splits_local:
                        auto_qty_map[cn] = alloc

                grades, colleges, majors, classes = sel_grades or [None], sel_colleges or [None], sel_majors or [None], sel_classes or [None]
                cnt = 0
                for g in grades:
                    for c in colleges:
                        for m in majors:
                            for cl in classes:
                                # 如果有按班级分配的数量，使用之；否则用统一的 o_qty
                                final_qty = auto_qty_map.get(cl, o_qty)
                                execute_sql("INSERT INTO textbook_orders (semester_id,textbook_id,grade,college,major,class_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                                    (sem_id[0], mid, g, c, m, cl, final_qty, o_remark))
                                # 同步到旧textbooks表保持兼容
                                execute_sql("INSERT INTO textbooks (semester_id,grade,college,major,class_name,name,isbn,publisher,editor,price,course_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                    (sem_id[0], g, c, m, cl, picked[1] if not is_new_book else bk_name,
                                     bk_isbn if is_new_book else next((r.get("isbn","") for _,r in master_books.iterrows() if r["id"]==picked[0]), ""),
                                     bk_publisher if is_new_book else next((r.get("publisher","") for _,r in master_books.iterrows() if r["id"]==picked[0]), ""),
                                     bk_editor if is_new_book else next((r.get("editor","") for _,r in master_books.iterrows() if r["id"]==picked[0]), ""),
                                     bk_price if is_new_book else next((r["price"] for _,r in master_books.iterrows() if r["id"]==picked[0]), 0),
                                     bk_course if is_new_book else next((r.get("course_name","") for _,r in master_books.iterrows() if r["id"]==picked[0]), ""),
                                     final_qty, o_remark))
                                cnt += 1
                st.success(f"✅ 已添加 {cnt} 条征订记录")
            st.rerun()
        if delete_btn:
            execute_sql("DELETE FROM textbook_orders WHERE id=%s", (edit_oid,))
            st.success("✅ 已删除")
            st.rerun()

    with tab3:
        st.markdown("#### 📥 从 Excel 导入教材征订")
        template_cols = ["学期", "年级", "学院", "专业", "班级", "教材名称", "书号(ISBN)", "出版社", "主编", "单价(元)", "征订数量", "备注"]
        template_df = make_template_df(template_cols)
        st.download_button("📄 下载导入模板", data=excel_export(template_df, "教材征订导入模板"),
            file_name="教材征订导入模板.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="secondary")
        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="tb_upload3")
        if uploaded:
            try:
                raw_df = read_excel_upload(uploaded)
                st.info(f"📄 检测到 {len(raw_df)} 行")
                col_map = {}
                for col in raw_df.columns:
                    cl = col.strip().lower()
                    if "教材" in cl: col_map["name"] = col
                    elif "学期" in cl: col_map["semester_name"] = col
                    elif "年级" in cl: col_map["grade"] = col
                    elif "学院" in cl: col_map["college"] = col
                    elif "专业" in cl and "年级" in cl: col_map["major_grade"] = col  # 专业年级（合并列）
                    elif "专业" in cl: col_map["major"] = col
                    elif "班级" in cl: col_map["class_name"] = col
                    elif "学生人数" in cl or "人数" in cl: col_map["student_count"] = col  # 学生人数
                    elif "教师" in cl: col_map["teacher_books"] = col  # 教师用书
                    elif "单价" in cl: col_map["price"] = col
                    elif "数量" in cl or "合计" in cl: col_map["quantity"] = col
                    elif "书号" in cl: col_map["isbn"] = col
                    elif "出版社" in cl: col_map["publisher"] = col
                    elif "主编" in cl or "作者" in cl: col_map["editor"] = col
                    elif "备注" in cl or "说明" in cl: col_map["remark"] = col
                if "name" not in col_map or "semester_name" not in col_map:
                    st.error("❌ 缺少必要列：教材名称、学期")
                else:
                    # 检测是否有"专业年级"合并列，有则提示自动拆班
                    has_major_grade = "major_grade" in col_map
                    if has_major_grade:
                        st.info("📌 检测到「专业年级」列，导入时将自动按各班级实际人数分摊征订数量")

                    if st.button("✅ 确认导入", use_container_width=True, type="primary", key="tb_import3"):
                        mapped = raw_df.rename(columns={v: k for k, v in col_map.items()})
                        # 统一列顺序
                        col_order = ["semester_name","grade","college","major","class_name","major_grade","name","isbn","publisher","editor","price","quantity","student_count","teacher_books","remark"]
                        mapped = mapped[[c for c in col_order if c in mapped.columns]]
                        sem_map = {r["name"]: r["id"] for _, r in semesters.iterrows()}
                        success, errors, total_rows = 0, [], len(mapped)
                        progress_bar = st.progress(0, text="正在导入...")

                        for i, (_, row) in enumerate(mapped.iterrows()):
                            sid = sem_map.get(str(row["semester_name"]))
                            if sid is None: errors.append(f"学期「{row['semester_name']}」不存在"); continue
                            try:
                                nm, p = safe_str(row["name"]), safe_float(row.get("price", 0))
                                # 查找或创建教材主记录
                                old = query_df("SELECT id FROM textbooks_master WHERE name=%s AND COALESCE(isbn,'')=%s", (nm, str(row.get("isbn","") or "")))
                                if old.empty:
                                    execute_sql("INSERT INTO textbooks_master (name,isbn,publisher,editor,price) VALUES (%s,%s,%s,%s,%s)",
                                        (nm, str(row.get("isbn","") or ""), str(row.get("publisher","") or ""), str(row.get("editor","") or ""), p))
                                    conn2 = sqlite3.connect(get_sqlite_path())
                                    c2 = conn2.cursor(); c2.execute("SELECT MAX(id) FROM textbooks_master"); tid = c2.fetchone()[0]; conn2.close()
                                else:
                                    tid = old.iloc[0]["id"]

                                college_val = str(row.get("college","") or "")
                                major_val = str(row.get("major","") or "")
                                class_val = str(row.get("class_name","") or "")
                                grade_val = str(row.get("grade","") or "")
                                mg_val = str(row.get("major_grade","") or "")
                                total_qty = safe_int(row.get("quantity", 0))

                                # 如果启用了自动拆班且有专业年级列
                                if has_major_grade and mg_val and total_qty > 0:
                                    # 尝试在 students 表中查找匹配的班级
                                    # 用 college + 从专业年级提取的专业名来过滤
                                    mg_major = ""
                                    # 尝试从专业年级提取专业名（如"电商1241-1244"→"电商"）
                                    import re
                                    mg_parts = re.split(r'(\d)', mg_val, maxsplit=1)
                                    if mg_parts:
                                        mg_major = mg_parts[0].strip()
                                    # 用学院+提取的专业名查找各班人数
                                    class_df = get_class_student_counts(college=college_val, major=major_val or mg_major)
                                    if not class_df.empty:
                                        total_students = int(class_df["student_count"].sum())
                                        teacher_cnt = safe_int(row.get("teacher_books", 0))
                                        # 总征订量 = 学生人数 + 教师用书（如果未单独指定教师用书列，就用合计数量）
                                        if "teacher_books" in row and pd.notna(row.get("teacher_books")):
                                            total_for_split = total_qty
                                        else:
                                            total_for_split = total_students + teacher_cnt if teacher_cnt > 0 else total_qty
                                        splits = [(row["class_name"], int(row["student_count"]), int(row["student_count"]))
                                                   for _, row in class_df.iterrows()]
                                        for cn, sc, alloc in splits:
                                            execute_sql("INSERT INTO textbook_orders (semester_id,textbook_id,grade,college,major,class_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                                                (sid, tid, grade_val or mg_major, college_val, major_val or mg_major, cn, alloc, str(row.get("remark","") or "")))
                                            execute_sql("INSERT INTO textbooks (semester_id,grade,college,major,class_name,name,isbn,publisher,editor,price,course_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                                (sid, grade_val or mg_major, college_val, major_val or mg_major, cn,
                                                 nm, str(row.get("isbn","") or ""), str(row.get("publisher","") or ""), str(row.get("editor","") or ""),
                                                 p, str(row.get("course_name","") or ""), alloc, str(row.get("remark","") or "")))
                                        success += 1
                                    else:
                                        # 数据库中没有匹配的班级，按原始数据导入
                                        execute_sql("INSERT INTO textbook_orders (semester_id,textbook_id,grade,college,major,class_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                                            (sid, tid, grade_val, college_val, major_val, mg_val, total_qty, str(row.get("remark","") or "")))
                                        execute_sql("INSERT INTO textbooks (semester_id,grade,college,major,class_name,name,isbn,publisher,editor,price,course_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                            (sid, grade_val, college_val, major_val, mg_val,
                                             nm, str(row.get("isbn","") or ""), str(row.get("publisher","") or ""), str(row.get("editor","") or ""),
                                             p, str(row.get("course_name","") or ""), total_qty, str(row.get("remark","") or "")))
                                        success += 1
                                else:
                                    # 标准导入逻辑（已有班级列）
                                    execute_sql("INSERT INTO textbook_orders (semester_id,textbook_id,grade,college,major,class_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                                        (sid, tid, grade_val, college_val, major_val, class_val, total_qty, str(row.get("remark","") or "")))
                                    execute_sql("INSERT INTO textbooks (semester_id,grade,college,major,class_name,name,isbn,publisher,editor,price,course_name,quantity,remark) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                        (sid, grade_val, college_val, major_val, class_val,
                                         nm, str(row.get("isbn","") or ""), str(row.get("publisher","") or ""), str(row.get("editor","") or ""),
                                         p, str(row.get("course_name","") or ""), total_qty, str(row.get("remark","") or "")))
                                    success += 1
                            except Exception as e: errors.append(f"{nm}: {str(e)[:100]}")
                            progress_bar.progress((i + 1) / total_rows, text=f"已处理 {i+1}/{total_rows}")
                        # 统一展示结果
                        mc_ok2, mc_err2 = st.columns(2)
                        mc_ok2.metric("✅ 导入成功", success)
                        mc_err2.metric("⚠️ 导入失败", len(errors))
                        if errors:
                            with st.expander(f"查看 {len(errors)} 条失败详情"):
                                for e in errors:
                                    st.caption(f"• {e}")
                        write_import_log("教材征订导入", uploaded.name, total_rows, success, errors)
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")

    with tab4:
        st.markdown("#### 📚 教材批量分配")
        st.caption("选择一本教材，批量分配给多个年级/学院/专业/班级")

        be_sem = st.selectbox("① 选择学期", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
            format_func=lambda x: x[1], key="be_sem")
        be_book = st.selectbox("② 选择教材", master_options, format_func=lambda x: x[1], key="be_book")
        st.caption("③~⑥ 至少选一项，不选则视为全部")

        with st.container():
            col_a, col_b = st.columns(2)
            with col_a:
                be_colleges = st.multiselect("③ 学院（可多选）", options=get_filtered_colleges(), key="be_cl")
                be_grades = st.multiselect("④ 年级（可多选）", options=get_filtered_grades(), key="be_gr")
            with col_b:
                # 专业：根据所选学院级联过滤
                if be_colleges:
                    be_major_where = " OR ".join(["college = %s"] * len(be_colleges))
                    be_major_opts = get_filtered_list("students", "major", be_major_where, tuple(be_colleges))
                else:
                    be_major_opts = get_filtered_majors()
                be_majors = st.multiselect("⑤ 专业（可多选）", options=be_major_opts, key="be_mj")
                # 班级：根据所选学院+专业级联过滤
                be_class_where = "1=1"; be_class_params = []
                if be_colleges:
                    be_class_where += " AND (" + " OR ".join(["college = %s"] * len(be_colleges)) + ")"
                    be_class_params.extend(be_colleges)
                if be_majors:
                    be_class_where += " AND (" + " OR ".join(["major = %s"] * len(be_majors)) + ")"
                    be_class_params.extend(be_majors)
                be_class_opts = get_filtered_list("students", "class_name", be_class_where, tuple(be_class_params)) if be_class_params else get_filtered_class_names()
                be_classes = st.multiselect("⑥ 班级（可多选）", options=be_class_opts, key="be_cls")
        
        be_qty = st.number_input("⑦ 征订数量", min_value=0, value=1, step=1, key="be_qty")
        
        if be_book[0] > 0:
            grades = be_grades or [None]
            colleges = be_colleges or [None]
            majors = be_majors or [None]
            classes = be_classes or [None]
            total_gen = len(grades) * len(colleges) * len(majors) * len(classes)
            
            if total_gen > 0:
                st.info(f"📌 将生成 **{total_gen}** 条征订记录")
                
                if st.button("✅ 确认批量分配", use_container_width=True, type="primary", key="be_confirm"):
                    cnt = 0
                    for g in grades:
                        for c in colleges:
                            for m in majors:
                                for cl in classes:
                                    execute_sql(
                                        "INSERT INTO textbook_orders (semester_id,textbook_id,grade,college,major,class_name,quantity) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                                        (be_sem[0], be_book[0], g, c, m, cl, be_qty))
                                    cnt += 1
                    st.success(f"✅ 已生成 {cnt} 条征订记录")
                    write_import_log("教材批量分配", be_book[1], total_gen, cnt, [])
                    st.rerun()
        else:
            st.warning("请先选择教材（不能选择「新增教材」）")
    
    # 底部快捷入口：教材表管理
    st.divider()
    st.caption("💡 如需添加/编辑教材信息，请前往侧边栏「📖 教材表管理」")

# 4. 教材发放表（打印版：从征订表带出，打印给发书人）
# ═════════════════════════════════════════════════════════

def distribution_management():
    show_header("📦 教材发放表", "按征订表带出教材清单，按班级打印领书单供学生签字确认")

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先添加学期")
        return

    class_names = get_filtered_class_names()
    colleges = get_filtered_colleges()
    majors = get_filtered_majors()

    tab1, tab2 = st.tabs(["📋 发书单", "📥 导入 Excel"])

    # ── Tab 1: 发书单 ──
    with tab1:
        col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1.5, 1.5, 1.5])
        with col_f1:
            f_sem = st.selectbox("学期", [(0, "全部")] + [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                                 format_func=lambda x: x[1], key="dl_sem", index=1)
        with col_f2:
            f_college = st.selectbox("学院", ["全部"] + colleges, key="dl_college")
        with col_f3:
            # 专业：基于所选学院级联过滤
            if f_college != "全部":
                dl_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (f_college,))
            else:
                dl_major_opts = ["全部"] + majors
            f_major = st.selectbox("专业", dl_major_opts, key="dl_major")
        with col_f4:
            # 班级：基于所选学院+专业级联过滤
            dl_class_where = "1=1"; dl_class_params = []
            if f_college != "全部":
                dl_class_where += " AND college = %s"; dl_class_params.append(f_college)
            if f_major != "全部":
                dl_class_where += " AND major = %s"; dl_class_params.append(f_major)
            dl_class_opts = ["全部"] + (get_filtered_list("students", "class_name", dl_class_where, tuple(dl_class_params)) if dl_class_params else get_filtered_class_names())
            f_class = st.selectbox("班级", dl_class_opts, key="dl_class")

        if f_sem[0] > 0:
            sem_id = f_sem[0]
            ay, semester_n = get_current_academic_info()
            sem_name = f_sem[1]
            sem_label = sem_name.replace(" ", "")

            # ── 录入发放记录 ──
            st.markdown("#### ➕ 录入发放记录")
            st.caption("选择班级→学生→教材，手动录入一条发放记录")

            # 录入专用的班级选择
            entry_class_where = "1=1"; entry_class_params = []
            if f_college != "全部":
                entry_class_where += " AND college = %s"; entry_class_params.append(f_college)
            if f_major != "全部":
                entry_class_where += " AND major = %s"; entry_class_params.append(f_major)
            entry_class_opts = get_filtered_list("students", "class_name", entry_class_where, tuple(entry_class_params)) if entry_class_params else get_filtered_class_names()
            if not entry_class_opts:
                entry_class_opts = class_names

            col_e1, col_e2, col_e3 = st.columns([1.5, 1.5, 1])
            with col_e1:
                entry_class = st.selectbox("发放班级 *", entry_class_opts, key="entry_class")
            with col_e2:
                if entry_class:
                    entry_stu = query_df(
                        "SELECT id, student_id, name FROM students WHERE class_name = %s ORDER BY student_id",
                        (entry_class,))
                    stu_opts = [(r["id"], f"{r['student_id']} {r['name']}") for _, r in entry_stu.iterrows()]
                    entry_student = st.selectbox("学生 *", stu_opts, format_func=lambda x: x[1], key="entry_student")
                else:
                    entry_student = None
            with col_e3:
                if entry_class:
                    # 查该班该学期的教材（从 textbook_orders + textbooks_master，确保有数据）
                    entry_tb = query_df("""
                        SELECT t.id, m.name, m.course_name, m.price
                        FROM textbook_orders o
                        JOIN textbooks_master m ON o.textbook_id = m.id
                        LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                        WHERE o.semester_id = %s AND o.class_name = %s
                        ORDER BY m.name
                    """, (sem_id, entry_class))
                    if entry_tb.empty:
                        entry_textbook = None
                        st.caption("⚠️ 该班暂无教材")
                    else:
                        # 确保 textbooks 表有记录（没有则自动创建），返回 textbooks.id
                        tb_opts = []
                        for _, r in entry_tb.iterrows():
                            tbid = r["id"]
                            if tbid is None or pd.isna(tbid):
                                # 自动创建 textbooks 记录
                                execute_sql(
                                    "INSERT INTO textbooks (semester_id, grade, college, major, class_name, name, price, course_name, quantity, remark) VALUES (%s, '', '', '', %s, %s, %s, %s, 0, '[逐条录入·自动创建]')",
                                    (sem_id, entry_class, r["name"], r["price"], r.get("course_name") or ""))
                                nr = query_df("SELECT id FROM textbooks WHERE semester_id=%s AND class_name=%s AND name=%s",
                                    (sem_id, entry_class, r["name"]))
                                tbid = int(nr.iloc[0]["id"]) if not nr.empty else None
                            if tbid:
                                tb_opts.append((tbid, f"{r['name']}" + (f"({r['course_name']})" if r.get('course_name') else "") + f" ¥{r['price']:.2f}"))
                        entry_textbook = st.selectbox("教材 *", tb_opts, format_func=lambda x: x[1], key="entry_textbook") if tb_opts else None
                else:
                    entry_textbook = None

            col_e4, col_e5, col_e6 = st.columns([1, 1.5, 1.5])
            with col_e4:
                entry_qty = st.number_input("数量", min_value=1, value=1, step=1, key="entry_qty")
            with col_e5:
                entry_date = st.date_input("发放日期", value=date.today(), key="entry_date")
            with col_e6:
                entry_handler = st.text_input("经手人", key="entry_handler", placeholder="录入人")

            if st.button("➕ 确认录入", use_container_width=True, type="primary", key="entry_submit"):
                if not entry_class:
                    st.warning("⚠️ 请选择发放班级")
                elif entry_student is None:
                    st.warning("⚠️ 请选择学生")
                elif entry_textbook is None:
                    st.warning("⚠️ 该班暂无教材，请先在「征订总表」中下发教材")
                else:
                    tb_id = int(entry_textbook[0])
                    execute_sql(
                        "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, %s, %s, %s)",
                        (int(entry_student[0]), tb_id, entry_qty, entry_date, entry_handler or ""))
                    st.success(f"✅ 已录入发放：{entry_student[1]} ← 《{entry_textbook[1]}》 ×{entry_qty}")
                    st.rerun()

            # ── 批量发放录入（从征订表带出数据，支持多班级）──
            st.divider()
            st.markdown("#### 📝 批量发放录入")
            st.caption('从征订表带出所选班级全部学生的教材清单，勾选「已发放」后一键批量确认。支持多班同时操作。')

            batch_class_where = "1=1"; batch_class_params = []
            if f_college != "全部":
                batch_class_where += " AND college = %s"; batch_class_params.append(f_college)
            if f_major != "全部":
                batch_class_where += " AND major = %s"; batch_class_params.append(f_major)
            batch_class_opts = get_filtered_list("students", "class_name", batch_class_where, tuple(batch_class_params)) if batch_class_params else get_filtered_class_names()
            if not batch_class_opts:
                batch_class_opts = class_names

            cc1, cc2 = st.columns([4, 1])
            with cc1:
                batch_classes = st.multiselect("批量录入班级（可多选）*", batch_class_opts, key="batch_classes",
                    placeholder="选择一个或多个班级...")
            with cc2:
                def _select_all_classes():
                    st.session_state["batch_classes"] = batch_class_opts
                st.button("☑ 全选班级", key="bsel_all_classes", on_click=_select_all_classes, use_container_width=True)

            if batch_classes:
                # ── 按班级收集数据 ──
                all_rows = []
                all_existing_set = set()
                total_stu_count = 0
                unique_tb_names = set()
                has_empty_stu = False
                has_empty_tb = False

                for bc in batch_classes:
                    # 查该班学生
                    bc_stu = query_df(
                        "SELECT id, student_id, name FROM students WHERE class_name = %s ORDER BY student_id",
                        (bc,))
                    if bc_stu.empty:
                        has_empty_stu = True
                        continue

                    # 查该班该学期的教材
                    raw_tb = query_df("""
                        SELECT o.textbook_id as master_id, m.name, m.course_name, m.isbn, m.publisher, m.editor, m.price
                        FROM textbook_orders o
                        JOIN textbooks_master m ON o.textbook_id = m.id
                        WHERE o.semester_id = %s AND o.class_name = %s
                        ORDER BY m.name
                    """, (sem_id, bc))

                    if raw_tb.empty:
                        has_empty_tb = True
                        continue

                    # 确保 textbooks 表存在对应记录
                    tb_id_map = {}
                    for _, rt in raw_tb.iterrows():
                        mid = int(rt["master_id"])
                        tbid_row = query_df(
                            "SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s AND name = %s",
                            (sem_id, bc, rt["name"]))
                        if not tbid_row.empty:
                            tb_id_map[mid] = int(tbid_row.iloc[0]["id"])
                        else:
                            execute_sql(
                                """INSERT INTO textbooks
                                (semester_id, grade, college, major, class_name, name, isbn, publisher, editor, price, course_name, quantity, remark)
                                VALUES (%s, '', '', '', %s, %s, %s, %s, %s, %s, %s, 0, '[自动创建·来自征订数据]')""",
                                (sem_id, bc, rt["name"], rt.get("isbn") or "",
                                 rt.get("publisher") or "", rt.get("editor") or "",
                                 rt["price"], rt.get("course_name") or ""))
                            new_row = query_df(
                                "SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s AND name = %s",
                                (sem_id, bc, rt["name"]))
                            if not new_row.empty:
                                tb_id_map[mid] = int(new_row.iloc[0]["id"])

                    # 重建 batch_tb
                    batch_tb = []
                    for _, rt in raw_tb.iterrows():
                        mid = int(rt["master_id"])
                        if mid in tb_id_map:
                            batch_tb.append({
                                "textbook_id": tb_id_map[mid],
                                "name": rt["name"],
                                "course_name": rt.get("course_name") or "",
                                "price": rt["price"],
                            })
                            unique_tb_names.add(rt["name"])

                    # 查已有发放记录
                    existing = query_df(
                        "SELECT student_id, textbook_id FROM distributions WHERE textbook_id IN (SELECT id FROM textbooks WHERE semester_id = %s AND class_name = %s)",
                        (sem_id, bc))
                    for _, er in existing.iterrows():
                        all_existing_set.add((int(er["student_id"]), int(er["textbook_id"])))

                    # 构建学生×教材矩阵
                    for _, stu in bc_stu.iterrows():
                        sid = int(stu["id"])
                        for tb in batch_tb:
                            tid = tb["textbook_id"]
                            already = (sid, tid) in all_existing_set
                            all_rows.append({
                                "班级": bc,
                                "student_id": sid,
                                "textbook_id": tid,
                                "学号": stu["student_id"],
                                "姓名": stu["name"],
                                "教材名称": tb["name"],
                                "课程": tb["course_name"],
                                "单价": tb["price"],
                                "已发放": already,
                            })
                    total_stu_count += len(bc_stu)

                if not all_rows:
                    if has_empty_stu:
                        st.warning("⚠️ 部分班级暂无学生")
                    if has_empty_tb:
                        st.warning("⚠️ 部分班级暂无教材，请先在「征订总表」中下发教材")
                else:
                    batch_df = pd.DataFrame(all_rows)
                    st.caption(f"共 {len(batch_classes)} 个班、{total_stu_count} 人 × {len(unique_tb_names)} 种教材 = {len(batch_df)} 条记录，已发放 {len(all_existing_set)} 条")

                    # 全选/取消全选 按钮
                    sel_col1, sel_col2, sel_col3, sel_col4 = st.columns([1, 1, 2, 2])
                    with sel_col1:
                        if st.button("☑ 全选", key="bsel_all", use_container_width=True):
                            st.session_state["batch_select_mode"] = "all"
                            st.rerun()
                    with sel_col2:
                        if st.button("☐ 全不选", key="bsel_none", use_container_width=True):
                            st.session_state["batch_select_mode"] = "none"
                            st.rerun()
                    with sel_col3:
                        if st.button("🔄 恢复已发放", key="bsel_restore", use_container_width=True):
                            st.session_state.pop("batch_select_mode", None)
                            st.rerun()
                    # 根据 session_state 调整初始勾选
                    sel_mode = st.session_state.get("batch_select_mode", "")
                    if sel_mode == "all":
                        batch_df["已发放"] = True
                    elif sel_mode == "none":
                        batch_df["已发放"] = False

                    with st.form("batch_dist_form"):
                        edited_batch = st.data_editor(
                            batch_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(35 * len(batch_df) + 38, 600),
                            column_config={
                                "班级": st.column_config.TextColumn("班级", disabled=True, width="small", alignment="center"),
                                "student_id": None,
                                "textbook_id": None,
                                "学号": st.column_config.TextColumn("学号", disabled=True, width="small", alignment="center"),
                                "姓名": st.column_config.TextColumn("姓名", disabled=True, width="small", alignment="center"),
                                "教材名称": st.column_config.TextColumn("教材名称", disabled=True, alignment="center"),
                                "课程": st.column_config.TextColumn("课程", disabled=True, alignment="center"),
                                "单价": st.column_config.NumberColumn("单价", disabled=True, format="¥%.2f", width="small", alignment="center"),
                                "已发放": st.column_config.CheckboxColumn("已发放", help="勾选=该生已领此书"),
                            },
                            disabled=["班级", "学号", "姓名", "教材名称", "课程", "单价"],
                            num_rows="fixed"
                        )

                        col_bf1, col_bf2 = st.columns([1.5, 1])
                        with col_bf1:
                            batch_date = st.date_input("发放日期", value=date.today(), key="batch_date")
                        with col_bf2:
                            batch_handler = st.text_input("经手人", key="batch_handler", placeholder="录入人")

                        batch_btn = st.form_submit_button("✅ 批量确认发放", use_container_width=True, type="primary")

                    if batch_btn:
                        to_insert = edited_batch[edited_batch["已发放"] == True]
                        new_records = []
                        for _, row in to_insert.iterrows():
                            key = (int(row["student_id"]), int(row["textbook_id"]))
                            if key not in all_existing_set:
                                new_records.append(row)

                        if not new_records:
                            st.info("📭 没有新的发放记录需要保存（全部已存在）")
                        else:
                            count = 0
                            for row in new_records:
                                tb_id = int(row["textbook_id"])
                                execute_sql(
                                    "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, 1, %s, %s)",
                                    (int(row["student_id"]), tb_id, batch_date, batch_handler or ""))
                                count += 1
                            st.success(f"✅ 批量录入 {count} 条发放记录")
                            st.rerun()

            st.divider()

            # ── 发书单预览 ──
            st.markdown("#### 📋 发书单预览")
            st.caption("下方展示各班的教材配发清单，供核对用")
            tb_sql = """SELECT o.class_name, o.grade, o.college, o.major, o.quantity,
                               m.name, m.course_name, m.isbn, m.publisher, m.editor, m.price,
                               o.id as order_id
                        FROM textbook_orders o
                        JOIN textbooks_master m ON o.textbook_id = m.id
                        WHERE o.semester_id = %s"""
            tb_params = [sem_id]
            if f_college != "全部":
                tb_sql += " AND o.college = %s"; tb_params.append(f_college)
            if f_major != "全部":
                tb_sql += " AND o.major = %s"; tb_params.append(f_major)
            if f_class != "全部":
                tb_sql += " AND o.class_name = %s"; tb_params.append(f_class)
            tb_sql += " ORDER BY o.class_name, m.name"

            textbooks = query_df(tb_sql, tuple(tb_params))

            if textbooks.empty:
                st.info("📭 当前条件下暂无已下发的教材领用数据，请先在「征订总表」中执行「一键下发」")
            else:
                # 按班级分组
                class_books = {}
                for _, tb in textbooks.iterrows():
                    cn = tb["class_name"]
                    if cn not in class_books:
                        class_books[cn] = []
                    class_books[cn].append(tb)

                # 发书单预览
                st.caption("📋 **发书单** — 按学生列出教材及费用，供发书核对")
                for cn, tbs in class_books.items():
                    students = query_df(
                        "SELECT student_id, name FROM students WHERE class_name = %s ORDER BY student_id", (cn,))
                    rows = []
                    for i, (_, stu) in enumerate(students.iterrows(), 1):
                        book_names = "、".join([f"《{b['name']}》" + (f"({b['course_name']})" if b.get('course_name') else "") for b in tbs])
                        total = sum(b["price"] for b in tbs)
                        rows.append({"序号": i, "学号": stu["student_id"], "姓名": stu["name"],
                                     "领取教材": book_names, "教材数": len(tbs), "合计(元)": round(total, 2),
                                     "领书人签字": "", "联系电话": "", "领书时间": ""})
                    lsd = pd.DataFrame(rows)
                    total_fee = sum(r["合计(元)"] for r in rows)
                    with st.expander(f"📦 {cn}（{len(students)}人 × {len(tbs)}种教材 = ¥{total_fee:.2f}）", expanded=False):
                        st.dataframe(lsd.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True,
                                     column_config={"合计(元)": st.column_config.NumberColumn("合计(元)", format="¥%.2f")})

                # 导出 - 教材发放清单（正式格式）
                st.divider()
                st.markdown("##### 📥 导出正式教材发放清单")
                col_sl1, col_sl2 = st.columns([2, 1])
                with col_sl1:
                    lq_school = st.text_input("学校名称", value="湖南理工职业技术学院", key="lq_school")
                with col_sl2:
                    lq_phone = st.text_input("联系电话", value="", key="lq_phone", placeholder="选填")

                def export_formal_bookreceipt():
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                    from openpyxl.worksheet.properties import PageSetupProperties
                    import io
                    output = io.BytesIO()
                    wb = Workbook()
                    default_ws = wb.active
                    wb.remove(default_ws)

                    cell_border = Border(
                        left=Side(style="medium"), right=Side(style="medium"),
                        top=Side(style="medium"), bottom=Side(style="medium")
                    )
                    blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                    hfont = Font(bold=True, size=10)
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    dfont = Font(size=10)

                    for cn, tbs in class_books.items():
                        ws = wb.create_sheet(title=str(cn)[:31])
                        nc = 9  # 班级、序号、课程、教材名称、主编、出版社、单价、实发、领书人

                        cq = query_df("SELECT COUNT(*) as c FROM students WHERE class_name = %s", (cn,))
                        stu_count = int(cq.iloc[0]["c"]) if not cq.empty else 0
                        college_q = query_df("SELECT DISTINCT college FROM students WHERE class_name = %s", (cn,))
                        tb_college = college_q.iloc[0]["college"] if not college_q.empty else ""

                        # Row 1: 大标题
                        for ci in range(1, nc + 1):
                            ws.cell(row=1, column=ci).border = cell_border
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                        t = ws.cell(row=1, column=1, value=f"{lq_school} {sem_label}教材发放清单")
                        t.font = Font(bold=True, size=14)
                        t.alignment = Alignment(horizontal="center", vertical="center")
                        # 合并后补设边框
                        for ci in range(1, nc + 1):
                            ws.cell(row=1, column=ci).border = cell_border
                        ws.row_dimensions[1].height = 30

                        # Row 2: 信息栏
                        for ci in range(1, nc + 1):
                            ws.cell(row=2, column=ci).border = cell_border
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
                        info_text = f"学院：{tb_college}    学期：{sem_name}    人数：{stu_count}人    电话：{lq_phone or '_____________'}"
                        ci2 = ws.cell(row=2, column=1, value=info_text)
                        ci2.font = Font(size=10)
                        ci2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        for ci in range(1, nc + 1):
                            ws.cell(row=2, column=ci).border = cell_border
                        ws.row_dimensions[2].height = 22

                        # Row 3: 表头
                        headers = ["班级", "序号", "课程", "教材名称", "主编", "出版社", "单价", "实发", "领书人"]
                        for i, h in enumerate(headers, 1):
                            c = ws.cell(row=3, column=i, value=h)
                            c.font = hfont; c.border = cell_border; c.fill = blue_fill
                            c.alignment = center
                        ws.row_dimensions[3].height = 22

                        # Row 4+: 数据
                        for idx, tb in enumerate(tbs):
                            rn = 4 + idx
                            ws.row_dimensions[rn].height = 26
                            vals = [
                                cn if idx == 0 else "",
                                idx + 1,
                                tb.get("course_name", ""),
                                tb["name"],
                                tb.get("editor", ""),
                                tb.get("publisher", ""),
                                round(float(tb.get("price", 0)), 2),
                                "",   # 实发：留空，打印后手填
                                ""    # 领书人：留空，打印后手签
                            ]
                            for ci, v in enumerate(vals, 1):
                                cell = ws.cell(row=rn, column=ci, value=v)
                                cell.font = dfont; cell.border = cell_border
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        # 合并班级列
                        if len(tbs) > 1:
                            ws.merge_cells(start_row=4, start_column=1, end_row=4 + len(tbs) - 1, end_column=1)
                        # 合并后重新补边框
                        for rrow in range(4, 4 + len(tbs)):
                            for ccol in range(1, nc + 1):
                                cell = ws.cell(row=rrow, column=ccol)
                                cell.border = cell_border
                                cell.alignment = center

                        # 列宽（适配A4打印，加宽文字列）
                        widths = [10, 6, 14, 28, 10, 18, 8, 6, 10]
                        for i, w in enumerate(widths, 1):
                            ws.column_dimensions[chr(64 + i)].width = w

                        # ── 打印页面设置（A4纸张，自适应一页宽）──
                        ws.page_setup.paperSize = 9       # A4
                        ws.page_setup.orientation = 'portrait'  # 纵向
                        ws.page_setup.fitToWidth = 1      # 强制一页宽
                        ws.page_setup.fitToHeight = 0     # 不限高
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                        ws.page_margins.left = 0.4
                        ws.page_margins.right = 0.4
                        ws.page_margins.top = 0.5
                        ws.page_margins.bottom = 0.5
                        ws.page_margins.header = 0.3
                        ws.page_margins.footer = 0.3
                        ws.print_title_rows = '1:3'  # 每页打印表头（标题+信息+列名）

                    wb.save(output)
                    return output.getvalue()

                if st.button("📥 导出正式教材发放清单", use_container_width=True, type="primary", key="lq_formal_export"):
                    st.session_state.lq_export_data = export_formal_bookreceipt()
                    st.session_state.lq_show_download = True
                    st.rerun()

                if st.session_state.get("lq_show_download"):
                    st.download_button(
                        "✅ 点击下载",
                        data=st.session_state.lq_export_data,
                        file_name=f"教材发放清单_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary", key="lq_dl"
                    )
        else:
            st.info("👆 请先选择学期")

    # ── Tab 2: 导入 Excel（批量导入发放记录）──
    with tab2:
        st.markdown("#### 📥 从 Excel 批量导入发放记录")
        st.caption("Excel 列名建议：学号（或身份证号）、教材名称、领取数量、发放日期、经手人")

        # 下载模板
        template_cols = ["学号", "身份证号", "教材名称", "领取数量", "发放日期", "经手人"]
        template_df = make_template_df(template_cols)
        template_bytes = excel_export(template_df, "发放记录导入模板")
        st.download_button(
            "📄 下载导入模板",
            data=template_bytes,
            file_name="发放记录导入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="secondary"
        )

        col1, col2 = st.columns(2)
        with col1:
            i_semester = st.selectbox("学期 *", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                              format_func=lambda x: x[1], key="di_sem")
        with col2:
            i_class = st.selectbox("班级 *", ["请选择班级"] + class_names, key="di_class")

        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="dist_upload")
        if uploaded:
            try:
                raw_df = read_excel_upload(uploaded)
                st.info(f"📄 检测到 {len(raw_df)} 行")

                col_map = {}
                for col in raw_df.columns:
                    cl = col.strip().lower()
                    if "学号" in cl or "student_id" in cl: col_map["student_id"] = col
                    elif "身份证" in cl or "id_card" in cl: col_map["id_card"] = col
                    elif "教材" in cl or "书名" in cl or "textbook" in cl: col_map["textbook_name"] = col
                    elif "数量" in cl or "quantity" in cl: col_map["quantity"] = col
                    elif "日期" in cl or "date" in cl: col_map["date"] = col
                    elif "经手" in cl or "handler" in cl: col_map["handler"] = col

                if not ("student_id" in col_map or "id_card" in col_map) or "textbook_name" not in col_map:
                    st.error("❌ 缺少必要列：学号/身份证号、教材名称")
                else:
                    if st.button("✅ 确认导入", use_container_width=True, type="primary"):
                        mapped = raw_df.rename(columns={v: k for k, v in col_map.items()})
                        sem_id = i_semester[0]

                        # 预加载映射
                        all_students = query_df("SELECT id, student_id, id_card FROM students WHERE class_name = %s", (i_class,))
                        students_map = {}
                        for _, s in all_students.iterrows():
                            students_map[str(s["student_id"])] = s["id"]
                            students_map[str(s["id_card"])] = s["id"]

                        all_textbooks = query_df(
                            "SELECT m.id, m.name FROM textbook_orders o "
                            "JOIN textbooks_master m ON o.textbook_id = m.id "
                            "WHERE o.semester_id = %s AND o.class_name = %s", (sem_id, i_class))
                        textbooks_map = {t["name"]: t["id"] for _, t in all_textbooks.iterrows()}

                        success, errors, total_rows = 0, [], len(mapped)
                        progress_bar = st.progress(0, text="正在导入...")
                        for i, (_, row) in enumerate(mapped.iterrows()):
                            sid_val = row.get("student_id")
                            lookup = safe_str(sid_val) if pd.notna(sid_val) else safe_str(row.get("id_card"))
                            if not lookup:
                                errors.append(f"空学号/身份证号"); continue
                            sid = students_map.get(lookup)
                            if sid is None:
                                errors.append(f"学生「{lookup}」未找到"); continue
                            tname = safe_str(row.get("textbook_name"))
                            tid = textbooks_map.get(tname)
                            if tid is None:
                                errors.append(f"教材「{tname}」未找到"); continue
                            try:
                                q = safe_int(row.get("quantity", 1), 1)
                                d = row.get("date")
                                h = row.get("handler", "")
                                if pd.notna(d):
                                    try:
                                        d = pd.Timestamp(d).date()
                                    except Exception:
                                        d = date.today()
                                else:
                                    d = date.today()
                                execute_sql(
                                    "INSERT INTO distributions (student_id,textbook_id,quantity,distribute_date,handler) VALUES (%s,%s,%s,%s,%s)",
                                    (sid, tid, q, d, safe_str(h)))
                                success += 1
                            except Exception as e:
                                errors.append(f"{lookup}×{tname}: {str(e)[:100]}")
                            progress_bar.progress((i + 1) / total_rows, text=f"已处理 {i+1}/{total_rows}")
                        # 统一展示结果
                        mc_ok3, mc_err3 = st.columns(2)
                        mc_ok3.metric("✅ 导入成功", success)
                        mc_err3.metric("⚠️ 导入失败", len(errors))
                        if errors:
                            with st.expander(f"查看 {len(errors)} 条失败详情"):
                                for err in errors:
                                    st.caption(f"• {err}")
                        write_import_log("发放记录导入", uploaded.name, total_rows, success, errors)
            except Exception as e:
                st.error(f"❌ 读取失败：{e}")

# ═════════════════════════════════════════════════════════
# 5. 学生领书确认表
# ═════════════════════════════════════════════════════════

def confirmation_page():
    show_header("✅ 学生领书确认表", "按班级确认学生是否需要教材，免领标记后该生费用统计将排除")
    
    # 页面跟踪：从其他页面切换进来时，重置班级选择
    prev_page = st.session_state.get("current_page", "")
    if prev_page != "confirmation":
        st.session_state.pop("c_class", None)
        st.session_state.pop("c_major", None)
    st.session_state["current_page"] = "confirmation"
    
    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先添加学期")
        return
    
    class_names = get_filtered_class_names()
    colleges = get_filtered_colleges()
    majors = get_filtered_majors()
    if not class_names:
        st.warning("⚠️ 请先在学生管理中导入学生数据")
        return
    
    tab1, tab2 = st.tabs(["📝 领书确认", "📋 确认汇总"])
    
    # ── Tab 1: 领书确认（查看领取情况，支持编辑备注）──
    with tab1:
        st.markdown("#### 📝 查看教材领取情况")
        st.caption('勾选「补领」可补发教材，勾选「退书」可退回。有退伍复学等特殊情况，请在「教材发放表」中补录。')

        col1, col2, col3, col4 = st.columns([2, 1.5, 1.5, 1.5])
        with col1:
            def _reset_c_class():
                st.session_state.pop("c_class", None)
            def _reset_c_major_and_class():
                st.session_state.pop("c_major", None)
                st.session_state.pop("c_class", None)
            c_semester = st.selectbox("学期 *", [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                               format_func=lambda x: x[1], key="c_sem", on_change=_reset_c_class)
        with col2:
            c_college = st.selectbox("学院", ["全部"] + colleges, key="c_college", on_change=_reset_c_major_and_class)
        with col3:
            # 专业：基于所选学院级联过滤
            if c_college != "全部":
                c_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (c_college,))
            else:
                c_major_opts = ["全部"] + majors
            c_major = st.selectbox("专业", c_major_opts, key="c_major", on_change=_reset_c_class)
        with col4:
            # 班级：基于所选学院+专业级联过滤
            c_class_where = "1=1"; c_class_params = []
            if c_college != "全部":
                c_class_where += " AND college = %s"; c_class_params.append(c_college)
            if c_major != "全部":
                c_class_where += " AND major = %s"; c_class_params.append(c_major)
            c_class_opts = get_filtered_list("students", "class_name", c_class_where, tuple(c_class_params)) if c_class_params else class_names
            # 处理从概览跳转到指定班级
            if st.session_state.get("_goto_class"):
                st.session_state["c_class"] = st.session_state.pop("_goto_class")
                st.rerun()
            c_class_opts = ["请选择班级"] + c_class_opts
            c_class = st.selectbox("班级 *", c_class_opts, key="c_class")
            st.caption("👆 请先选择班级，再查看领取情况")

        # ── 班级领取概览：选学期后自动展示（折叠）──
        if c_semester and (not c_class or c_class == "请选择班级"):
            sem_id_ov = c_semester[0]
            # 查所有已下发教材的班级及其学生/教材数
            ov_classes = query_df("""
                SELECT o.class_name, COUNT(DISTINCT s.id) as stu_count,
                       COUNT(DISTINCT o.textbook_id) as tb_count
                FROM textbook_orders o
                LEFT JOIN students s ON s.class_name = o.class_name
                WHERE o.semester_id = %s
                GROUP BY o.class_name
                ORDER BY o.class_name
            """, (sem_id_ov,))

            if not ov_classes.empty:
                # 查各班级的发放记录数
                ov_dist = query_df("""
                    SELECT s.class_name, COUNT(*) as dist_count
                    FROM distributions d
                    JOIN textbooks t ON d.textbook_id = t.id
                    JOIN students s ON d.student_id = s.id
                    WHERE t.semester_id = %s
                    GROUP BY s.class_name
                """, (sem_id_ov,))
                dist_map = {}
                for _, dr in ov_dist.iterrows():
                    dist_map[dr["class_name"]] = int(dr["dist_count"])

                # 按学院/专业筛选 + 逐班计算完成率
                ov_rows = []
                for _, oc in ov_classes.iterrows():
                    cn = oc["class_name"]
                    # 按学院筛选
                    if c_college != "全部":
                        cs = query_df("SELECT college FROM students WHERE class_name = %s LIMIT 1", (cn,))
                        if cs.empty or cs.iloc[0]["college"] != c_college:
                            continue
                    if c_major != "全部":
                        cs = query_df("SELECT major FROM students WHERE class_name = %s LIMIT 1", (cn,))
                        if cs.empty or cs.iloc[0]["major"] != c_major:
                            continue
                    stu = int(oc["stu_count"])
                    tb = int(oc["tb_count"])
                    total_pairs = stu * tb
                    actual = dist_map.get(cn, 0)
                    pct = round(actual / total_pairs * 100, 1) if total_pairs > 0 else 0
                    ov_rows.append({
                        "班级": cn,
                        "学生": stu,
                        "教材种数": tb,
                        "应发条数": total_pairs,
                        "已发条数": actual,
                        "完成率": pct,
                    })

                if ov_rows:
                    st.divider()
                    st.markdown("##### 📊 各班级领取情况概览")
                    st.caption("已发条数 = 学生×教材的发放记录数，100% 表示该班全部教材均已发放完毕")
                    for r in sorted(ov_rows, key=lambda x: x["完成率"]):
                        pct = r["完成率"]
                        if pct >= 100:
                            badge, icon = "🟢", "✅ 已全部完成"
                        elif pct > 0:
                            badge, icon = "🟡", "⏳ 部分完成"
                        else:
                            badge, icon = "🔴", "❌ 未发放"
                        with st.expander(f"{badge} {r['班级']} — {icon}（{r['已发条数']}/{r['应发条数']}，{pct}%）", expanded=False):
                            c1, c2, c3, c4, c5 = st.columns(5)
                            c1.metric("学生数", r["学生"])
                            c2.metric("教材种数", r["教材种数"])
                            c3.metric("应发条数", r["应发条数"])
                            c4.metric("已发条数", r["已发条数"])
                            c5.metric("完成率", f"{pct}%")
                            # 快速跳转：点击按钮自动选中该班级
                            if st.button(f"🔍 查看 {r['班级']} 详情", key=f"goto_{r['班级']}"):
                                st.session_state["_goto_class"] = r["班级"]
                                st.rerun()
                else:
                    if c_college != "全部" or c_major != "全部":
                        st.info("📭 当前筛选条件下暂无已下发教材的班级")
            else:
                st.info("📭 当前学期暂无已下发教材的班级，请先在「征订总表」中执行「一键下发」")

        if c_semester and c_class and c_class != "请选择班级":
            sem_id = c_semester[0]

            students_df = query_df(
                "SELECT id, id_card, student_id, name, grade, college, major FROM students WHERE class_name = %s ORDER BY name",
                (c_class,))
            textbooks_df = query_df(
                """SELECT m.name, m.publisher, m.price, t.id
                   FROM textbook_orders o
                   JOIN textbooks_master m ON o.textbook_id = m.id
                   LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                   WHERE o.semester_id = %s AND o.class_name = %s
                   ORDER BY m.name""",
                (sem_id, c_class))

            if students_df.empty:
                st.info(f"📭 班级「{c_class}」暂无学生")
                return
            if textbooks_df.empty:
                st.warning("⚠️ 该班级在当前学期暂无教材")
                return

            # 对没有 textbooks.id 的教材自动创建 textbooks 记录
            for _, tr in textbooks_df.iterrows():
                if tr["id"] is None or pd.isna(tr["id"]):
                    execute_sql(
                        "INSERT INTO textbooks (semester_id, grade, college, major, class_name, name, publisher, price, quantity, remark) VALUES (%s, '', '', '', %s, %s, %s, %s, 0, '[确认表·自动创建]')",
                        (sem_id, c_class, tr["name"], tr.get("publisher") or "", tr["price"]))
            # 重新查询以获取新的 textbooks.id
            textbooks_df = query_df(
                """SELECT m.name, m.publisher, m.price, t.id
                   FROM textbook_orders o
                   JOIN textbooks_master m ON o.textbook_id = m.id
                   LEFT JOIN textbooks t ON t.semester_id = o.semester_id AND t.class_name = o.class_name AND t.name = m.name
                   WHERE o.semester_id = %s AND o.class_name = %s
                   ORDER BY m.name""",
                (sem_id, c_class))

            # ── 选教材 ──
            st.markdown("##### 📖 选择教材查看领取情况")
            tb_options = [(r["id"], r["name"]) for _, r in textbooks_df.iterrows()]
            sel_tb = st.selectbox("教材", tb_options, format_func=lambda x: x[1], key="c_tb",
                              label_visibility="collapsed")

            # 该教材的发放情况
            dist_stu = query_df(
                "SELECT DISTINCT student_id FROM distributions WHERE textbook_id = %s",
                (sel_tb[0],))
            dist_ids = set(int(r["student_id"]) for _, r in dist_stu.iterrows())

            # 统计
            total = len(students_df)
            got = len(dist_ids)
            not_got = total - got

            m1, m2, m3 = st.columns(3)
            with m1:
                st.metric("👨‍🎓 班级人数", total)
            with m2:
                st.metric("✅ 已领", got)
            with m3:
                st.metric("❌ 未领", not_got)

            # 学生列表（带补领+退书操作）—— 用 form 确保 checkbox 状态正确捕获
            st.markdown("##### 👨‍🎓 学生领取情况")
            st.caption('未领可「补领」，已领可「退书」，勾选后点击下方对应按钮')

            rows = []
            for i, (_, stu) in enumerate(students_df.iterrows(), 1):
                sid = int(stu["id"])
                has_book = sid in dist_ids
                rows.append({
                    "编号": i,
                    "_sid": sid,
                    "学号": stu["student_id"],
                    "姓名": stu["name"],
                    "领取状态": "✅ 已领" if has_book else "❌ 未领",
                    "补领": False,
                    "退书": False,
                })

            status_df = pd.DataFrame(rows)

            with st.form(f"confirm_book_form_{c_class}_{sel_tb[0]}"):
                edited_df = st.data_editor(
                    status_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "编号": st.column_config.NumberColumn("编号", disabled=True, width="small", alignment="center"),
                        "_sid": None,
                        "学号": st.column_config.TextColumn("学号", disabled=True, alignment="center"),
                        "姓名": st.column_config.TextColumn("姓名", disabled=True, alignment="center"),
                        "领取状态": st.column_config.TextColumn("领取状态", disabled=True, alignment="center"),
                        "补领": st.column_config.CheckboxColumn("补领", help="勾选=补发此书（仅未领）"),
                        "退书": st.column_config.CheckboxColumn("退书", help="勾选=退回此书（仅已领）"),
                    },
                    disabled=["编号", "学号", "姓名", "领取状态"],
                    num_rows="fixed"
                )

                # ── 补领/退书操作区（在 form 内）──
                to_supplement = edited_df[(edited_df["领取状态"] == "❌ 未领") & (edited_df["补领"] == True)]
                to_refund = edited_df[(edited_df["领取状态"] == "✅ 已领") & (edited_df["退书"] == True)]

                c_date, c_handler, c_btn1, c_btn2 = st.columns([1, 1, 1, 1])
                with c_date:
                    op_date = st.date_input("操作日期", value=date.today(), key="op_date_f")
                with c_handler:
                    op_handler = st.text_input("经手人", key="op_handler_f", placeholder="录入人")

                with c_btn1:
                    sup_label = f"📦 补领 {len(to_supplement)} 人" if len(to_supplement) > 0 else "📦 补领"
                    sup_btn = st.form_submit_button(sup_label, use_container_width=True, type="primary")
                with c_btn2:
                    ref_label = f"🔙 退书 {len(to_refund)} 人" if len(to_refund) > 0 else "🔙 退书"
                    ref_btn = st.form_submit_button(ref_label, use_container_width=True, type="secondary")

            # ── 处理 form 提交 ──
            if sup_btn:
                if len(to_supplement) == 0:
                    st.warning("⚠️ 请先在表格中勾选需要「补领」的学生")
                else:
                    count = 0
                    tb_id = sel_tb[0]
                    for _, row in to_supplement.iterrows():
                        execute_sql(
                            "INSERT INTO distributions (student_id, textbook_id, quantity, distribute_date, handler) VALUES (%s, %s, 1, %s, %s)",
                            (int(row["_sid"]), tb_id, op_date, op_handler or ""))
                        count += 1
                    st.success(f"✅ 成功补领 {count} 人")
                    st.rerun()

            if ref_btn:
                if len(to_refund) == 0:
                    st.warning("⚠️ 请先在表格中勾选需要「退书」的学生")
                else:
                    count = 0
                    for _, row in to_refund.iterrows():
                        execute_sql(
                            "DELETE FROM distributions WHERE student_id=%s AND textbook_id=%s",
                            (int(row["_sid"]), sel_tb[0]))
                        count += 1
                    st.success(f"✅ 成功退书 {count} 人")
                    st.rerun()

            # 提示信息（form 外）
            if not sup_btn and not ref_btn:
                sup_edited = edited_df[(edited_df["领取状态"] == "❌ 未领") & (edited_df["补领"] == True)]
                ref_edited = edited_df[(edited_df["领取状态"] == "✅ 已领") & (edited_df["退书"] == True)]
                if len(sup_edited) == 0 and len(ref_edited) == 0 and not_got > 0:
                    st.info(f"💡 剩 {not_got} 人未领→勾选「补领」；已领的可勾选「退书」")

            # ── 导出正式领书单 ──
            st.markdown("##### 📥 导出领书单")
            col_exp1, col_exp2, col_exp3, col_exp4 = st.columns([1.5, 1, 1, 1])
            with col_exp1:
                school_name = st.text_input("学校名称", value="湖南理工职业技术学院", key="c_school")
            with col_exp2:
                handler_name = st.text_input("辅导员", value="", key="c_handler", placeholder="必填")
            with col_exp3:
                receiver_name = st.text_input("领书人", value="", key="c_receiver", placeholder="班级负责人")
            with col_exp4:
                handler_phone = st.text_input("联系电话", value="", key="c_phone", placeholder="选填")
            semester_label = c_semester[1].replace(" ","")
            college_val = students_df.iloc[0].get("college", "") if not students_df.empty else ""

            def export_formal_booksheet():
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side
                import io, math
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "领书单"
                nc = 10  # 固定10列（左右各5列: 序号、学号、姓名、领取状态、签字）

                # Row 1: 大标题
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                t = ws.cell(row=1, column=1, value=f"{school_name} {semester_label}领书单")
                t.font = Font(bold=True, size=18); t.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40

                # Row 2: 信息栏
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
                info_text = f"学院：{college_val}    班级：{c_class}    辅导员：{handler_name}    联系电话：{handler_phone or '________'}    领书人：{receiver_name or '___'}    班级人数：{total}人    领书人数：{got}人"
                inf = Font(bold=False, size=11)
                ws.cell(row=2, column=1, value=info_text).font = inf
                ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 25

                # Row 3+: 双栏表格（表头直接在信息栏下方）
                left_headers = ["序号", "领书人", "学号", "领取", "签字"]
                right_headers = ["序号", "领书人", "学号", "领取", "签字"]
                half = math.ceil(total / 2)
                left_stu = students_df.iloc[:half].reset_index(drop=True)
                right_stu = students_df.iloc[half:].reset_index(drop=True)

                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                hfont = Font(bold=True, size=10)

                for i, h in enumerate(left_headers):
                    c = ws.cell(row=3, column=i + 1, value=h)
                    c.font = hfont; c.border = border; c.alignment = Alignment(horizontal="center", vertical="center")
                for i, h in enumerate(right_headers):
                    c = ws.cell(row=3, column=i + 6, value=h)
                    c.font = hfont; c.border = border; c.alignment = Alignment(horizontal="center", vertical="center")

                # 数据行
                dfont = Font(size=10)
                max_rows = max(len(left_stu), len(right_stu))
                for r_idx in range(max_rows):
                    row_num = 4 + r_idx
                    ws.row_dimensions[row_num].height = 22

                    if r_idx < len(left_stu):
                        stu = left_stu.iloc[r_idx]
                        sid = int(stu["id"]); has_book = sid in dist_ids
                        for c_idx, val in enumerate([r_idx + 1, stu["name"], stu["student_id"],
                                          "✓" if has_book else "", ""], 1):
                            cell = ws.cell(row=row_num, column=c_idx, value=val)
                            cell.font = dfont; cell.border = border
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    if r_idx < len(right_stu):
                        stu = right_stu.iloc[r_idx]
                        sid = int(stu["id"]); has_book = sid in dist_ids
                        for c_idx, val in enumerate([half + r_idx + 1, stu["name"], stu["student_id"],
                                          "✓" if has_book else "", ""], 6):
                            cell = ws.cell(row=row_num, column=c_idx, value=val)
                            cell.font = dfont; cell.border = border
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                # 列宽
                col_widths = [6, 10, 14, 8, 10, 6, 10, 14, 8, 10]
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = w

                wb.save(output)
                return output.getvalue()

            if st.button("📥 导出正式领书单", use_container_width=True, type="primary", key="c_formal_export"):
                st.session_state.c_export_data = export_formal_booksheet()
                st.session_state.c_show_download = True
                st.rerun()

            if st.session_state.get("c_show_download"):
                st.download_button(
                    "✅ 点击下载",
                    data=st.session_state.c_export_data,
                    file_name=f"领书单_{c_class}_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary", key="c_dl_formal"
                )

            st.divider()
            st.caption("💡 **退伍复学等新到学生**：先到「学生管理」新增，再到此处补领")
    
    # ── Tab 2: 发放情况汇总（含领书人明细）──
    with tab2:
        st.markdown("#### 📋 发放明细（含领书人信息）")

        col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1.5, 1.5, 1.5])
        with col_f1:
            r_semester = st.selectbox("学期", [(0, "全部")] + [(r["id"], r["name"]) for _, r in semesters.iterrows()],
                               format_func=lambda x: x[1], key="cr_sem", index=1)
        with col_f2:
            r_college = st.selectbox("学院", ["全部"] + get_filtered_colleges(), key="cr_college")
        with col_f3:
            if r_college != "全部":
                r_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (r_college,))
            else:
                r_major_opts = ["全部"] + get_filtered_majors()
            r_major = st.selectbox("专业", r_major_opts, key="cr_major")
        with col_f4:
            r_class_where = "1=1"; r_class_params = []
            if r_college != "全部":
                r_class_where += " AND college = %s"; r_class_params.append(r_college)
            if r_major != "全部":
                r_class_where += " AND major = %s"; r_class_params.append(r_major)
            r_class_opts = ["全部"] + (get_filtered_list("students", "class_name", r_class_where, tuple(r_class_params)) if r_class_params else class_names)
            r_class = st.selectbox("班级", r_class_opts, key="cr_class")

        # 获取发放明细（结算价格实时从 textbooks_master 计算）
        sql = f"""
            SELECT sem.name as semester_name, s.class_name, s.grade, s.college, s.major,
                   s.student_id, s.name as student_name,
                   t.name as textbook_name,
                   {PRICE_CALC} as calc_price,
                   d.quantity,
                   {PRICE_CALC} * d.quantity as subtotal,
                   d.distribute_date as 领书时间, d.handler as 经手人
            FROM distributions d
            JOIN students s ON d.student_id = s.id
            JOIN textbooks t ON d.textbook_id = t.id
            JOIN semesters sem ON t.semester_id = sem.id
            {PRICE_JOIN}
            WHERE 1=1
        """
        params = []
        if r_semester[0] > 0:
            sql += " AND t.semester_id = %s"; params.append(r_semester[0])
        if r_college != "全部":
            sql += " AND s.college = %s"; params.append(r_college)
        if r_major != "全部":
            sql += " AND s.major = %s"; params.append(r_major)
        if r_class != "全部":
            sql += " AND s.class_name = %s"; params.append(r_class)
        sql += " ORDER BY sem.id DESC, s.class_name, s.student_id, t.name"

        df = query_df(sql, tuple(params) if params else None)

        if not df.empty:
            df["subtotal"] = pd.to_numeric(df["subtotal"], errors="coerce").fillna(0)
            total = df["subtotal"].sum()

            col_a, col_b = st.columns(2)
            with col_a:
                st.caption(f"共 **{len(df)}** 条发放记录")
            with col_b:
                st.metric("💰 发放总金额", f"¥{total:,.2f}")

            # 展示明细（含领书人信息，结算价=实洋>单价×折扣率）
            display_df = df.rename(columns={
                "semester_name": "学期", "class_name": "班级", "grade": "年级",
                "college": "学院", "major": "专业",
                "student_id": "学号", "student_name": "领书人", "textbook_name": "教材名称",
                "calc_price": "结算价", "quantity": "数量", "subtotal": "小计",
                "领书时间": "领书时间", "经手人": "经手人"
            })
            st.dataframe(
                display_df[["学期", "年级", "学院", "专业", "班级", "学号", "领书人", "教材名称", "结算价", "数量", "小计", "领书时间", "经手人"]].style.set_properties(**{"text-align": "center"}),
                use_container_width=True, hide_index=True,
                column_config={
                    "结算价": st.column_config.NumberColumn("结算价", format="¥%.2f"),
                    "小计": st.column_config.NumberColumn("小计", format="¥%.2f"),
                }
            )

            # 导出 - 共性字段放表头，表格只保留明细
            export_df = display_df[["学期", "年级", "学院", "专业", "班级", "学号", "领书人", "教材名称", "结算价", "数量", "小计", "领书时间", "经手人"]].copy()
            # 添加空白联系电话列（打印后手写）
            export_df.insert(export_df.columns.get_loc("领书人") + 1, "联系电话", "")

            # 计算班级级别的共性数据
            class_stats = {}
            for class_name in export_df["班级"].unique():
                # 班级人数
                stu_cnt = query_df("SELECT COUNT(*) as c FROM students WHERE class_name = %s", (class_name,))
                class_size = int(stu_cnt.iloc[0]["c"]) if not stu_cnt.empty else 0
                # 征订总数（该班该学期所有教材的征订量之和）
                if r_semester[0] > 0:
                    sub_q = query_df(
                        "SELECT COALESCE(SUM(quantity),0) as total FROM textbooks WHERE semester_id = %s AND class_name = %s",
                        (r_semester[0], class_name))
                else:
                    sub_q = query_df(
                        "SELECT COALESCE(SUM(quantity),0) as total FROM textbooks WHERE class_name = %s",
                        (class_name,))
                total_sub = int(sub_q.iloc[0]["total"]) if not sub_q.empty else 0
                class_stats[class_name] = {"班级人数": class_size, "征订总数": total_sub}
            export_df["班级人数"] = export_df["班级"].map(lambda c: class_stats[c]["班级人数"])
            export_df["征订总数"] = export_df["班级"].map(lambda c: class_stats[c]["征订总数"])

            def export_detail(df, class_col):
                from openpyxl.styles import Font, Alignment
                import io
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    classes = df[class_col].unique()
                    for cls in classes:
                        sheet_df = df[df[class_col] == cls].copy()
                        first_row = sheet_df.iloc[0]
                        sem_val = first_row.get("学期", ""); grade_val = first_row.get("年级", "")
                        college_val = first_row.get("学院", ""); major_val = first_row.get("专业", "")
                        sub_qty = first_row.get("征订总数", 0); class_size = first_row.get("班级人数", 0)

                        body_df = sheet_df.drop(columns=[class_col, "学期", "年级", "学院", "专业", "班级人数", "征订总数"])
                        body_df.to_excel(writer, sheet_name=str(cls)[:31], index=False, startrow=3)

                        ws = writer.sheets[str(cls)[:31]]; nc = len(body_df.columns)
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
                        tc = ws.cell(row=1, column=1, value=f"{sem_val} {cls} 教材发放明细")
                        tc.font = Font(bold=True, size=14); tc.alignment = Alignment(horizontal="center", vertical="center")
                        inf = Font(bold=False, size=10)
                        for i, (l, v) in enumerate([("学期", sem_val), ("年级", grade_val), ("班级", cls), ("班级人数", class_size), ("征订总数", sub_qty)]):
                            ws.cell(row=2, column=i*2+1, value=f"{l}：{v}").font = inf
                        for j, (l, v) in enumerate([("学院", college_val), ("专业", major_val)]):
                            ws.cell(row=3, column=j*2+1, value=f"{l}：{v}").font = inf
                        data_rows = len(body_df) + 1
                        apply_excel_borders(ws, 4, 4 + data_rows - 1, 1, nc)

                return output.getvalue()

            excel_data = export_detail(export_df, class_col="班级")
            st.download_button(
                "📥 导出发放明细（含领书人+联系电话栏，按班级分页）",
                data=excel_data,
                file_name=f"发放明细_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary"
            )
        else:
            st.info("📭 暂无发放数据")

# ═════════════════════════════════════════════════════════
# 6. 费用统计（管理员）—— 支持多选 + 按学期分列
# ═════════════════════════════════════════════════════════

def statistics_page():
    show_header("📊 费用统计", "多维度汇总学生教材费用，支持按学期分列展示，可导出 Excel")

    semesters = query_df("SELECT id, name FROM semesters ORDER BY id DESC")
    if semesters.empty:
        st.warning("⚠️ 请先添加学期")
        return

    class_names = get_filtered_class_names()

    # 筛选条件
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        sel_semester_names = st.multiselect(
            "选择学期（可多选）",
            options=list(semesters["name"]),
            default=list(semesters["name"])[:1] if not semesters.empty else [],
            format_func=lambda x: x
        )
    with col2:
        stat_college = st.selectbox("学院", ["全部"] + get_filtered_colleges(), key="stat_college")
    with col3:
        # 专业：基于所选学院级联过滤
        if stat_college != "全部":
            stat_major_opts = ["全部"] + get_filtered_list("students", "major", "college = %s", (stat_college,))
        else:
            stat_major_opts = ["全部"] + get_filtered_majors()
        stat_major = st.selectbox("专业", stat_major_opts, key="stat_major")
    with col4:
        # 班级：基于所选学院+专业级联过滤
        stat_class_where = "1=1"; stat_class_params = []
        if stat_college != "全部":
            stat_class_where += " AND college = %s"; stat_class_params.append(stat_college)
        if stat_major != "全部":
            stat_class_where += " AND major = %s"; stat_class_params.append(stat_major)
        stat_class_opts = ["全部"] + (get_filtered_list("students", "class_name", stat_class_where, tuple(stat_class_params)) if stat_class_params else class_names)
        sel_classes = st.multiselect(
            "选择班级（可多选）",
            options=stat_class_opts,
            default=[],
            format_func=lambda x: x
        )

    col_group = st.columns(1)[0]
    with col_group:
        group_by = st.selectbox("汇总维度", ["按学生", "按班级", "按专业", "按年级", "按学院"], key="stat_group")

    # 查询
    if not sel_semester_names:
        st.info("📭 请至少选择一个学期")
        return

    sem_id_map = {r["name"]: r["id"] for _, r in semesters.iterrows()}
    sel_sem_ids = [sem_id_map[n] for n in sel_semester_names]

    # 构建查询：获取明细数据
    ph = ",".join(["%s"] * len(sel_sem_ids))
    sql = f"""
            SELECT s.id as student_id_pk, s.student_id, s.name as student_name,
                   s.grade, s.college, s.major, s.class_name,
                   t.name as textbook_name,
                   {PRICE_CALC} as calc_price,
                   d.quantity,
                   {PRICE_CALC} * d.quantity as subtotal,
                   sem.name as semester_name, sem.id as semester_id,
                   d.distribute_date
            FROM distributions d
            JOIN students s ON d.student_id = s.id
            JOIN textbooks t ON d.textbook_id = t.id
            JOIN semesters sem ON t.semester_id = sem.id
            {PRICE_JOIN}
            LEFT JOIN student_exemptions e ON e.semester_id = sem.id AND e.student_id = s.id
            WHERE t.semester_id IN ({ph})
              AND (e.id IS NULL OR e.is_exempt = 0)
        """

    params = list(sel_sem_ids)

    # 学院筛选
    if stat_college != "全部":
        sql += " AND s.college = %s"; params.append(stat_college)
    # 专业筛选
    if stat_major != "全部":
        sql += " AND s.major = %s"; params.append(stat_major)
    # 班级筛选
    if sel_classes and "全部" not in sel_classes:
        placeholders_c = ",".join(["%s"] * len(sel_classes))
        sql += f" AND s.class_name IN ({placeholders_c})"
        params.extend(sel_classes)

    sql += " ORDER BY s.class_name, s.name, sem.id, t.name"

    raw_df = query_df(sql, tuple(params))

    if raw_df.empty:
        st.info("📭 暂无发放数据可统计")
        return

    raw_df["subtotal"] = pd.to_numeric(raw_df["subtotal"], errors="coerce").fillna(0)
    raw_df["calc_price"] = pd.to_numeric(raw_df["calc_price"], errors="coerce").fillna(0)

    # ── 按学生维度：生成透视表（学生 × 学期）──
    if group_by == "按学生":
        # 每个学生：各学期费用 + 总计
        pivot = raw_df.groupby(["student_id_pk", "student_id", "student_name", "class_name", "grade", "major", "college", "semester_name"])["subtotal"].sum().reset_index()
        
        # 透视：行=学生，列=学期
        pivot_table = pivot.pivot_table(
            index=["student_id_pk", "student_id", "student_name", "class_name", "grade", "major", "college"],
            columns="semester_name",
            values="subtotal",
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        # 总计列
        sem_cols = [c for c in pivot_table.columns if c not in ["student_id_pk", "student_id", "student_name", "class_name", "grade", "major", "college"]]
        pivot_table["总计"] = pivot_table[sem_cols].sum(axis=1)
        pivot_table["总计"] = pivot_table["总计"].round(2)

        # 格式化学期列名
        display_cols = ["学号", "姓名", "班级", "年级", "专业", "学院"] + sem_cols + ["总计"]
        result_df = pivot_table.rename(columns={
            "student_id": "学号",
            "student_name": "姓名",
            "class_name": "班级",
            "grade": "年级",
            "major": "专业",
            "college": "学院"
        })
        result_df = result_df[display_cols]

        # 指标
        total_fee = result_df["总计"].sum()
        record_count = len(result_df)
        avg_fee = total_fee / record_count if record_count > 0 else 0

        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("💰 总费用", f"¥{total_fee:,.2f}")
        with m2:
            st.metric("📊 学生数", f"{record_count}")
        with m3:
            st.metric("📈 人均费用", f"¥{avg_fee:,.2f}")

        st.dataframe(result_df.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

        # 导出：全部在一个 sheet
        excel_data = excel_export(result_df, "费用统计_按学生")
        st.download_button(
            "📥 导出费用统计",
            data=excel_data,
            file_name=f"费用统计_按学生_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        # 明细展开
        with st.expander("📋 查看费用明细列表"):
            detail_df = raw_df[["student_id","student_name","class_name","textbook_name",
                            "calc_price","quantity","subtotal","semester_name","distribute_date"]]
            detail_df = detail_df.rename(columns={
                "student_id":"学号","student_name":"姓名","class_name":"班级",
                "textbook_name":"教材名称","calc_price":"结算价","quantity":"数量",
                "subtotal":"小计","semester_name":"学期","distribute_date":"发放日期"
            })
            st.dataframe(detail_df.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

    else:
        # 其他维度：按班级/专业/年级/学院
        group_map = {
            "按班级": ["class_name"],
            "按专业": ["major"],
            "按年级": ["grade"],
            "按学院": ["college"],
        }
        gcols = group_map[group_by]
        
        # 先按维度+学期分组
        pivot = raw_df.groupby(gcols + ["semester_name"])["subtotal"].sum().reset_index()
        
        # 透视
        pivot_table = pivot.pivot_table(
            index=gcols,
            columns="semester_name",
            values="subtotal",
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        sem_cols = [c for c in pivot_table.columns if c not in gcols]
        pivot_table["总计"] = pivot_table[sem_cols].sum(axis=1)
        pivot_table["总计"] = pivot_table["总计"].round(2)

        # 重命名列
        rename_map = {
            "class_name": "班级",
            "major": "专业",
            "grade": "年级",
            "college": "学院"
        }
        result_df = pivot_table.rename(columns=rename_map)

        total_fee = result_df["总计"].sum()
        record_count = len(result_df)

        m1, m2 = st.columns(2)
        with m1:
            st.metric("💰 总费用", f"¥{total_fee:,.2f}")
        with m2:
            st.metric(f"📊 {group_by}条目", f"{record_count}")

        st.dataframe(result_df.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True)

        excel_data = excel_export(result_df, "费用统计")
        st.download_button(
            "📥 导出费用统计",
            data=excel_data,
            file_name=f"费用统计_{group_by}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

# ═════════════════════════════════════════════════════════
# 7. 学生费用查询
# ═════════════════════════════════════════════════════════

def student_query_page():
    student = st.session_state.user
    show_header("🎓 我的教材费用", f"查看个人各学期教材费用明细与汇总")

    st.markdown(f"""
    <div class="student-card">
        <div style="display:flex; gap:24px; align-items:center; flex-wrap:wrap;">
            <div style="text-align:center;">
                <div style="font-size:48px;">👤</div>
            </div>
            <div>
                <h2 style="margin:0 0 4px 0; color:#1e40af;">{student['name']}</h2>
                <p style="margin:0; color:#6b7280;">学号：{student['student_id']} &nbsp;|&nbsp; 班级：{student.get('class_name', '-')} &nbsp;|&nbsp; 专业：{student.get('major', '-')}</p>
                <p style="margin:4px 0 0 0; color:#9ca3af; font-size:13px;">学院：{student.get('college', '-')} &nbsp;|&nbsp; 年级：{student.get('grade', '-')}</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    sql = f"""
        SELECT t.name as textbook_name, t.publisher, t.editor,
               {PRICE_CALC} as calc_price,
               d.quantity,
               {PRICE_CALC} * d.quantity as subtotal,
               sem.name as semester_name, d.distribute_date
        FROM distributions d
        JOIN textbooks t ON d.textbook_id = t.id
        JOIN semesters sem ON t.semester_id = sem.id
        {PRICE_JOIN}
        LEFT JOIN student_exemptions e ON e.semester_id = sem.id AND e.student_id = d.student_id
        WHERE d.student_id = %s
          AND (e.id IS NULL OR e.is_exempt = 0)
        ORDER BY sem.id, t.name
    """
    df = query_df(sql, (student["id"],))

    if df.empty:
        st.info("📭 暂无教材发放记录")
        return

    df["subtotal"] = pd.to_numeric(df["subtotal"], errors="coerce").fillna(0)
    df["calc_price"] = pd.to_numeric(df["calc_price"], errors="coerce").fillna(0)

    total_fee = df["subtotal"].sum()
    book_count = len(df["textbook_name"].unique())

    # 汇总指标
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("📚 总费用", f"¥{total_fee:,.2f}")
    with m2:
        st.metric("📖 教材种类", f"{book_count} 种")
    with m3:
        st.metric("📝 记录条数", f"{len(df)} 条")

    st.divider()

    # 按学期汇总
    st.markdown("### 📊 按学期汇总")
    sem_summary = df.groupby("semester_name")["subtotal"].sum().reset_index()
    sem_summary.columns = ["学期", "费用合计"]
    sem_summary["费用合计"] = sem_summary["费用合计"].round(2)

    cols_sem = st.columns(len(sem_summary) if len(sem_summary) > 0 else 1)
    for i, (_, row) in enumerate(sem_summary.iterrows()):
        with cols_sem[i % len(cols_sem)]:
            st.metric(row["学期"], f"¥{row['费用合计']:,.2f}")

    st.divider()

    # 明细
    st.markdown("### 📋 费用明细")
    detail_df = df.rename(columns={
        "textbook_name":"教材名称","publisher":"出版社","editor":"主编",
        "calc_price":"结算价(元)","quantity":"数量","subtotal":"小计(元)",
        "semester_name":"学期","distribute_date":"发放日期"
    })
    st.dataframe(detail_df.style.set_properties(**{"text-align": "center"}), use_container_width=True, hide_index=True,
                 column_config={
                     "结算价(元)": st.column_config.NumberColumn("结算价(元)", format="¥%.2f"),
                     "小计(元)": st.column_config.NumberColumn("小计(元)", format="¥%.2f"),
                 })

    # 导出
    excel_data = excel_export(detail_df, "费用明细")
    st.download_button("📥 导出费用明细", data=excel_data,
                       file_name=f"教材费用_{student['student_id']}_{student['name']}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═════════════════════════════════════════════════════════
# 8. 系统日志页面
# ═════════════════════════════════════════════════════════

def system_logs_page():
    show_header("📋 系统日志", "查看导入操作记录和错误详情")
    
    col1, col2 = st.columns([0.7, 0.3])
    with col2:
        if st.button("🔄 刷新日志", use_container_width=True):
            st.rerun()
    
    logs = read_import_logs(50)
    if not logs or logs == ["暂无日志记录"]:
        st.info("暂无日志记录")
        return
    
    for log_text in logs:
        st.code(log_text.strip(), language="text")
    
    st.caption("日志文件位置：logs/import.log")

# ═════════════════════════════════════════════════════════
# 主入口
# ═════════════════════════════════════════════════════════

def main():
    # 初始化 session
    for k, v in [("role", None), ("page", None), ("show_semester_form", False), ("show_gen_confirm", False), ("pending_action", "")]:
        if k not in st.session_state:
            st.session_state[k] = v

    # 初始化数据库
    try:
        init_db()
    except Exception as e:
        st.error(f"❌ 数据库连接失败，请检查 config.ini 配置：\n{e}")
        st.stop()

    # 路由
    if st.session_state.role is None:
        login_page()
    elif st.session_state.role == "admin":
        admin_sidebar()
        pages = {
            "semester": semester_management,
            "students": student_management,
            "textbook_master": textbook_master_management,
            "subscriptions": subscription_management,
            "textbooks": textbook_management,
            "distribution": distribution_management,
            "confirmation": confirmation_page,
            "statistics": statistics_page,
            "logs": system_logs_page,
        }
        page_func = pages.get(st.session_state.page, semester_management)
        page_func()
    elif st.session_state.role == "student":
        student_sidebar()
        student_query_page()

if __name__ == "__main__":
    main()
